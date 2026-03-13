# -*- coding: utf-8 -*-
"""
Dış Alan Çalışma Excel Şablonu Otomatik Üretici

Kullanım:
    python scripts/create_dis_alan_calisma_template.py

Bu script, aktif katsayı protokollerini DB'den çekerek
Dis_Alan_Calisma_Sablonu_v4.xlsx dosyasını üretir.
"""
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import sqlite3
import sys

# DB yolu (gerekirse değiştirin)
DB_PATH = "data/local.db"
TEMPLATE_PATH = "data/templates/Dis_Alan_Calisma_Sablonu_v4.xlsx"

# Aktif katsayı protokollerini DB'den çek

def get_aktif_birimler():
    try:
        conn = sqlite3.connect(DB_PATH)
        cur = conn.cursor()
        cur.execute("""
            SELECT AnaBilimDali, Birim
            FROM Dis_Alan_Katsayi_Protokol
            WHERE Aktif=1 AND (GecerlilikBitis IS NULL OR date(GecerlilikBitis) >= date('now'))
            GROUP BY AnaBilimDali, Birim
            ORDER BY AnaBilimDali, Birim
        """)
        birimler = [f"{row[0]} - {row[1]}" for row in cur.fetchall()]
        conn.close()
        return birimler
    except Exception as e:
        print(f"DB bağlantı/okuma hatası: {e}")
        return ["Birim1", "Birim2"]

def create_template():
    wb = Workbook()
    ws = wb.active
    if ws is None:
        # This should not happen with a new workbook, but it's good practice to check.
        print("Hata: Aktif çalışma sayfası oluşturulamadı.", file=sys.stderr)
        return

    ws.title = "Çalışma"
    # Başlıklar
    headers = [
        "TC Kimlik No",
        "Ad Soyad",
        "Çalışılan Alan",  # Dropdown olacak
        "Ort. Süre (dk)",
        "Vaka Sayısı",
        "Katsayı",
        "Hesaplanan Saat",
        # "Tutanak No" KALDIRILDI
        "Tarih"
    ]
    for col_idx, header_text in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header_text)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1A5490", end_color="1A5490", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Çalışılan Alan dropdown
    birimler = get_aktif_birimler()
    from openpyxl.worksheet.datavalidation import DataValidation
    dv = DataValidation(type="list", formula1=f'"{','.join(birimler)}"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add("C2:C100")

    # Kılavuz sayfası
    ws2 = wb.create_sheet("Kılavuz")
    ws2["A1"] = "Çalışılan Alan dropdown'ı, sistemdeki aktif katsayı protokollerinden otomatik üretilir."
    ws2["A2"] = "Katsayılar ve formüller için kurumsal protokolü inceleyin."
    ws2["A3"] = "Tutanak No alanı kaldırıldı, sistem otomatik üretir."
    ws2["A4"] = "Ort. Süre (dk) ve Katsayı alanlarını protokole göre doldurun."
    for i in range(1, 5):
        ws2[f"A{i}"].font = Font(bold=True, color="1A5490")

    wb.save(TEMPLATE_PATH)
    print(f"Şablon oluşturuldu: {TEMPLATE_PATH}")

if __name__ == "__main__":
    create_template()
