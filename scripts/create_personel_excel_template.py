# -*- coding: utf-8 -*-
"""
Personel Bilgi Excel Şablonment Creator
════════════════════════════════════════════════════════════════════════════

Bu script, personel_bilgi.xlsx şablonunu otomatik olarak oluşturur.
RaporServisi kullanmak için bu şablonun data/templates/excel/ dizininde
olması gerekir.

Çalıştırma:
    python scripts/create_personel_excel_template.py
"""

from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side,
    DEFAULT_FONT
)
from core.paths import BASE_DIR



def create_personel_template_excel():
    """Personel bilgi Excel şablonunu oluşturur."""
    
    # Çıkış dizini
    templates_dir = Path(BASE_DIR) / "data" / "templates" / "excel"
    templates_dir.mkdir(parents=True, exist_ok=True)
    output_file = templates_dir / "personel_bilgi.xlsx"
    
    # Yeni workbook oluştur
    wb = Workbook()
    ws = wb.active
    ws.title = "Personel"
    
    # Sütun genişlikleri
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 30
    
    # Renkler ve stiller
    header_fill = PatternFill(start_color="1A5490", end_color="1A5490", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    
    section_fill = PatternFill(start_color="D9E8F5", end_color="D9E8F5", fill_type="solid")
    section_font = Font(bold=True, color="1A5490", size=11)
    
    label_font = Font(bold=True, color="1A5490", size=10)
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    row = 1
    
    # ── BAŞLIK ─────────────────────────────────────────────────────────
    ws.merge_cells(f"A{row}:D{row}")
    cell = ws[f"A{row}"]
    cell.value = "{{sirket_adi}}"
    cell.font = Font(bold=True, size=16, color="FFFFFF")
    cell.fill = header_fill
    cell.alignment = center_align
    ws.row_dimensions[row].height = 30
    row += 1
    
    ws.merge_cells(f"A{row}:D{row}")
    cell = ws[f"A{row}"]
    cell.value = "PERSONEL BİLGİ FORMU"
    cell.font = Font(bold=True, size=13, color="FFFFFF")
    cell.fill = header_fill
    cell.alignment = center_align
    ws.row_dimensions[row].height = 25
    row += 2
    
    # Rapor tarihi
    ws[f"A{row}"].value = "Rapor Tarihi:"
    ws[f"A{row}"].font = label_font
    ws[f"B{row}"].value = "{{rapor_tarihi}}"
    row += 2
    
    # ── KİMLİK BİLGİLERİ ──────────────────────────────────────────────
    ws.merge_cells(f"A{row}:D{row}")
    cell = ws[f"A{row}"]
    cell.value = "KİMLİK BİLGİLERİ"
    cell.font = section_font
    cell.fill = section_fill
    cell.alignment = left_align
    row += 1
    
    # İki sütun format
    def _add_field(r, col_a, label_a, col_b, label_b):
        ws[f"{col_a}{r}"].value = f"{label_a}:"
        ws[f"{col_a}{r}"].font = label_font
        ws[f"{col_b}{r}"].value = f"{{{{{label_a.lower().replace(' ', '_').replace('.', '')}}}}}"
        
        row_height = 18
        ws.row_dimensions[r].height = row_height
    
    _add_field(row, "A", "Adı Soyadı", "C", "Tc Kimlik No")
    row += 1
    _add_field(row, "A", "Doğum Tarihi", "C", "Cinsiyet")
    row += 1
    _add_field(row, "A", "Medeni Durum", "C", "Telefon")
    row += 2
    
    # ── İŞ BİLGİLERİ ──────────────────────────────────────────────────
    ws.merge_cells(f"A{row}:D{row}")
    cell = ws[f"A{row}"]
    cell.value = "İŞ BİLGİLERİ"
    cell.font = section_font
    cell.fill = section_fill
    ws.row_dimensions[row].height = 20
    row += 1
    
    _add_field(row, "A", "Personel Numarası", "C", "Departman")
    row += 1
    _add_field(row, "A", "Pozisyon", "C", "İşe Başlama Tarihi")
    row += 1
    _add_field(row, "A", "İş Yeri", "C", "Durum")
    row += 2
    
    # ── SAĞLIK KONTROL GEÇMİŞİ ────────────────────────────────────────
    ws.merge_cells(f"A{row}:D{row}")
    cell = ws[f"A{row}"]
    cell.value = "SAĞLIK KONTROL GEÇMİŞİ"
    cell.font = section_font
    cell.fill = section_fill
    ws.row_dimensions[row].height = 20
    row += 1
    
    # Tablo başlığı
    headers = ["#", "Muayene Sınıfı", "Muayene Tarihi", "Durum"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border
    
    ws.row_dimensions[row].height = 20
    row += 1
    
    # Tablo satır şablonu ({{ROW}} işareti ile)
    table_headers = ["{{ROW}}", "{{muayene_sinifi}}", "{{tarih}}", "{{durum}}"]
    for col_idx, val in enumerate(table_headers, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = val
        cell.alignment = left_align
        cell.border = border
        cell.font = DEFAULT_FONT
    
    ws.row_dimensions[row].height = 18
    row += 1
    
    # Boş satır ekle (örnek)
    for col_idx in range(1, 5):
        cell = ws.cell(row=row, column=col_idx)
        cell.border = border
    ws.row_dimensions[row].height = 18
    row += 2
    
    # ── NOTLAR ─────────────────────────────────────────────────────────
    ws.merge_cells(f"A{row}:D{row}")
    cell = ws[f"A{row}"]
    cell.value = "NOTLAR:"
    cell.font = label_font
    row += 1
    
    ws.merge_cells(f"A{row}:D{row + 2}")
    cell = ws[f"A{row}"]
    cell.value = "{{notlar}}"
    cell.alignment = left_align
    cell.border = border
    ws.row_dimensions[row].height = 60
    row += 3
    
    # ── İMZA ───────────────────────────────────────────────────────────
    row += 1
    ws.merge_cells(f"A{row}:B{row}")
    cell = ws[f"A{row}"]
    cell.value = "Hazırlayan"
    cell.font = Font(bold=True, size=10)
    cell.alignment = center_align
    
    ws.merge_cells(f"C{row}:D{row}")
    cell = ws[f"C{row}"]
    cell.value = "Onaylayan"
    cell.font = Font(bold=True, size=10)
    cell.alignment = center_align
    
    row += 1
    ws.merge_cells(f"A{row}:B{row}")
    cell = ws[f"A{row}"]
    cell.value = "{{hazirlayan}}"
    cell.alignment = center_align
    
    ws.merge_cells(f"C{row}:D{row}")
    cell = ws[f"C{row}"]
    cell.value = "{{onaylayan}}"
    cell.alignment = center_align
    
    row += 2
    ws.merge_cells(f"A{row}:B{row}")
    cell = ws[f"A{row}"]
    cell.value = "Tarih: {{cikis_tarihi}}"
    cell.alignment = center_align
    
    # Dosyayı kaydet
    wb.save(str(output_file))
    print(f"✓ Personel bilgi şablonu oluşturuldu: {output_file}")
    return output_file


if __name__ == "__main__":
    create_personel_template_excel()
