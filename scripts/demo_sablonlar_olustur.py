#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Demo şablonları oluşturur — data/templates/excel/ ve data/templates/pdf/
Çalıştır: python scripts/demo_sablonlar_olustur.py

Bu şablonlar hem kullanıma hazır rapor örnekleri hem de kendi
şablonlarını tasarlayacak olanlar için referans gösterir.

Özellikler gösterilen teknikler:
  Excel:
    • {{skalar_deger}} — tek hücre yer tutucusu
    • {{ROW}} + {{kolon}} — otomatik genişleyen tablo satırı
    • {{#}} — otomatik sıra numarası
    • Biçimlendirme (renk, font, kenarlık, birleştirme, sütun genişliği)
    • Çok sayfalı çalışma kitabı (opsiyonel)
  PDF (HTML):
    • {{ degisken }} Jinja2 değişkeni
    • {% for satir in tablo %} döngüsü
    • Koşullu stil ({{ satir.Durum | e }})
"""
import sys, os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side,
    GradientFill
)
from openpyxl.utils import get_column_letter

from core.rapor_servisi import EXCEL_TMPL_DIR, PDF_TMPL_DIR

# ── Renk paleti ──────────────────────────────────────────────────────────────
KOYU_MAVI  = "1a3a5c"
ORTA_MAVI  = "2563a8"
ACIK_MAVI  = "dbeafe"
GRI_ZEMIN  = "f1f5f9"
SARI       = "fef9c3"
KIRMIZI    = "fee2e2"
YESIL      = "dcfce7"
BEYAZ      = "FFFFFF"

def _yarim_ince(renk="b0bec5"):
    s = Side(style="thin", color=renk)
    return Border(left=s, right=s, top=s, bottom=s)

def _ince_alt(renk="dde1e7"):
    return Border(bottom=Side(style="thin", color=renk))

def _hdr(ws, row, col, val, bg=KOYU_MAVI, fg=BEYAZ, bold=True, hizala="center"):
    h = ws.cell(row=row, column=col, value=val)
    h.font = Font(bold=bold, color=fg, size=10)
    h.fill = PatternFill(fill_type="solid", fgColor=bg)
    h.alignment = Alignment(horizontal=hizala, vertical="center", wrap_text=True)
    h.border = _yarim_ince()
    return h

def _row_tmpl(ws, row, vals_dict):
    """
    vals_dict: {col_index: "{{placeholder}}" veya literal}
    """
    for col, val in vals_dict.items():
        h = ws.cell(row=row, column=col, value=val)
        h.fill = PatternFill(fill_type="solid", fgColor=GRI_ZEMIN)
        h.alignment = Alignment(horizontal="center", vertical="center")
        h.border = _ince_alt()


# ════════════════════════════════════════════════════════════════════════════
# 1. KALİBRASYON LİSTESİ ŞABLONU
# ════════════════════════════════════════════════════════════════════════════

def olustur_kalibrasyon_excel():
    wb = Workbook()
    ws = wb.active
    ws.title = "Kalibrasyon"

    # ── Üst başlık (birleştirilmiş) ─────────────────────────────────────────
    ws.merge_cells("A1:H1")
    h1 = ws["A1"]
    h1.value = "{{birim}} — Kalibrasyon Takip Raporu"
    h1.font = Font(bold=True, size=14, color=BEYAZ)
    h1.fill = PatternFill(fill_type="solid", fgColor=KOYU_MAVI)
    h1.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # ── Meta satırı ─────────────────────────────────────────────────────────
    ws.merge_cells("A2:C2")
    ws["A2"] = "Rapor Tarihi: {{tarih}}"
    ws["A2"].font = Font(size=10, color="444444")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    ws.merge_cells("D2:F2")
    ws["D2"] = "Toplam Kayıt: {{toplam}}"
    ws["D2"].font = Font(size=10, color="444444")
    ws["D2"].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("G2:H2")
    ws["G2"] = "Hazırlayan: {{hazirlayan}}"
    ws["G2"].font = Font(size=10, color="444444", italic=True)
    ws["G2"].alignment = Alignment(horizontal="right", vertical="center")
    ws.row_dimensions[2].height = 18

    # ── Boş ayırıcı ─────────────────────────────────────────────────────────
    ws.row_dimensions[3].height = 6

    # ── Tablo başlıkları (satır 4) ───────────────────────────────────────────
    basliklar = [
        (1, "#",             5),
        (2, "Cihaz Adı",    22),
        (3, "Cihaz No",     14),
        (4, "Bitiş Tarihi", 14),
        (5, "Geçerlilik",   14),
        (6, "Durum",        16),
        (7, "Yapıldı Tarihi",14),
        (8, "Açıklama",     30),
    ]
    for col, baslik, genislik in basliklar:
        _hdr(ws, 4, col, baslik)
        ws.column_dimensions[get_column_letter(col)].width = genislik

    ws.row_dimensions[4].height = 22

    # ── Tablo satır şablonu (satır 5) ────────────────────────────────────────
    # {{ROW}} → bu satırın şablon satırı olduğunu belirtir
    # {{#}}   → otomatik sıra numarası (1, 2, 3, ...)
    _row_tmpl(ws, 5, {
        1: "{{ROW}}",          # <-- zorunlu işaretçi
        2: "{{CihazAdi}}",
        3: "{{CihazNo}}",
        4: "{{BitisTarihi}}",
        5: "{{Gecerlilik}}",
        6: "{{Durum}}",
        7: "{{YapilanTarih}}",
        8: "{{Aciklama}}",
    })
    # Sıra no hücresine {{#}} ekle (ROW işaretçisinin yanında)
    ws.cell(5, 1).value = "{{ROW}}"   # RaporServisi bu hücreyi temizler

    ws.row_dimensions[5].height = 18

    # ── Dondur (başlık satırları sabit kalsın) ───────────────────────────────
    ws.freeze_panes = "A5"

    yol = EXCEL_TMPL_DIR / "kalibrasyon_listesi.xlsx"
    wb.save(str(yol))
    print(f"  ✅ {yol.name}")
    return yol


# ════════════════════════════════════════════════════════════════════════════
# 2. PERSONEL LİSTESİ ŞABLONU
# ════════════════════════════════════════════════════════════════════════════

def olustur_personel_excel():
    wb = Workbook()
    ws = wb.active
    ws.title = "Personel"

    ws.merge_cells("A1:F1")
    h1 = ws["A1"]
    h1.value = "{{birim}} — Personel Listesi"
    h1.font = Font(bold=True, size=13, color=BEYAZ)
    h1.fill = PatternFill(fill_type="solid", fgColor="0f3460")
    h1.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws["A2"] = "Tarih: {{tarih}}"
    ws["D2"] = "Aktif Personel: {{aktif_personel}}"
    ws["F2"] = "Toplam: {{toplam}}"
    for hcr in ["A2","D2","F2"]:
        ws[hcr].font = Font(size=10, color="555555")
        ws[hcr].alignment = Alignment(vertical="center")
    ws.row_dimensions[2].height = 16

    basliklar = [
        (1, "#",             5),
        (2, "Ad Soyad",     24),
        (3, "Unvan",        20),
        (4, "Birim",        18),
        (5, "Durum",        12),
        (6, "İletişim",     22),
    ]
    for col, baslik, genislik in basliklar:
        _hdr(ws, 3, col, baslik, bg="0f3460")
        ws.column_dimensions[get_column_letter(col)].width = genislik
    ws.row_dimensions[3].height = 20

    _row_tmpl(ws, 4, {
        1: "{{ROW}}",
        2: "{{AdSoyad}}",
        3: "{{Unvan}}",
        4: "{{BirimAdi}}",
        5: "{{Durum}}",
        6: "{{Telefon}}",
    })
    ws.row_dimensions[4].height = 17
    ws.freeze_panes = "A4"

    yol = EXCEL_TMPL_DIR / "personel_listesi.xlsx"
    wb.save(str(yol))
    print(f"  ✅ {yol.name}")
    return yol


# ════════════════════════════════════════════════════════════════════════════
# 3. RKE MUAYENE ŞABLONU
# ════════════════════════════════════════════════════════════════════════════

def olustur_rke_excel():
    wb = Workbook()
    ws = wb.active
    ws.title = "RKE Muayene"

    ws.merge_cells("A1:G1")
    ws["A1"].value = "RKE Muayene Raporu — {{donem}}"
    ws["A1"].font = Font(bold=True, size=13, color=BEYAZ)
    ws["A1"].fill = PatternFill(fill_type="solid", fgColor="1e3a5f")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws["A2"] = "Rapor Tarihi: {{tarih}}"
    ws["D2"] = "Toplam Muayene: {{toplam}}"
    ws["F2"] = "Hazırlayan: {{hazirlayan}}"
    for h in ["A2","D2","F2"]:
        ws[h].font = Font(size=10)
        ws[h].alignment = Alignment(vertical="center")
    ws.row_dimensions[2].height = 15

    basliklar = [
        (1, "#",                 5),
        (2, "Ekipman No",       14),
        (3, "Cins",             16),
        (4, "Pb Değeri (mm)",   16),
        (5, "Kontrol Tarihi",   14),
        (6, "Sonuç",            20),
        (7, "Kontrol Eden",     20),
    ]
    for col, baslik, genislik in basliklar:
        _hdr(ws, 3, col, baslik, bg="1e3a5f")
        ws.column_dimensions[get_column_letter(col)].width = genislik
    ws.row_dimensions[3].height = 20

    _row_tmpl(ws, 4, {
        1: "{{ROW}}",
        2: "{{EkipmanNo}}",
        3: "{{Cins}}",
        4: "{{Pb}}",
        5: "{{KontrolTarihi}}",
        6: "{{Sonuc}}",
        7: "{{KontrolEden}}",
    })
    ws.row_dimensions[4].height = 17
    ws.freeze_panes = "A4"

    yol = EXCEL_TMPL_DIR / "rke_muayene.xlsx"
    wb.save(str(yol))
    print(f"  ✅ {yol.name}")
    return yol


# ════════════════════════════════════════════════════════════════════════════
# 4. ARIZA LİSTESİ ŞABLONU
# ════════════════════════════════════════════════════════════════════════════

def olustur_ariza_excel():
    wb = Workbook()
    ws = wb.active
    ws.title = "Arıza"

    ws.merge_cells("A1:G1")
    ws["A1"].value = "Cihaz Arıza Raporu — {{tarih_aralik}}"
    ws["A1"].font = Font(bold=True, size=13, color=BEYAZ)
    ws["A1"].fill = PatternFill(fill_type="solid", fgColor="7c1c1c")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws["A2"] = "Rapor Tarihi: {{tarih}}"
    ws["C2"] = "Toplam: {{toplam}}"
    ws["E2"] = "Açık: {{acik}}"
    ws["F2"] = "Kapalı: {{kapali}}"
    for h in ["A2","C2","E2","F2"]:
        ws[h].font = Font(size=10)
    ws.row_dimensions[2].height = 15

    basliklar = [
        (1, "#",               5),
        (2, "Arıza No",       12),
        (3, "Cihaz",          22),
        (4, "Başlangıç",      14),
        (5, "Bitiş",          14),
        (6, "Durum",          14),
        (7, "Açıklama",       30),
    ]
    for col, baslik, genislik in basliklar:
        _hdr(ws, 3, col, baslik, bg="7c1c1c")
        ws.column_dimensions[get_column_letter(col)].width = genislik
    ws.row_dimensions[3].height = 20

    _row_tmpl(ws, 4, {
        1: "{{ROW}}",
        2: "{{ArizaNo}}",
        3: "{{CihazAdi}}",
        4: "{{Baslangic}}",
        5: "{{Bitis}}",
        6: "{{Durum}}",
        7: "{{Aciklama}}",
    })
    ws.row_dimensions[4].height = 17
    ws.freeze_panes = "A4"

    yol = EXCEL_TMPL_DIR / "ariza_listesi.xlsx"
    wb.save(str(yol))
    print(f"  ✅ {yol.name}")
    return yol


# ════════════════════════════════════════════════════════════════════════════
# 5. PDF ŞABLONLARI
# ════════════════════════════════════════════════════════════════════════════

BASE_CSS = """
body  { font-family: 'Segoe UI', Arial, sans-serif; font-size: 11px;
        color: #1a1a2e; margin: 0; padding: 0; }
.header { background: {bg}; color: #fff; padding: 12px 16px 8px; }
.header h1 { margin: 0 0 4px; font-size: 16px; }
.header .meta { font-size: 10px; opacity: 0.85; }
.content { padding: 12px 16px; }
table { width: 100%; border-collapse: collapse; font-size: 10.5px; }
thead th {{ background: {bg}; color: #fff; padding: 6px 8px;
            text-align: left; font-weight: 600; }}
tbody td {{ padding: 5px 8px; border-bottom: 1px solid #e2e8f0; }}
tbody tr:nth-child(even) td {{ background: #f8fafc; }}
.badge {{ display: inline-block; padding: 2px 7px; border-radius: 10px;
          font-size: 9.5px; font-weight: 600; }}
.badge-ok    {{ background: #dcfce7; color: #166534; }}
.badge-warn  {{ background: #fef9c3; color: #854d0e; }}
.badge-err   {{ background: #fee2e2; color: #991b1b; }}
.footer {{ margin-top: 16px; font-size: 9px; color: #94a3b8;
           text-align: right; border-top: 1px solid #e2e8f0; padding-top: 6px; }}
"""

def _pdf_tmpl(dosya: str, baslik: str, meta_satirlar: list[str],
              tablo_basliklar: list[str], tablo_satirlar: str, bg: str):
    meta = " &nbsp;|&nbsp; ".join(f"{{ {m} }}" if "{" not in m else m for m in meta_satirlar)
    ths  = "".join(f"<th>{t}</th>" for t in tablo_basliklar)
    css  = BASE_CSS.replace("{bg}", f"#{bg}")
    html = f"""<!DOCTYPE html>
<html lang="tr">
<head><meta charset="utf-8">
<style>
{css}
</style>
</head>
<body>
<div class="header">
  <h1>{baslik}</h1>
  <div class="meta">{meta}</div>
</div>
<div class="content">
<table>
  <thead><tr>{ths}</tr></thead>
  <tbody>
{tablo_satirlar}
  </tbody>
</table>
</div>
<div class="footer">ITF Desktop &mdash; {{{{ tarih }}}}</div>
</body>
</html>"""
    yol = PDF_TMPL_DIR / dosya
    yol.write_text(html, encoding="utf-8")
    print(f"  ✅ {yol.name}")
    return yol


def olustur_kalibrasyon_pdf():
    satirlar = """\
    {% for satir in tablo %}
    <tr>
      <td>{{ loop.index }}</td>
      <td>{{ satir.CihazAdi }}</td>
      <td>{{ satir.CihazNo }}</td>
      <td>{{ satir.BitisTarihi }}</td>
      <td>{{ satir.Gecerlilik }}</td>
      <td>
        {% if satir.Durum == 'Tamamlandı' %}
          <span class="badge badge-ok">{{ satir.Durum }}</span>
        {% elif satir.Durum == 'Planlandı' %}
          <span class="badge badge-warn">{{ satir.Durum }}</span>
        {% else %}
          <span class="badge badge-err">{{ satir.Durum }}</span>
        {% endif %}
      </td>
      <td>{{ satir.Aciklama | e }}</td>
    </tr>
    {% else %}
    <tr><td colspan="7" style="text-align:center;color:#94a3b8">Kayıt bulunamadı</td></tr>
    {% endfor %}"""
    return _pdf_tmpl(
        "kalibrasyon_listesi.html",
        "{{ birim }} — Kalibrasyon Takip Raporu",
        ["Tarih: {{ tarih }}", "Toplam: {{ toplam }}", "Hazırlayan: {{ hazirlayan }}"],
        ["#","Cihaz Adı","Cihaz No","Bitiş Tarihi","Geçerlilik","Durum","Açıklama"],
        satirlar, "1a3a5c",
    )


def olustur_rke_pdf():
    satirlar = """\
    {% for satir in tablo %}
    <tr>
      <td>{{ loop.index }}</td>
      <td>{{ satir.EkipmanNo }}</td>
      <td>{{ satir.Cins }}</td>
      <td>{{ satir.Pb }}</td>
      <td>{{ satir.KontrolTarihi }}</td>
      <td>
        {% if 'Uygun Değil' in satir.Sonuc %}
          <span class="badge badge-err">{{ satir.Sonuc }}</span>
        {% else %}
          <span class="badge badge-ok">{{ satir.Sonuc }}</span>
        {% endif %}
      </td>
      <td>{{ satir.KontrolEden }}</td>
    </tr>
    {% else %}
    <tr><td colspan="7" style="text-align:center;color:#94a3b8">Kayıt bulunamadı</td></tr>
    {% endfor %}"""
    return _pdf_tmpl(
        "rke_muayene.html",
        "RKE Muayene Raporu — {{ donem }}",
        ["Tarih: {{ tarih }}", "Toplam: {{ toplam }}", "Hazırlayan: {{ hazirlayan }}"],
        ["#","Ekipman No","Cins","Pb (mm)","Kontrol Tarihi","Sonuç","Kontrol Eden"],
        satirlar, "1e3a5f",
    )


def olustur_ariza_pdf():
    satirlar = """\
    {% for satir in tablo %}
    <tr>
      <td>{{ loop.index }}</td>
      <td>{{ satir.ArizaNo }}</td>
      <td>{{ satir.CihazAdi }}</td>
      <td>{{ satir.Baslangic }}</td>
      <td>{{ satir.Bitis }}</td>
      <td>
        {% if satir.Durum == 'Açık' %}
          <span class="badge badge-err">{{ satir.Durum }}</span>
        {% elif satir.Durum == 'Kapatıldı' %}
          <span class="badge badge-ok">{{ satir.Durum }}</span>
        {% else %}
          <span class="badge badge-warn">{{ satir.Durum }}</span>
        {% endif %}
      </td>
      <td>{{ satir.Aciklama | e }}</td>
    </tr>
    {% else %}
    <tr><td colspan="7" style="text-align:center;color:#94a3b8">Kayıt bulunamadı</td></tr>
    {% endfor %}"""
    return _pdf_tmpl(
        "ariza_listesi.html",
        "Cihaz Arıza Raporu — {{ tarih_aralik }}",
        ["Tarih: {{ tarih }}", "Toplam: {{ toplam }}", "Açık: {{ acik }}", "Kapalı: {{ kapali }}"],
        ["#","Arıza No","Cihaz","Başlangıç","Bitiş","Durum","Açıklama"],
        satirlar, "7c1c1c",
    )


# ════════════════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    print("\n📊 Excel şablonları:")
    olustur_kalibrasyon_excel()
    olustur_personel_excel()
    olustur_rke_excel()
    olustur_ariza_excel()

    print("\n📄 PDF şablonları:")
    olustur_kalibrasyon_pdf()
    olustur_rke_pdf()
    olustur_ariza_pdf()

    from core.rapor_servisi import RaporServisi
    print(f"\nMevcut Excel şablonları: {RaporServisi.sablon_listesi('excel')}")
    print(f"Mevcut PDF  şablonları: {RaporServisi.sablon_listesi('pdf')}")
    print("\n✅ Tüm şablonlar oluşturuldu.")
