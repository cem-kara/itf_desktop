# -*- coding: utf-8 -*-
"""
core/rapor_servisi.py unit testleri
=====================================
Kapsam:
  1. Excel şablon doldurma — skalar yer tutucular
  2. Excel tablo satırı genişletme ({{ROW}} mekanizması)
  3. Excel stil kopyalama (renk, font, kenarlık)
  4. Excel boş tablo ve tek satır
  5. PDF Jinja2 render doğruluğu
  6. PDF Qt-stub ortamında HTML fallback
  7. Hata senaryoları (olmayan şablon → None)
  8. Yardımcı metotlar (sablon_listesi, sablon_yolu)

Qt bağımlılığı yoktur — PySide6 stub ile test edilir.
"""
import sys
import os
# Proje kökünü Python path'e ekler; pytest hangi dizinden çalıştırılırsa çalıştırılsın modüller bulunur.
sys.path.insert(0, os.path.dirname(__file__))

# PySide6 stub — Qt kurulmayan CI ortamları için
import types as _types
for _mod in ["PySide6","PySide6.QtCore","PySide6.QtWidgets","PySide6.QtGui"]:
    if _mod not in sys.modules:
        _stub = _types.ModuleType(_mod)
        for _sym in ["QTextDocument","QPdfWriter","QPageSize","QPageLayout",
                     "QMarginsF","QWidget","QFileDialog","QFont","QColor",
                     "QThread","Signal","QPushButton","QLabel","QHBoxLayout",
                     "QVBoxLayout","Qt","QCursor","QSizePolicy"]:
            setattr(_stub, _sym, type(_sym, (), {"__init__": lambda s,*a,**k: None}))
        sys.modules[_mod] = _stub

import shutil
import tempfile
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# ── Sabitler ────────────────────────────────────────────────────────────────

CONTEXT = {
    "baslik"  : "Test Raporu",
    "tarih"   : "16.02.2026",
    "toplam"  : 3,
    "birim"   : "Radyoloji",
}

TABLO = [
    {"Cihaz": "CT-1",  "Tarih": "2026-01-15", "Durum": "Tamamlandı", "Aciklama": "Periyodik"},
    {"Cihaz": "MR-2",  "Tarih": "2026-02-01", "Durum": "Tamamlandı", "Aciklama": "Yıllık"},
    {"Cihaz": "XR-3",  "Tarih": "2026-02-10", "Durum": "Planlandı",  "Aciklama": "Bekliyor"},
]

HTML_TMPL = """\
<!DOCTYPE html>
<html lang="tr">
<head><meta charset="utf-8">
<style>
body{font-family:Arial,sans-serif;font-size:11px}
h1{color:#1a3a5c}
table{width:100%;border-collapse:collapse}
th{background:#1a3a5c;color:#fff;padding:6px 8px}
td{padding:5px 8px;border-bottom:1px solid #dde}
tr:nth-child(even) td{background:#f4f6fb}
</style>
</head>
<body>
<h1>{{ baslik }}</h1>
<p>Tarih: {{ tarih }} | Toplam: {{ toplam }} | Birim: {{ birim }}</p>
<table>
  <thead><tr><th>#</th><th>Cihaz</th><th>Tarih</th><th>Durum</th><th>Açıklama</th></tr></thead>
  <tbody>
  {% for satir in tablo %}
  <tr>
    <td>{{ loop.index }}</td>
    <td>{{ satir.Cihaz }}</td>
    <td>{{ satir.Tarih }}</td>
    <td>{{ satir.Durum }}</td>
    <td>{{ satir.Aciklama }}</td>
  </tr>
  {% endfor %}
  </tbody>
</table>
</body>
</html>"""


# ── Fixture: geçici şablon dizinleri ────────────────────────────────────────

@pytest.fixture(scope="module")
def sablonlar(tmp_path_factory):
    """
    Gerçek openpyxl şablonu + Jinja2 HTML şablonu oluşturur.
    Projenin data/templates klasörlerini kullanmak yerine izole tmp dizini.
    """
    root = tmp_path_factory.mktemp("templates")
    excel_dir = root / "excel"
    pdf_dir   = root / "pdf"
    excel_dir.mkdir(); pdf_dir.mkdir()

    # ── Excel şablonu ───────────────────────────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "{{baslik}}"
    ws["A1"].font = Font(bold=True, size=14)
    ws["B1"] = "Tarih: {{tarih}}"
    ws["C1"] = "Birim: {{birim}}"

    # Tablo başlığı (satır 2)
    for col, val in enumerate(["#","Cihaz","Tarih","Durum","Açıklama"], start=1):
        h = ws.cell(row=2, column=col, value=val)
        h.font = Font(bold=True, color="FFFFFF")
        h.fill = PatternFill(fill_type="solid", fgColor="1a3a5c")
        h.alignment = Alignment(horizontal="center")

    # Tablo satır şablonu (satır 3)
    for col, val in enumerate(["{{ROW}}","{{Cihaz}}","{{Tarih}}","{{Durum}}","{{Aciklama}}"], start=1):
        h = ws.cell(row=3, column=col, value=val)
        h.fill = PatternFill(fill_type="solid", fgColor="f0f4f8")
        h.border = Border(bottom=Side(style="thin", color="cccccc"))
        h.alignment = Alignment(horizontal="center")

    excel_yol = excel_dir / "test_sablon.xlsx"
    wb.save(str(excel_yol))

    # ── HTML şablonu ────────────────────────────────────────────────────
    pdf_yol = pdf_dir / "test_sablon.html"
    pdf_yol.write_text(HTML_TMPL, encoding="utf-8")

    return {"excel_dir": excel_dir, "pdf_dir": pdf_dir}


@pytest.fixture
def servis(sablonlar, monkeypatch):
    """
    EXCEL_TMPL_DIR ve PDF_TMPL_DIR'i geçici fixture dizinlerine yönlendirir.
    """
    import core.rapor_servisi as rs
    monkeypatch.setattr(rs, "EXCEL_TMPL_DIR", sablonlar["excel_dir"])
    monkeypatch.setattr(rs, "PDF_TMPL_DIR",   sablonlar["pdf_dir"])
    return rs.RaporServisi


# ════════════════════════════════════════════════════════════════
# 1. Excel — skalar yer tutucular
# ════════════════════════════════════════════════════════════════

class TestExcelSkalarler:

    def test_baslik_doldurulur(self, servis, tmp_path):
        yol = servis.excel("test_sablon", CONTEXT, [], str(tmp_path / "r.xlsx"))
        ws = load_workbook(yol).active
        assert ws["A1"].value == "Test Raporu"

    def test_tarih_doldurulur(self, servis, tmp_path):
        yol = servis.excel("test_sablon", CONTEXT, [], str(tmp_path / "r.xlsx"))
        ws = load_workbook(yol).active
        assert ws["B1"].value == "Tarih: 16.02.2026"

    def test_birim_doldurulur(self, servis, tmp_path):
        yol = servis.excel("test_sablon", CONTEXT, [], str(tmp_path / "r.xlsx"))
        ws = load_workbook(yol).active
        assert ws["C1"].value == "Birim: Radyoloji"

    def test_bilinmeyen_ph_bos_string(self, servis, tmp_path):
        """Context'te olmayan yer tutucu boş string olmalı."""
        ctx = {"baslik": "X"}   # tarih ve birim yok
        yol = servis.excel("test_sablon", ctx, [], str(tmp_path / "r.xlsx"))
        ws = load_workbook(yol).active
        assert ws["B1"].value == "Tarih: "


# ════════════════════════════════════════════════════════════════
# 2. Excel — tablo satırı genişletme
# ════════════════════════════════════════════════════════════════

class TestExcelTablo:

    def test_3_satir_yazilir(self, servis, tmp_path):
        yol = servis.excel("test_sablon", CONTEXT, TABLO, str(tmp_path / "r.xlsx"))
        ws = load_workbook(yol).active
        assert ws.cell(3,2).value == "CT-1"
        assert ws.cell(4,2).value == "MR-2"
        assert ws.cell(5,2).value == "XR-3"

    def test_durum_dogru_yazilir(self, servis, tmp_path):
        yol = servis.excel("test_sablon", CONTEXT, TABLO, str(tmp_path / "r.xlsx"))
        ws = load_workbook(yol).active
        assert ws.cell(5,4).value == "Planlandı"

    def test_row_hucresi_temizlenir(self, servis, tmp_path):
        """{{ROW}} hücresi None olmalı (sıra no için {{#}} kullanılır)."""
        yol = servis.excel("test_sablon", CONTEXT, TABLO, str(tmp_path / "r.xlsx"))
        ws = load_workbook(yol).active
        assert ws.cell(3,1).value is None

    def test_tek_satir(self, servis, tmp_path):
        yol = servis.excel("test_sablon", CONTEXT, [TABLO[0]], str(tmp_path / "r.xlsx"))
        ws = load_workbook(yol).active
        assert ws.cell(3,2).value == "CT-1"

    def test_bos_tablo_sablon_satiri_temizlenir(self, servis, tmp_path):
        yol = servis.excel("test_sablon", CONTEXT, [], str(tmp_path / "r.xlsx"))
        ws = load_workbook(yol).active
        # Tüm şablon satırı hücreleri boş olmalı
        assert all(h.value is None for h in ws[3])

    def test_bos_tablo_skalarler_korunur(self, servis, tmp_path):
        yol = servis.excel("test_sablon", CONTEXT, [], str(tmp_path / "r.xlsx"))
        ws = load_workbook(yol).active
        assert ws["A1"].value == "Test Raporu"


# ════════════════════════════════════════════════════════════════
# 3. Excel — stil kopyalama
# ════════════════════════════════════════════════════════════════

class TestExcelStil:

    def test_renk_kopyalanir(self, servis, tmp_path):
        yol = servis.excel("test_sablon", CONTEXT, TABLO, str(tmp_path / "r.xlsx"))
        ws = load_workbook(yol).active
        # Satır 3 ve 4'ün fill rengi aynı olmalı (şablondan kopyalandı)
        renk3 = ws.cell(3,2).fill.fgColor.rgb
        renk4 = ws.cell(4,2).fill.fgColor.rgb
        assert renk3 == renk4

    def test_hizalama_kopyalanir(self, servis, tmp_path):
        yol = servis.excel("test_sablon", CONTEXT, TABLO, str(tmp_path / "r.xlsx"))
        ws = load_workbook(yol).active
        hz3 = ws.cell(3,2).alignment.horizontal
        hz4 = ws.cell(4,2).alignment.horizontal
        assert hz3 == hz4


# ════════════════════════════════════════════════════════════════
# 4. Excel — dosya çıktısı
# ════════════════════════════════════════════════════════════════

class TestExcelCikti:

    def test_dosya_olusur(self, servis, tmp_path):
        yol = servis.excel("test_sablon", CONTEXT, TABLO, str(tmp_path / "r.xlsx"))
        assert Path(yol).exists()

    def test_boyut_pozitif(self, servis, tmp_path):
        yol = servis.excel("test_sablon", CONTEXT, TABLO, str(tmp_path / "r.xlsx"))
        assert Path(yol).stat().st_size > 0

    def test_gecici_dizin_none(self, servis):
        """kayit_yolu=None verilince geçici dosyaya kaydedilmeli."""
        yol = servis.excel("test_sablon", CONTEXT, TABLO, None)
        assert yol is not None
        assert Path(yol).exists()

    def test_olmayan_sablon_none(self, servis):
        assert servis.excel("yok_sablon", {}) is None


# ════════════════════════════════════════════════════════════════
# 5. PDF — Jinja2 render
# ════════════════════════════════════════════════════════════════

class TestPDFRender:

    def _html(self, servis, tmp_path, ctx=None, tablo=None):
        """PDF üretir, HTML fallback dosyasını okuyup döndürür."""
        yol = str(tmp_path / "r.pdf")
        servis.pdf("test_sablon", ctx or CONTEXT, tablo or TABLO, yol)
        fb = yol.replace(".pdf", "_preview.html")
        return Path(fb).read_text(encoding="utf-8")

    def test_baslik_render(self, servis, tmp_path):
        html = self._html(servis, tmp_path)
        assert "Test Raporu" in html

    def test_tarih_render(self, servis, tmp_path):
        html = self._html(servis, tmp_path)
        assert "16.02.2026" in html

    def test_tum_satirlar_render(self, servis, tmp_path):
        html = self._html(servis, tmp_path)
        for kayit in TABLO:
            assert kayit["Cihaz"] in html

    def test_bos_tablo_render(self, servis, tmp_path):
        html = self._html(servis, tmp_path, tablo=[])
        assert "Test Raporu" in html

    def test_olmayan_sablon_none(self, servis):
        assert servis.pdf("yok_pdf", {}) is None


# ════════════════════════════════════════════════════════════════
# 6. Yardımcı metotlar
# ════════════════════════════════════════════════════════════════

class TestYardimciMetotlar:

    def test_sablon_listesi_excel(self, servis):
        assert "test_sablon" in servis.sablon_listesi("excel")

    def test_sablon_listesi_pdf(self, servis):
        assert "test_sablon" in servis.sablon_listesi("pdf")

    def test_sablon_yolu_excel_uzanti(self, servis):
        assert str(servis.sablon_yolu("test", "excel")).endswith(".xlsx")

    def test_sablon_yolu_pdf_uzanti(self, servis):
        assert str(servis.sablon_yolu("test", "pdf")).endswith(".html")

    def test_sablon_yolu_dogru_dizin(self, servis, sablonlar):
        yol = servis.sablon_yolu("test_sablon", "excel")
        assert yol.parent == sablonlar["excel_dir"]
