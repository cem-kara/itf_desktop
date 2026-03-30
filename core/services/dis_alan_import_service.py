# core/services/dis_alan_import_service.py
# -*- coding: utf-8 -*-
"""
DisAlanImportService — Dış alan birimlerinden gelen Excel şablonlarını
okur, doğrular ve Dis_Alan_Calisma tablosuna yazar.

Tutanak Akışı:
    Import edilen Excel dosyası önce Dokumanlar tablosuna kaydedilir.
    Oluşan DokumanId → tüm satırlar için TutanakNo olur.
    Import tarihi → TutanakTarihi olur.
    Böylece her şablon arşivlenir ve izlenebilir olur.

NOT: Bu servis Personel tablosundan TAMAMEN BAĞIMSIZDIR.
     TC kimlik no salt tanımlayıcıdır — sistemde kayıtlı olması gerekmez.
"""
from __future__ import annotations

import re
import uuid
import shutil
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Optional

from core.logger import logger

from database.repository_registry import RepositoryRegistry
from core.hata_yonetici import SonucYonetici
from core.services.dis_alan_katsayi_service import DisAlanKatsayiService


KOLON_HARITA = {
    "TC Kimlik No":     "TCKimlik",
    "Ad Soyad":         "AdSoyad",
    "Ad Soyad *":       "AdSoyad",
    "Çalışılan Alan":   "IslemTipi",
    "Çalışılan Alan *": "IslemTipi",
    "Vaka Sayısı":      "VakaSayisi",
    "Vaka Sayısı *":    "VakaSayisi",
}

TC_REGEX = re.compile(r"^\d{11}$")

# Türkçe ay adları → numara
_AY_MAP = {
    "oca": 1, "ocak": 1,
    "şub": 2, "sub": 2, "şubat": 2, "subat": 2,
    "mar": 3, "mart": 3,
    "nis": 4, "nisan": 4,
    "may": 5, "mayıs": 5, "mayis": 5,
    "haz": 6, "haziran": 6,
    "tem": 7, "temmuz": 7,
    "ağu": 8, "agu": 8, "ağustos": 8, "agustos": 8,
    "eyl": 9, "eylül": 9, "eylul": 9,
    "eki": 10, "ekim": 10,
    "kas": 11, "kasım": 11, "kasim": 11,
    "ara": 12, "aralık": 12, "aralik": 12,
}


def _parse_ay(deger) -> int:
    """
    D3 hücresinden sadece ay numarasını çıkarır.
    Desteklenen formatlar:
      - Sayı:  1, 3, 12
      - Ay adı (TR): "Ocak", "Mart", "oca", "mar"
      - Kısa+nokta: "Oca.", "Mar."
    Returns: 1–12 veya 0 (parse edilemedi)
    """
    if deger is None:
        return 0
    import datetime as _dt
    if isinstance(deger, (_dt.datetime, _dt.date)):
        return deger.month
    try:
        v = int(float(str(deger).strip()))
        if 1 <= v <= 12:
            return v
    except (ValueError, TypeError):
        pass
    s = str(deger).strip().rstrip(".").lower()
    return _AY_MAP.get(s, 0)
    """
    F3 hücresinden ay ve yılı çıkarır.

    Desteklenen formatlar:
      - "3/2026"  veya  "03/2026"
      - "3-2026"
      - "Mart 2026"  /  "Mar.26"  /  "Oca.26"
      - datetime / date nesnesi
      - Excel serial float

    Returns:
        (ay, yil)  — başarısız olursa (0, 0)
    """
    import datetime as _dt

    if deger is None:
        return 0, 0

    # datetime / date
    if isinstance(deger, (_dt.datetime, _dt.date)):
        return deger.month, deger.year

    # float → Excel seri tarihi
    if isinstance(deger, float):
        try:
            base = _dt.date(1899, 12, 30)
            d = base + _dt.timedelta(days=int(deger))
            return d.month, d.year
        except Exception:
            pass

    s = str(deger).strip()

    # "3/2026"  "03/2026"  "3-2026"
    m = re.match(r"^(\d{1,2})[/\-\.](\d{4})$", s)
    if m:
        return int(m.group(1)), int(m.group(2))

    # "2026/3"  "2026-03"
    m = re.match(r"^(\d{4})[/\-\.](\d{1,2})$", s)
    if m:
        return int(m.group(2)), int(m.group(1))

    # "Oca.26"  "Mar.26"  — Excel otomatik kısa format
    m = re.match(r"^([A-Za-zÇçĞğİıÖöŞşÜü]{3,})[\.\s\-](\d{2,4})$", s)
    if m:
        ay_str = m.group(1).lower()
        yil_str = m.group(2)
        ay = _AY_MAP.get(ay_str, 0)
        yil = int(yil_str)
        if yil < 100:
            yil += 2000
        if ay:
            return ay, yil

    # "Mart 2026"  "OCAK 2026"
    m = re.match(r"^([A-Za-zÇçĞğİıÖöŞşÜü]+)\s+(\d{4})$", s)
    if m:
        ay_str = m.group(1).lower()
        ay = _AY_MAP.get(ay_str, 0)
        if ay:
            return ay, int(m.group(2))

    return 0, 0


# ─────────────────────────────────────────────────────────────
#  Veri sınıfları
# ─────────────────────────────────────────────────────────────

@dataclass
class SatirSonucu:
    satir_no: int
    veri: dict
    hatalar: list[str] = field(default_factory=list)
    uyarilar: list[str] = field(default_factory=list)
    gecerli: bool = True
    kullanici_onayladi: bool = False

    @property
    def durum(self) -> str:
        if not self.gecerli: return "HATA"
        if self.uyarilar:    return "UYARI"
        return "TAMAM"



@dataclass
class ImportSonucu:
    dosya: str
    anabilim_dali: str = ""
    birim: str = ""
    donem_etiket: str = ""
    donem_ay: int = 0
    donem_yil: int = 0
    dokuman_id: str = ""        # Dokumanlar tablosuna kaydedilince dolar
    tutanak_no: str = ""        # dokuman_id ile aynı — tüm satırlar bunu kullanır
    import_tarihi: str = ""
    toplam_satir: int = 0
    gecerli: int = 0
    hatali: int = 0
    uyarili: int = 0
    kaydedilen: int = 0
    atlanan: int = 0
    satirlar: list[SatirSonucu] = field(default_factory=list)
    conflict_report_text: str = ""
    conflict_report_path: str = ""
    conflict_report_pdf_path: str = ""

    @property
    def basarili(self) -> bool:
        return self.kaydedilen > 0


# ─────────────────────────────────────────────────────────────
#  Yardımcılar
# ─────────────────────────────────────────────────────────────

def _parse_tc(deger) -> str:
    if deger is None:
        return ""
    try:
        s = str(int(float(str(deger).strip())))
    except (ValueError, TypeError):
        s = str(deger).strip()
    return s.zfill(11) if s.isdigit() and len(s) <= 11 else s


def _parse_int(deger) -> Optional[int]:
    try:
        return int(float(str(deger).strip()))
    except (ValueError, TypeError):
        return None


# ─────────────────────────────────────────────────────────────
#  Servis
# ─────────────────────────────────────────────────────────────


class DisAlanImportService:
    def __init__(self, registry: Optional[RepositoryRegistry] = None, arsiv_dizin: Optional[str] = None, katsayi_service: Optional[DisAlanKatsayiService] = None):
        self._r          = registry
        self._arsiv_dizin = Path(arsiv_dizin) if arsiv_dizin else None
        if katsayi_service:
            self._katsayi_service = katsayi_service
        elif registry:
            self._katsayi_service = DisAlanKatsayiService(registry)
        else:
            self._katsayi_service = None
        # Denetim servisi DI ile alınabilir, yoksa None

    # ── 1. Excel Okuma ───────────────────────────────────────

    def excel_oku(
        self,
        dosya_yolu: str,
        donem_ay: int = 0,
        donem_yil: int = 0,
        katsayi_cache: Optional[dict] = None,
    ) -> SonucYonetici:
        """
        Excel şablonunu okur ve doğrular. DB'ye yazmaz.

        Başlık satırı: 6. satır
        Veri satırları: 7–56
        Üst bilgi:
          A3="Anabilim Dalı" etiketi, B3=değer
          A4="Birim" etiketi,         B4=değer
          C3="Dönem Ay" etiketi,      D3=değer (ör: Ocak veya 1)
          C4="Dönem Yıl" etiketi,     D4=değer (ör: 2026)

        donem_ay / donem_yil: 0 verilirse D3/D4'ten otomatik okunur.
        """
        try:
            import openpyxl
        except ImportError:
            return SonucYonetici.hata(RuntimeError("openpyxl kurulu değil: pip install openpyxl"), "DisAlanImportService.excel_oku")

        import_tarihi = datetime.now().strftime("%Y-%m-%d")
        sonuc = ImportSonucu(
            dosya=Path(dosya_yolu).name,
            import_tarihi=import_tarihi,
        )

        try:
            wb = openpyxl.load_workbook(dosya_yolu, data_only=True)
        except Exception as e:
            return SonucYonetici.hata(ValueError(f"Dosya açılamadı: {e}"), "DisAlanImportService.excel_oku")

        if "Veri Girişi" not in wb.sheetnames:
            return SonucYonetici.hata(ValueError(
                "'Veri Girişi' sayfası bulunamadı — "
                "lütfen resmi şablonu kullanın."
            ), "DisAlanImportService.excel_oku")

        ws = wb["Veri Girişi"]

        # Üst bilgi — B3:AnaBilimDali, B4:Birim, F3:Dönem
        anabilim = str(ws["B3"].value or "").strip()
        birim    = str(ws["B4"].value or "").strip()

        # Dönem: D3=Ay, D4=Yıl
        d3_val = ws["D3"].value
        d4_val = ws["D4"].value

        # Yıl — D4
        if not donem_yil:
            try:
                donem_yil = int(float(str(d4_val).strip()))
            except (ValueError, TypeError):
                donem_yil = 0

        # Ay — D3: sadece ay adı veya numarası
        if not donem_ay:
            donem_ay = _parse_ay(d3_val)

        if not (donem_ay and donem_yil):
            return SonucYonetici.hata(ValueError(
                f"Dönem bilgisi okunamadı.\n"
                f"D3 hücresine ay adını (ör: Ocak) veya numarasını (ör: 1),\n"
                f"D4 hücresine yılı (ör: 2026) girin.\n"
                f"Okunan — D3: '{d3_val}'  D4: '{d4_val}'"
            ), "DisAlanImportService.excel_oku")

        donem_et = f"{donem_ay}/{donem_yil}"

        sonuc.anabilim_dali = anabilim
        sonuc.birim         = birim
        sonuc.donem_etiket  = donem_et
        sonuc.donem_ay      = donem_ay
        sonuc.donem_yil     = donem_yil

        # Kolon haritası
        kolon_idx = self._kolon_haritasi_bul(ws, baslik_satiri=6)
        if not kolon_idx:
            return SonucYonetici.hata(ValueError(
                "Başlık satırı tanınamadı. "
                "Lütfen resmi şablonu kullanın."
            ), "DisAlanImportService.excel_oku")

        # Veri satırları
        # Aynı kişi (TCKimlik) + dönem için tutarlılık kontrolü
        kisi_donem_ad_map: dict[tuple[str, int, int], dict] = {}
        conflict_items: list[dict] = []
        for row in range(7, 57):
            row_vals = {
                alan: ws.cell(row=row, column=col).value
                for alan, col in kolon_idx.items()
            }
            if all(v is None or str(v).strip() == "" for v in row_vals.values()):
                continue

            satir = self._satiri_isle(
                row, row_vals, donem_ay, donem_yil,
                anabilim_dali=sonuc.anabilim_dali,
                birim=sonuc.birim,
                katsayi_cache=katsayi_cache,
            )

            # TCKimlik + dönem eşleşmesi kontrolü (tutanak no boş gelebilir)
            tc = str(satir.veri.get("TCKimlik", "")).strip()
            ad = str(satir.veri.get("AdSoyad", "")).strip()
            if tc:
                key = (tc, donem_ay, donem_yil)
                prev = kisi_donem_ad_map.get(key)
                prev_ad = prev.get("ad") if prev else ""
                if prev and ad and prev_ad != ad:
                    conflict_items.append({
                        "tc": tc,
                        "donem_ay": donem_ay,
                        "donem_yil": donem_yil,
                        "eski": {
                            "satir_no": prev.get("satir_no"),
                            "veri": prev.get("veri", {}),
                        },
                        "yeni": {
                            "satir_no": row,
                            "veri": satir.veri,
                        },
                    })
                    satir.hatalar.append(
                        f"Aynı kişi+ dönem için farklı Ad Soyad: '{prev_ad}' / '{ad}'"
                    )
                    satir.gecerli = False
                elif ad:
                    kisi_donem_ad_map[key] = {"ad": ad, "satir_no": row, "veri": satir.veri}

            sonuc.satirlar.append(satir)
            sonuc.toplam_satir += 1
            if not satir.gecerli:    sonuc.hatali  += 1
            elif satir.uyarilar:     sonuc.uyarili += 1
            else:                    sonuc.gecerli += 1

        # Çakışma raporu hazırla
        if conflict_items:
            lines = []
            lines.append("DIS ALAN IMPORT HATA BILDIRIMI")
            lines.append(f"Dosya: {sonuc.dosya}")
            lines.append(f"Anabilim Dali/Birim: {sonuc.anabilim_dali} / {sonuc.birim}")
            lines.append(f"Donem: {donem_ay}/{donem_yil}")
            lines.append(f"Rapor Tarihi: {import_tarihi}")
            lines.append("")
            lines.append("Ayni kisi (TCKimlik) + donem icin farkli bilgiler tespit edildi.")
            lines.append("Lutfen dogru bilgiyi teyit edip guncel tutanagi gonderiniz.")
            lines.append("")

            def _fmt_row(veri: dict) -> str:
                return (
                    f"AdSoyad={veri.get('AdSoyad', '')} | "
                    f"Alan={veri.get('IslemTipi', '')} | "
                    f"Vaka={veri.get('VakaSayisi', '')} | "
                    f"Katsayi={veri.get('HesaplananSaat', '')}"
                )

            for i, c in enumerate(conflict_items, 1):
                lines.append(f"{i}. TC: {c['tc']} | Donem: {c['donem_ay']}/{c['donem_yil']}")
                lines.append(
                    f"Eski Kayit (Satir {c['eski']['satir_no']}): "
                    f"{_fmt_row(c['eski']['veri'])}"
                )
                lines.append(
                    f"Yeni Kayit (Satir {c['yeni']['satir_no']}): "
                    f"{_fmt_row(c['yeni']['veri'])}"
                )
                lines.append("")

            report_text = "\n".join(lines).strip() + "\n"
            sonuc.conflict_report_text = report_text

            report_dir = Path("logs")
            report_name = f"dis_alan_import_hata_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
            try:
                report_dir.mkdir(parents=True, exist_ok=True)
                report_path = report_dir / report_name
                report_path.write_text(report_text, encoding="utf-8")
                sonuc.conflict_report_path = str(report_path)
            except Exception as e:
                logger.error(f"Cakismanin raporu yazilamadi: {e}")

            # PDF raporu (kurumsal çıktı için)
            try:
                from reportlab.lib.pagesizes import A4
                from reportlab.lib.styles import getSampleStyleSheet
                from reportlab.lib.units import mm
                from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer
                from reportlab.pdfbase import pdfmetrics
                from reportlab.pdfbase.ttfonts import TTFont

                try:
                    pdfmetrics.registerFont(TTFont("Arial", "arial.ttf"))
                    pdfmetrics.registerFont(TTFont("ArialBold", "arialbd.ttf"))
                    font_name, font_bold = "Arial", "ArialBold"
                except Exception:
                    font_name, font_bold = "Helvetica", "Helvetica-Bold"

                pdf_name = report_name.replace(".txt", ".pdf")
                pdf_path = report_dir / pdf_name

                styles = getSampleStyleSheet()
                title_style = styles["Title"]
                title_style.fontName = font_bold
                title_style.fontSize = 14
                body_style = styles["Normal"]
                body_style.fontName = font_name
                body_style.fontSize = 9

                doc = SimpleDocTemplate(
                    str(pdf_path),
                    pagesize=A4,
                    leftMargin=18 * mm,
                    rightMargin=18 * mm,
                    topMargin=18 * mm,
                    bottomMargin=18 * mm,
                )

                elements = []
                elements.append(Paragraph("DIS ALAN IMPORT HATA BILDIRIMI", title_style))
                elements.append(Spacer(1, 4 * mm))
                elements.append(Paragraph(f"Dosya: {sonuc.dosya}", body_style))
                elements.append(Paragraph(f"Anabilim Dali/Birim: {sonuc.anabilim_dali} / {sonuc.birim}", body_style))
                elements.append(Paragraph(f"Donem: {donem_ay}/{donem_yil}", body_style))
                elements.append(Paragraph(f"Rapor Tarihi: {import_tarihi}", body_style))
                elements.append(Spacer(1, 4 * mm))
                elements.append(Paragraph(
                    "Ayni kisi (TCKimlik) + donem icin farkli bilgiler tespit edildi.",
                    body_style
                ))
                elements.append(Paragraph(
                    "Lutfen dogru bilgiyi teyit edip guncel tutanagi gonderiniz.",
                    body_style
                ))
                elements.append(Spacer(1, 6 * mm))

                for i, c in enumerate(conflict_items, 1):
                    elements.append(Paragraph(
                        f"{i}. TC: {c['tc']} | Donem: {c['donem_ay']}/{c['donem_yil']}",
                        body_style
                    ))
                    elements.append(Paragraph(
                        f"Eski Kayit (Satir {c['eski']['satir_no']}): {_fmt_row(c['eski']['veri'])}",
                        body_style
                    ))
                    elements.append(Paragraph(
                        f"Yeni Kayit (Satir {c['yeni']['satir_no']}): {_fmt_row(c['yeni']['veri'])}",
                        body_style
                    ))
                    elements.append(Spacer(1, 4 * mm))

                doc.build(elements)
                sonuc.conflict_report_pdf_path = str(pdf_path)
            except ImportError:
                logger.warning("reportlab kurulu degil; PDF raporu olusturulamadi")
            except Exception as e:
                logger.error(f"PDF raporu yazilamadi: {e}")

        logger.info(
            f"Excel okundu | {sonuc.dosya} | "
            f"Birim:{sonuc.anabilim_dali}/{sonuc.birim} | "
            f"Toplam:{sonuc.toplam_satir} "
            f"Geçerli:{sonuc.gecerli} "
            f"Hatalı:{sonuc.hatali} "
            f"Uyarılı:{sonuc.uyarili}"
        )
        return SonucYonetici.tamam(veri=sonuc)

    def _kolon_haritasi_bul(self, ws, baslik_satiri: int) -> dict[str, int]:
        harita = {}
        for col_idx in range(1, 15):
            val = ws.cell(row=baslik_satiri, column=col_idx).value
            if val is None:
                continue
            val_str = str(val).strip()
            for baslik, ic_ad in KOLON_HARITA.items():
                if baslik.strip() == val_str:
                    harita[ic_ad] = col_idx
                    break
        return harita

    # ── 2. Satır doğrulama ───────────────────────────────────

    def _satiri_isle(
        self,
        row_idx: int,
        row_vals: dict,
        donem_ay: int,
        donem_yil: int,
        anabilim_dali: str = "",
        birim: str = "",
        katsayi_cache: Optional[dict] = None,
    ) -> SatirSonucu:
        hatalar, uyarilar = [], []
        veri = {
            "DonemAy":      donem_ay,
            "DonemYil":     donem_yil,
            "AnaBilimDali": anabilim_dali,
            "Birim":        birim,
        }

        # TC Kimlik — opsiyonel
        tc = _parse_tc(row_vals.get("TCKimlik"))
        if tc and not TC_REGEX.match(tc):
            uyarilar.append(f"TC Kimlik format hatası: '{tc}' (11 rakam olmalı)")
        veri["TCKimlik"] = tc

        # Ad Soyad — zorunlu
        ad = str(row_vals.get("AdSoyad") or "").strip()
        if not ad:
            hatalar.append("Ad Soyad boş")
        else:
            veri["AdSoyad"] = ad

        # Çalışılan Alan — zorunlu (Excel'de birim adını tanımlar, aynı zamanda gösterim için saklanır)
        alan = str(row_vals.get("IslemTipi") or "").strip()
        if not alan:
            hatalar.append("Çalışılan Alan boş")
        else:
            veri["IslemTipi"] = alan

        # Katsayı — önce cache'e bak (thread-safe), yoksa servise düş
        katsayi_kayit = None
        if katsayi_cache is not None:
            katsayi_kayit = katsayi_cache.get((anabilim_dali, birim))
        elif self._katsayi_service:
            sonuc_katsayi = self._katsayi_service.get_aktif_katsayi(anabilim_dali, birim)
            if sonuc_katsayi.basarili:
                katsayi_kayit = sonuc_katsayi.data

        if katsayi_kayit:
            veri["Katsayi"]   = katsayi_kayit.get("Katsayi")
            veri["OrtSureDk"] = katsayi_kayit.get("OrtSureDk", 0)
        else:
            hatalar.append(
                f"{anabilim_dali} / {birim} için aktif katsayı protokolü yok. "
                f"Lütfen önce Katsayı Protokolleri ekranından protokol tanımlayın."
            )
        
        # Vaka Sayısı — zorunlu
        vaka = _parse_int(row_vals.get("VakaSayisi"))
        if vaka is None:
            hatalar.append("Vaka Sayısı boş veya sayı değil")
        elif vaka <= 0:
            hatalar.append("Vaka Sayısı 0 olamaz")
        else:
            if vaka > 500:
                uyarilar.append(f"Vaka Sayısı yüksek ({vaka}) — kontrol edin")
            veri["VakaSayisi"] = vaka

        # Hesaplanan Saat + Toplam İşlem Süresi
        if "VakaSayisi" in veri and "Katsayi" in veri:
            vaka    = veri["VakaSayisi"]
            katsayi = veri["Katsayi"]
            if vaka is not None and katsayi is not None:
                veri["HesaplananSaat"] = round(vaka * katsayi, 2)
                veri["ToplamSureDk"]   = vaka * veri.get("OrtSureDk", 0)

        # ── K1 Matematiksel Denetim ──────────────────────────
        # HBYS gerekmez — sadece sayılarla fiziksel imkânsızlıkları yakala
        if "VakaSayisi" in veri and donem_ay and donem_yil:
            try:
                import calendar as _cal
                from datetime import date as _date
                toplam_gun = _cal.monthrange(donem_yil, donem_ay)[1]
                is_gunu    = sum(
                    1 for g in range(1, toplam_gun + 1)
                    if _date(donem_yil, donem_ay, g).weekday() < 5
                )
                vaka_k1     = veri["VakaSayisi"]
                ort_sure_k1 = veri.get("OrtSureDk") or 0
                saat_k1     = veri.get("HesaplananSaat") or 0

                # Fiziksel imkânsız: tek kişinin toplam işlem süresi > aylık iş kapasitesi
                if ort_sure_k1 > 0:
                    toplam_islem_dk = vaka_k1 * ort_sure_k1
                    kapasite_dk     = is_gunu * 8 * 60   # 8 saatlik iş günü
                    if toplam_islem_dk > kapasite_dk:
                        hatalar.append(
                            f"K1 FİZİKSEL İMKNSIZ: {vaka_k1} vaka × {ort_sure_k1} dk = "
                            f"{toplam_islem_dk} dk > {kapasite_dk} dk ({is_gunu} iş günü × 8 saat)"
                        )

                # Hesaplanan saat > günlük mesai × iş günü
                if saat_k1 > is_gunu * 8:
                    hatalar.append(
                        f"K1 SAAT AŞIMI: {saat_k1:.1f} saat > "
                        f"{is_gunu * 8} saat ({is_gunu} iş günü × 8 saat)"
                    )

                # Vaka sayısı makul uyarı sınırı (ort_sure biliniyorsa)
                if ort_sure_k1 > 0:
                    makul_vaka = (is_gunu * 8 * 60) // ort_sure_k1
                    if vaka_k1 > makul_vaka * 0.8:
                        uyarilar.append(
                            f"K1 YÜKSEK VAKA: {vaka_k1} vaka, "
                            f"bu dönem için makul üst sınır ~{int(makul_vaka)}"
                        )
            except Exception:
                pass

        # TutanakNo ve TutanakTarihi → kaydet() aşamasında doldurulur

        return SatirSonucu(
            satir_no=row_idx,
            veri=veri,
            hatalar=hatalar,
            uyarilar=uyarilar,
            gecerli=len(hatalar) == 0,
        )
    def get_birim_listesi(self):
        if not self._katsayi_service:
            return []
        return self._katsayi_service.get_birim_listesi()

    # ── 3. Kaydetme ──────────────────────────────────────────

    def kaydet(
        self,
        sonuc: ImportSonucu,
        dosya_yolu: str,
        kaydeden: Optional[str] = None,
    ) -> SonucYonetici:
        """
        Onaylanan satırları DB'ye yazar.

        Adımlar:
          1. Excel dosyasını Dokumanlar tablosuna kaydet → DokumanId al
          2. DokumanId'yi TutanakNo, import tarihini TutanakTarihi olarak
             tüm satırlara yaz
          3. Dis_Alan_Calisma tablosuna satırları ekle
        """

        if not self._r:
            return SonucYonetici.hata(RuntimeError("RepositoryRegistry bağlı değil"), "DisAlanImportService.kaydet")

        try:
            # ── Adım 1: Dokumanlar kaydı ─────────────────────────
            dokuman_id    = self._dokumanlar_kaydet(sonuc, dosya_yolu, kaydeden)
            tutanak_no    = dokuman_id
            tutanak_tarihi = datetime.now().strftime("%Y-%m-%d")

            sonuc.dokuman_id   = dokuman_id
            sonuc.tutanak_no   = tutanak_no
            sonuc.import_tarihi = tutanak_tarihi

            logger.info(f"Tutanak oluşturuldu | DokumanId: {dokuman_id}")

            # ── Adım 2–3: Satır kayıtları ─────────────────────────
            repo = self._r.get("Dis_Alan_Calisma")
            kaydedilen = atlanan = 0

            for satir in sonuc.satirlar:
                if not satir.gecerli and not satir.kullanici_onayladi:
                    atlanan += 1
                    continue

                veri = {
                    **satir.veri,
                    "TutanakNo":          tutanak_no,
                    "TutanakTarihi":      tutanak_tarihi,
                    "KaydedenKullanici":  kaydeden or "Import",
                }

                try:
                    pk = (
                        str(veri.get("TCKimlik",  "")),
                        str(veri.get("DonemAy",   "")),
                        str(veri.get("DonemYil",  "")),
                        str(veri.get("TutanakNo", "")),
                    )
                    if repo.get_by_pk(pk):
                        satir.uyarilar.append("Zaten kayıtlı — atlandı")
                        atlanan += 1
                        continue

                    repo.insert(veri)
                    kaydedilen += 1
                    logger.info(
                        f"Import | {veri.get('AdSoyad')} | "
                        f"{veri.get('IslemTipi')} | "
                        f"Vaka:{veri.get('VakaSayisi')} | "
                        f"{veri.get('HesaplananSaat', 0):.2f} saat"
                    )
                except Exception as e:
                    satir.hatalar.append(f"DB hatası: {e}")
                    satir.gecerli = False
                    atlanan += 1
                    logger.error(f"Import satır {satir.satir_no}: {e}")

            sonuc.kaydedilen = kaydedilen
            sonuc.atlanan    = atlanan
            logger.info(
                f"Import tamamlandı | "
                f"TutanakNo:{tutanak_no[:8]}… | "
                f"Kaydedilen:{kaydedilen} Atlanan:{atlanan}"
            )
            return SonucYonetici.tamam(veri=sonuc)
        except Exception as e:
            return SonucYonetici.hata(e, "DisAlanImportService.kaydet")

    def _dokumanlar_kaydet(
        self,
        sonuc: ImportSonucu,
        dosya_yolu: str,
        kaydeden: Optional[str],
    ) -> str:
        """
        Excel dosyasını Dokumanlar tablosına kaydeder.
        Dosyayı arşiv dizinine kopyalar (varsa).
        Üretilen DokumanId'yi döndürür.
        """
        dokuman_id = str(uuid.uuid4())
        dosya_adi  = Path(dosya_yolu).name
        tarih      = datetime.now().strftime("%Y-%m-%d")

        # Arşiv dizinine kopyala
        local_path = dosya_yolu
        if self._arsiv_dizin:
            self._arsiv_dizin.mkdir(parents=True, exist_ok=True)
            hedef = self._arsiv_dizin / f"{dokuman_id}_{dosya_adi}"
            try:
                shutil.copy2(dosya_yolu, hedef)
                local_path = str(hedef)
            except Exception as e:
                logger.warning(f"Dosya arşivlenemedi: {e}")

        # EntityId: dönem + birim (arama kolaylığı için)
        ay   = sonuc.satirlar[0].veri.get("DonemAy",  "?") if sonuc.satirlar else "?"
        yil  = sonuc.satirlar[0].veri.get("DonemYil", "?") if sonuc.satirlar else "?"
        birim_kisa = f"{sonuc.anabilim_dali}_{sonuc.birim}".replace(" ", "_")[:30]
        entity_id  = f"{ay}_{yil}_{birim_kisa}"

        dokuman_kayit = {
            "EntityType":       "DisAlanCalisma",
            "EntityId":         entity_id,
            "BelgeTuru":        "Tutanak",
            "Belge":            dosya_adi,
            "DokumanId":        dokuman_id,
            "DocType":          "xlsx",
            "DisplayName":      (
                f"Dış Alan Bildirim — "
                f"{sonuc.anabilim_dali} / {sonuc.birim} — "
                f"{sonuc.donem_etiket}"
            ),
            "LocalPath":        local_path,
            "BelgeAciklama":    (
                f"Import: {kaydeden or 'sistem'} | "
                f"Satır: {sonuc.toplam_satir} | "
                f"Geçerli: {sonuc.gecerli}"
            ),
            "YuklenmeTarihi":   tarih,
            "DrivePath":        "",
            "IliskiliBelgeID":  "",
            "IliskiliBelgeTipi": "",
        }

        try:
            if not self._r:
                raise RuntimeError("RepositoryRegistry bağlı değil")
            self._r.get("Dokumanlar").insert(dokuman_kayit)
            logger.info(
                f"Dokumanlar kaydı oluşturuldu | "
                f"DokumanId:{dokuman_id} | {dosya_adi}"
            )
        except Exception as e:
            logger.error(f"Dokumanlar kayıt hatası: {e}")
            # Döküman kaydı başarısız olsa bile import devam edebilir
            # TutanakNo olarak yine de dokuman_id kullanılır

        return dokuman_id
