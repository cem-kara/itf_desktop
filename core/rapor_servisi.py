# -*- coding: utf-8 -*-
"""
Rapor Servisi  —  core/rapor_servisi.py
═══════════════════════════════════════════════════════════════════════════════
Tüm uygulama genelinde tek nokta: şablon dosyasını yükle → veriyi doldur →
Excel (.xlsx) veya PDF (.pdf) olarak kaydet.

─── ŞABLON KONUMU ───────────────────────────────────────────────────────────
    data/
    └── templates/
        ├── excel/   *.xlsx      (openpyxl ile doldurulur)
        └── pdf/     *.html      (Jinja2 render → Qt QPdfWriter)

─── EXCEL ŞABLON KURALLARI ──────────────────────────────────────────────────
  • Skalar yer tutucular  : hücre içeriği = {{alan_adi}}
                            Context'ten karşılığı yazılır.
  • Tablo satır şablonu   : İlk sütunu tam olarak {{ROW}} içeren satır.
                            O satırın diğer hücrelerinde {{kolon_adi}} yer
                            tutucuları bulunur.
                            Servis bu satırı N kez tekrarlayarak tabloyu
                            genişletir; biçimlendirme (renk, font, kenarlık,
                            hizalama) kopyalanır.
  • Satır numarası        : {{#}} yer tutucusu → 1'den başlayan sıra numarası.

  Örnek (A sütunu = {{ROW}}, B = {{Cihaz}}, C = {{Tarih}}, D = {{Durum}}):
    A2: {{ROW}}   B2: {{Cihaz}}   C2: {{Tarih}}   D2: {{Durum}}

─── PDF ŞABLON KURALLARI ────────────────────────────────────────────────────
  • Standart Jinja2 sözdizimi: {{ degisken }}, {% for satir in tablo %}...{% endfor %}
  • Context anahtarları doğrudan erişilebilir.
  • `tablo` listesi otomatik olarak context'e eklenir.

─── KULLANIM ─────────────────────────────────────────────────────────────────
    from core.rapor_servisi import RaporServisi

    # Excel
    yol = RaporServisi.excel(
        sablon      = "kalibrasyon_listesi",     # templates/excel/kalibrasyon_listesi.xlsx
        context     = {"baslik": "Rapor", "tarih": "16.02.2026"},
        tablo       = [{"Cihaz": "CT", "Tarih": "2026-01-15", "Durum": "OK"}],
        kayit_yolu  = "/tmp/rapor.xlsx",         # None → geçici dosyaya yazar
    )

    # PDF
    yol = RaporServisi.pdf(
        sablon     = "kalibrasyon_listesi",      # templates/pdf/kalibrasyon_listesi.html
        context    = {"baslik": "Rapor", ...},
        tablo      = [...],
        kayit_yolu = "/tmp/rapor.pdf",
    )

    # Dosyayı OS'un varsayılan programıyla aç
    RaporServisi.ac(yol)

    # QWidget üzerinde kayıt diyaloğu
    yol = RaporServisi.kaydet_diyalogu(parent, "rapor", tur="excel")
═══════════════════════════════════════════════════════════════════════════════
"""
from __future__ import annotations

import copy
import os
import re
import shutil
import subprocess
import sys
import tempfile
from pathlib import Path
from typing import Any

from core.logger import logger
from core.paths import BASE_DIR

# ── Şablon dizinleri ─────────────────────────────────────────────────────────
TEMPLATES_DIR  = Path(BASE_DIR) / "data" / "templates"
EXCEL_TMPL_DIR = TEMPLATES_DIR / "excel"
PDF_TMPL_DIR   = TEMPLATES_DIR / "pdf"

# Dizinler yoksa oluştur (kurulum gerektirmez)
for _d in [EXCEL_TMPL_DIR, PDF_TMPL_DIR]:
    _d.mkdir(parents=True, exist_ok=True)

# ── Regex: {{placeholder}} ───────────────────────────────────────────────────
_PH = re.compile(r"\{\{([^}]+)\}\}")


# ═══════════════════════════════════════════════════════════════════════════
#  İÇ YARDIMCILAR
# ═══════════════════════════════════════════════════════════════════════════

def _deger_str(value: Any) -> str:
    """Hücreye yazılacak değeri güvenli stringe çevirir."""
    if value is None:
        return ""
    return str(value)


def _ph_doldur(text: str, context: dict, satir_no: int | None = None) -> Any:
    """
    Bir hücre metnindeki {{placeholder}} yer tutucularını context'ten doldurur.
    Metin *sadece* bir yer tutucudan oluşuyorsa orijinal Python tipini korur
    (sayı, tarih vb. Excel'de doğru tip olarak görünsün diye).
    {{#}} → satir_no
    """
    if not isinstance(text, str):
        return text

    # {{#}} → sıra numarası
    if satir_no is not None:
        text = text.replace("{{#}}", str(satir_no))

    eslesme = _PH.fullmatch(text.strip())
    if eslesme:
        anahtar = eslesme.group(1).strip()
        return context.get(anahtar, "")

    def _yedek(m):
        return _deger_str(context.get(m.group(1).strip(), ""))

    return _PH.sub(_yedek, text)


def _hucre_stilini_kopyala(kaynak, hedef) -> None:
    """openpyxl hücresi stilini (font, fill, border, alignment, number_format) kopyalar."""
    if kaynak.has_style:
        hedef.font        = copy.copy(kaynak.font)
        hedef.fill        = copy.copy(kaynak.fill)
        hedef.border      = copy.copy(kaynak.border)
        hedef.alignment   = copy.copy(kaynak.alignment)
        hedef.number_format = kaynak.number_format


# ═══════════════════════════════════════════════════════════════════════════
#  EXCEL ŞABLON İŞLEMCİSİ
# ═══════════════════════════════════════════════════════════════════════════

class _ExcelSablon:
    """
    Bir .xlsx şablon dosyasını yükler, {{ROW}} satırını N kez
    genişletir, tüm yer tutucuları doldurur ve sonucu kaydeder.
    """

    def __init__(self, sablon_yolu: Path):
        from openpyxl import load_workbook
        self._wb = load_workbook(str(sablon_yolu))
        self._sablon_yolu = sablon_yolu

    # ── Genel arayüz ─────────────────────────────────────────────────────

    def doldur_ve_kaydet(
        self,
        context: dict,
        tablo: list[dict],
        kayit_yolu: str,
    ) -> str:
        for ws in self._wb.worksheets:
            self._sayfayi_isle(ws, context, tablo)
        self._wb.save(kayit_yolu)
        return kayit_yolu

    # ── İç işleyiciler ───────────────────────────────────────────────────

    def _sayfayi_isle(self, ws, context: dict, tablo: list[dict]) -> None:
        row_satir = self._row_satiri_bul(ws)
        if row_satir is not None:
            self._tabloyu_genislet(ws, row_satir, tablo)
        self._skalarlari_doldur(ws, context)

    def _row_satiri_bul(self, ws) -> int | None:
        """{{ROW}} yer tutucusunu içeren ilk satırın indeksini döndürür."""
        for satir in ws.iter_rows():
            for hucre in satir:
                if isinstance(hucre.value, str) and hucre.value.strip() == "{{ROW}}":
                    return hucre.row
        return None

    def _tabloyu_genislet(self, ws, sablon_satir: int, tablo: list[dict]) -> None:
        """
        sablon_satir'daki tablo satır şablonunu tablo uzunluğu kadar tekrarlar.
        Şablon satırının biçimlendirmesi her veri satırına kopyalanır.
        """
        if not tablo:
            # Boş tablo: sadece şablon satırını temizle
            for hucre in ws[sablon_satir]:
                hucre.value = None
            return

        # Şablon satırının hücrelerini kaydet
        sablon_hucreleri = [
            {"deger": h.value, "kolon": h.column, "hucre": h}
            for h in ws[sablon_satir]
            if h.value is not None
        ]

        # Şablon satırından sonraki satırları aşağı kaydır
        n = len(tablo)
        if n > 1:
            ws.insert_rows(sablon_satir + 1, n - 1)

        # Her veri satırını doldur
        for i, kayit in enumerate(tablo):
            hedef_satir = sablon_satir + i
            for sh in sablon_hucreleri:
                hedef = ws.cell(row=hedef_satir, column=sh["kolon"])
                if i > 0:
                    # Stil kopyala (insert_rows yeni boş satır ekler)
                    _hucre_stilini_kopyala(sh["hucre"], hedef)

                metin = sh["deger"]
                if isinstance(metin, str) and metin.strip() == "{{ROW}}":
                    hedef.value = None   # {{ROW}} hücresini temizle
                    continue

                # Satır numarası desteği
                kayit_ctx = {**kayit, "#": i + 1}
                hedef.value = _ph_doldur(metin, kayit_ctx, satir_no=i + 1)

    def _skalarlari_doldur(self, ws, context: dict) -> None:
        """{{placeholder}} içeren tüm hücreleri context'ten doldurur."""
        for satir in ws.iter_rows():
            for hucre in satir:
                if isinstance(hucre.value, str) and _PH.search(hucre.value):
                    yeni = _ph_doldur(hucre.value, context)
                    hucre.value = yeni


# ═══════════════════════════════════════════════════════════════════════════
#  PDF ŞABLON İŞLEMCİSİ  (Jinja2 → QPdfWriter)
# ═══════════════════════════════════════════════════════════════════════════

class _PDFSablon:
    """
    Jinja2 HTML şablon dosyasını render edip Qt'nin QPdfWriter'ı ile PDF üretir.
    Qt olmayan ortamda (test) graceful fail eder.
    """

    def __init__(self, sablon_yolu: Path):
        self._sablon_yolu = sablon_yolu

    def doldur_ve_kaydet(
        self,
        context: dict,
        tablo: list[dict],
        kayit_yolu: str,
    ) -> str:
        html = self._render(context, tablo)
        self._pdf_yaz(html, kayit_yolu)
        return kayit_yolu

    def _render(self, context: dict, tablo: list[dict]) -> str:
        from jinja2 import Environment, FileSystemLoader, select_autoescape
        env = Environment(
            loader=FileSystemLoader(str(self._sablon_yolu.parent)),
            autoescape=select_autoescape(["html"]),
        )
        sablon = env.get_template(self._sablon_yolu.name)
        return sablon.render(**context, tablo=tablo)

    @staticmethod
    def _pdf_yaz(html: str, kayit_yolu: str) -> None:
        """
        HTML içeriğini PDF olarak yazar.
        Qt tam kurulu değilse (CI, test) _preview.html olarak kaydeder.
        """
        def _html_fallback(reason: str) -> None:
            logger.warning(f"PDF yazarken {reason} — HTML olarak kaydediliyor")
            html_yol = kayit_yolu.replace(".pdf", "_preview.html")
            Path(html_yol).write_text(html, encoding="utf-8")

        try:
            from PySide6.QtGui import QTextDocument, QPdfWriter, QPageSize
            from PySide6.QtCore import QMarginsF

            # Gerçek Qt nesnesi mi yoksa stub mu kontrol et
            doc = QTextDocument()
            if not callable(getattr(doc, "setHtml", None)):
                _html_fallback("Qt stub tespit edildi")
                return

            from PySide6.QtGui import QPageLayout
            doc.setHtml(html)

            writer = QPdfWriter(kayit_yolu)
            writer.setPageSize(QPageSize(QPageSize.A4))
            writer.setPageMargins(QMarginsF(15, 15, 15, 15))
            writer.setResolution(150)

            doc.print_(writer)
            logger.info(f"PDF kaydedildi: {kayit_yolu}")

        except ImportError:
            _html_fallback("Qt bulunamadı")
        except Exception as e:
            logger.error(f"QPdfWriter hatası: {e}")
            _html_fallback(f"hata: {e}")


# ═══════════════════════════════════════════════════════════════════════════
#  ANA SERVİS  (public API)
# ═══════════════════════════════════════════════════════════════════════════

class RaporServisi:
    """
    Tüm uygulamadan çağrılan statik rapor servisi.
    Instance oluşturmak gerekmez.
    """

    # ── Excel ─────────────────────────────────────────────────────────────

    @staticmethod
    def excel(
        sablon: str,
        context: dict,
        tablo: list[dict] | None = None,
        kayit_yolu: str | None = None,
    ) -> str | None:
        """
        Excel raporu üretir.

        Parameters
        ----------
        sablon      : Şablon adı (uzantısız). ``data/templates/excel/<sablon>.xlsx``
        context     : Skalar yer tutucular. ``{"tarih": "16.02.2026", ...}``
        tablo       : Tablo satır listesi. Her dict bir satırı temsil eder.
        kayit_yolu  : Tam dosya yolu. None → geçici dizine kaydedilir.

        Returns
        -------
        str | None  : Oluşturulan dosyanın yolu; hata durumunda None.
        """
        sablon_yolu = EXCEL_TMPL_DIR / f"{sablon}.xlsx"
        if not sablon_yolu.exists():
            logger.error(f"Excel şablonu bulunamadı: {sablon_yolu}")
            return None

        if kayit_yolu is None:
            kayit_yolu = str(Path(tempfile.mkdtemp()) / f"{sablon}.xlsx")

        try:
            return _ExcelSablon(sablon_yolu).doldur_ve_kaydet(
                context    = context,
                tablo      = tablo or [],
                kayit_yolu = kayit_yolu,
            )
        except Exception as e:
            logger.error(f"Excel rapor hatası [{sablon}]: {e}")
            return None

    # ── PDF ───────────────────────────────────────────────────────────────

    @staticmethod
    def pdf(
        sablon: str,
        context: dict,
        tablo: list[dict] | None = None,
        kayit_yolu: str | None = None,
    ) -> str | None:
        """
        PDF raporu üretir.

        Parameters
        ----------
        sablon      : Şablon adı (uzantısız). ``data/templates/pdf/<sablon>.html``
        context     : Jinja2 değişkenler. ``{"baslik": "...", "tarih": "..."}``
        tablo       : ``{% for satir in tablo %}`` döngüsü için veri.
        kayit_yolu  : Tam dosya yolu. None → geçici dizine kaydedilir.
        """
        sablon_yolu = PDF_TMPL_DIR / f"{sablon}.html"
        if not sablon_yolu.exists():
            logger.error(f"PDF şablonu bulunamadı: {sablon_yolu}")
            return None

        if kayit_yolu is None:
            kayit_yolu = str(Path(tempfile.mkdtemp()) / f"{sablon}.pdf")

        try:
            return _PDFSablon(sablon_yolu).doldur_ve_kaydet(
                context    = context,
                tablo      = tablo or [],
                kayit_yolu = kayit_yolu,
            )
        except Exception as e:
            logger.error(f"PDF rapor hatası [{sablon}]: {e}")
            return None

    # ── Yardımcılar ───────────────────────────────────────────────────────

    @staticmethod
    def ac(dosya_yolu: str) -> None:
        """Dosyayı işletim sisteminin varsayılan programıyla açar."""
        try:
            if sys.platform == "win32":
                os.startfile(dosya_yolu)
            elif sys.platform == "darwin":
                subprocess.run(["open", dosya_yolu], check=True)
            else:
                subprocess.run(["xdg-open", dosya_yolu], check=True)
        except Exception as e:
            logger.warning(f"Dosya açma hatası: {e}")

    @staticmethod
    def kaydet_diyalogu(
        parent,
        varsayilan_isim: str,
        tur: str = "excel",
    ) -> str | None:
        """
        Qt dosya kaydetme diyaloğu gösterir.

        Parameters
        ----------
        parent          : QWidget — diyalog için ebeveyn
        varsayilan_isim : Önerilen dosya adı (uzantısız)
        tur             : "excel" | "pdf"

        Returns
        -------
        str | None : Seçilen dosya yolu; iptal edilirse None.
        """
        try:
            from PySide6.QtWidgets import QFileDialog
            if tur == "excel":
                filtre = "Excel Dosyası (*.xlsx)"
                uzanti = ".xlsx"
            else:
                filtre = "PDF Dosyası (*.pdf)"
                uzanti = ".pdf"

            yol, _ = QFileDialog.getSaveFileName(
                parent,
                "Raporu Kaydet",
                f"{varsayilan_isim}{uzanti}",
                filtre,
            )
            return yol if yol else None
        except Exception as e:
            logger.error(f"Kaydet diyaloğu hatası: {e}")
            return None

    @staticmethod
    def sablon_listesi(tur: str = "excel") -> list[str]:
        """
        Mevcut şablon adlarını (uzantısız) döndürür.
        tur: "excel" | "pdf"
        """
        dizin = EXCEL_TMPL_DIR if tur == "excel" else PDF_TMPL_DIR
        uzanti = ".xlsx" if tur == "excel" else ".html"
        return [p.stem for p in dizin.glob(f"*{uzanti}") if p.is_file()]

    @staticmethod
    def sablon_yolu(sablon: str, tur: str = "excel") -> Path:
        """Şablon dosyasının tam yolunu döndürür (var olup olmadığına bakmaz)."""
        dizin  = EXCEL_TMPL_DIR if tur == "excel" else PDF_TMPL_DIR
        uzanti = ".xlsx" if tur == "excel" else ".html"
        return dizin / f"{sablon}{uzanti}"
