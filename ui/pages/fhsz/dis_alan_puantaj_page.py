# ui/pages/fhsz/dis_alan_puantaj_page.py
# -*- coding: utf-8 -*-
"""
Dış Alan Puantaj Raporu — Kişi Bazlı Dönem Özeti

Her kişi için:
  - Aylık vaka ve hesaplanan saat
  - Yıllık kümülatif saat
  - İzin günü hakkı
  - RKS onay durumu

Çıktı: PDF (resmi dönem özeti) + Excel
"""
from datetime import datetime

from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton,
    QComboBox, QFrame, QTableWidget, QTableWidgetItem,
    QHeaderView, QAbstractItemView, QMessageBox, QFileDialog,
)
from PySide6.QtCore import Qt, QDate
from PySide6.QtGui import QColor, QBrush

from ui.styles.icons import IconRenderer
from core.di import get_dis_alan_service
from core.services.dis_alan_service import _izin_gunu_hesapla
from core.hata_yonetici import soru_sor, bilgi_goster, hata_goster
from core.logger import logger

AY_ADLARI = ["Ocak","Şubat","Mart","Nisan","Mayıs","Haziran",
              "Temmuz","Ağustos","Eylül","Ekim","Kasım","Aralık"]

def _int(v):
    try: return int(v)
    except: return 0

def _float(v):
    try: return float(str(v).replace(",","."))
    except: return 0.0

def _kisi_key(r):
    tc = str(r.get("TCKimlik","")).strip()
    return tc if tc else str(r.get("AdSoyad","")).strip().upper()


class DisAlanPuantajPage(QWidget):
    def __init__(self, db=None, parent=None):
        super().__init__(parent)
        self.setProperty("bg-role", "page")
        self._db   = db
        self._svc  = get_dis_alan_service(db) if db else None
        self._rows_cache: list[dict] = []
        self._tablo_rows: list[dict] = []
        self._setup_ui()
        self._connect_signals()
        self._load_all_rows()

    # ── UI ─────────────────────────────────────────────────────

    def _setup_ui(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(20, 15, 20, 15)
        root.setSpacing(10)

        # Filtre
        top = QFrame()
        top.setProperty("bg-role", "panel")
        top.setMaximumHeight(56)
        tl = QHBoxLayout(top)
        tl.setContentsMargins(12,6,12,6); tl.setSpacing(10)

        lbl = QLabel("Puantaj Raporu")
        lbl.setProperty("style-role", "section-title")
        tl.addWidget(lbl); tl.addStretch()

        tl.addWidget(QLabel("Anabilim Dalı:"))
        self.cmb_ana = QComboBox(); self.cmb_ana.setFixedWidth(190)
        tl.addWidget(self.cmb_ana)

        tl.addWidget(QLabel("Birim:"))
        self.cmb_bir = QComboBox(); self.cmb_bir.setFixedWidth(150)
        tl.addWidget(self.cmb_bir)

        tl.addWidget(QLabel("Dönem:"))
        self.cmb_ay = QComboBox()
        self.cmb_ay.addItems(AY_ADLARI)
        self.cmb_ay.setCurrentIndex(QDate.currentDate().month()-1)
        self.cmb_ay.setFixedWidth(90)
        tl.addWidget(self.cmb_ay)

        self.cmb_yil = QComboBox()
        self.cmb_yil.addItems([str(y) for y in range(
            QDate.currentDate().year()-2, QDate.currentDate().year()+2)])
        self.cmb_yil.setCurrentText(str(QDate.currentDate().year()))
        self.cmb_yil.setFixedWidth(80)
        tl.addWidget(self.cmb_yil)

        self.btn_getir = QPushButton("Getir")
        self.btn_getir.setProperty("style-role", "action")
        self.btn_getir.setFixedHeight(36); self.btn_getir.setFixedWidth(80)
        tl.addWidget(self.btn_getir)
        root.addWidget(top)

        # Özet kartlar
        kl = QHBoxLayout(); kl.setSpacing(8)
        self.k_personel = self._kart("Personel",          "—", "#457B9D")
        self.k_vaka     = self._kart("Toplam Vaka",        "—", "#457B9D")
        self.k_saat     = self._kart("Toplam Saat",        "—", "#457B9D")
        self.k_ort      = self._kart("Kişi Başı Ort.",     "—", "#457B9D")
        self.k_izin     = self._kart("Max İzin Günü",      "—", "#E65100")
        self.k_onay     = self._kart("RKS Onaylı",         "—", "#1B5E20")
        for k in [self.k_personel, self.k_vaka, self.k_saat,
                  self.k_ort, self.k_izin, self.k_onay]:
            kl.addWidget(k)
        root.addLayout(kl)

        # Tablo
        kolonlar = [
            ("TC Kimlik",     120), ("Ad Soyad",     200),
            ("Birim",         130), ("Vaka",           60),
            ("Aylık Saat",     85), ("Küm. Saat",      85),
            ("İzin Günü",      80), ("Onay",           80),
        ]
        self.tablo = QTableWidget()
        self.tablo.setColumnCount(len(kolonlar))
        self.tablo.setHorizontalHeaderLabels([c[0] for c in kolonlar])
        self.tablo.verticalHeader().setVisible(False)
        self.tablo.setProperty("style-role", "table")
        self.tablo.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.tablo.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.tablo.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)
        hdr = self.tablo.horizontalHeader()
        for i, (_, w) in enumerate(kolonlar):
            if kolonlar[i][0] == "Ad Soyad":
                hdr.setSectionResizeMode(i, QHeaderView.ResizeMode.Stretch)
            else:
                hdr.setSectionResizeMode(i, QHeaderView.ResizeMode.Fixed)
                self.tablo.setColumnWidth(i, w)
        root.addWidget(self.tablo)

        # Alt butonlar
        bot = QFrame()
        bot.setProperty("bg-role", "panel")
        bot.setMaximumHeight(50)
        bl = QHBoxLayout(bot)
        bl.setContentsMargins(12,6,12,6); bl.setSpacing(10)

        self.btn_ozet = QPushButton("Dönem Özetini Kaydet")
        self.btn_ozet.setProperty("style-role", "action")
        self.btn_ozet.setFixedHeight(34)
        self.btn_ozet.setEnabled(False)
        IconRenderer.set_button_icon(self.btn_ozet, "save", color="#FFFFFF")

        self.btn_onayla = QPushButton("RKS Onayla")
        self.btn_onayla.setProperty("style-role", "success-filled")
        self.btn_onayla.setFixedHeight(34)
        self.btn_onayla.setEnabled(False)

        self.btn_pdf = QPushButton("PDF Rapor")
        self.btn_pdf.setProperty("style-role", "danger")
        self.btn_pdf.setFixedHeight(34)
        self.btn_pdf.setEnabled(False)

        self.btn_excel = QPushButton("Excel'e Aktar")
        self.btn_excel.setProperty("style-role", "secondary")
        self.btn_excel.setFixedHeight(34)
        self.btn_excel.setEnabled(False)

        bl.addWidget(self.btn_ozet)
        bl.addWidget(self.btn_onayla)
        bl.addStretch()
        bl.addWidget(self.btn_pdf)
        bl.addWidget(self.btn_excel)

        self.lbl_durum = QLabel("")
        self.lbl_durum.setProperty("style-role", "info")
        bl.addWidget(self.lbl_durum)
        root.addWidget(bot)

    def _kart(self, baslik, val, renk="#457B9D"):
        f = QFrame()
        f.setProperty("bg-role", "panel")
        lay = QVBoxLayout(f)
        lay.setContentsMargins(10,6,10,6); lay.setSpacing(2)
        lb = QLabel(baslik)
        lb.setProperty("style-role", "stat-label")
        lv = QLabel(str(val))
        lv.setProperty("style-role", "stat-value")
        lay.addWidget(lb); lay.addWidget(lv)
        setattr(f, "_lv", lv)
        return f

    def _kset(self, k, v):
        lv = getattr(k, "_lv", None)
        if lv: lv.setText(str(v))

    # ── Sinyaller ──────────────────────────────────────────────

    def _connect_signals(self):
        self.cmb_ana.currentIndexChanged.connect(self._ana_degisti)
        self.btn_getir.clicked.connect(self._load_data)
        self.btn_ozet.clicked.connect(self._ozet_kaydet)
        self.btn_onayla.clicked.connect(self._rks_onayla)
        self.btn_pdf.clicked.connect(self._pdf)
        self.btn_excel.clicked.connect(self._excel)
        self.tablo.itemSelectionChanged.connect(self._secim_degisti)

    # ── Filtre ─────────────────────────────────────────────────

    def _load_all_rows(self):
        if not self._svc:
            return
        try:
            self._rows_cache = self._svc._r.get("Dis_Alan_Calisma").get_all() or []
        except Exception as e:
            logger.error(f"PuantajPage._load_all_rows: {e}")
            self._rows_cache = []
        anabilimler = sorted({str(r.get("AnaBilimDali","")).strip()
                              for r in self._rows_cache if r.get("AnaBilimDali")})
        self.cmb_ana.blockSignals(True)
        self.cmb_ana.clear()
        self.cmb_ana.addItem("Tümü")
        self.cmb_ana.addItems(anabilimler)
        self.cmb_ana.blockSignals(False)
        self._ana_degisti()

    def _ana_degisti(self):
        ana = self.cmb_ana.currentText()
        birimler = sorted({str(r.get("Birim","")).strip()
                           for r in self._rows_cache
                           if r.get("Birim") and
                           (ana == "Tümü" or r.get("AnaBilimDali","") == ana)})
        self.cmb_bir.blockSignals(True)
        self.cmb_bir.clear()
        self.cmb_bir.addItem("Tümü")
        self.cmb_bir.addItems(birimler)
        self.cmb_bir.blockSignals(False)

    # ── Veri ───────────────────────────────────────────────────

    def _load_data(self):
        self._load_all_rows()
        ay    = self.cmb_ay.currentIndex() + 1
        yil   = int(self.cmb_yil.currentText())
        ana   = self.cmb_ana.currentText()
        birim = self.cmb_bir.currentText()

        rows = [r for r in self._rows_cache
                if _int(r.get("DonemAy")) == ay and _int(r.get("DonemYil")) == yil]
        if ana != "Tümü":
            rows = [r for r in rows if r.get("AnaBilimDali","") == ana]
        if birim != "Tümü":
            rows = [r for r in rows if r.get("Birim","") == birim]

        yil_rows = [r for r in self._rows_cache
                    if _int(r.get("DonemYil")) == yil and _int(r.get("DonemAy")) <= ay]
        if ana != "Tümü":
            yil_rows = [r for r in yil_rows if r.get("AnaBilimDali","") == ana]
        if birim != "Tümü":
            yil_rows = [r for r in yil_rows if r.get("Birim","") == birim]

        # Kümülatif
        kum: dict[str,float] = {}
        for r in yil_rows:
            k = _kisi_key(r)
            kum[k] = kum.get(k, 0.0) + _float(r.get("HesaplananSaat", 0))

        # Kişi özet
        ozet: dict[str, dict] = {}
        for r in rows:
            k = _kisi_key(r)
            if k not in ozet:
                ozet[k] = {
                    "tc":    str(r.get("TCKimlik","")).strip(),
                    "ad":    str(r.get("AdSoyad","")).strip(),
                    "ana":   str(r.get("AnaBilimDali","")).strip(),
                    "birim": str(r.get("Birim","")).strip(),
                    "vaka":  0, "saat": 0.0,
                }
            ozet[k]["vaka"] += _int(r.get("VakaSayisi", 0))
            ozet[k]["saat"] += _float(r.get("HesaplananSaat", 0))

        # Onay + kümülatif
        ozet_repo = None
        if self._svc:
            try:
                ozet_repo = self._svc._r.get("Dis_Alan_Izin_Ozet")
            except Exception:
                pass

        for k, o in ozet.items():
            o["saat"]   = round(o["saat"], 2)
            o["kum"]    = round(kum.get(k, 0.0), 2)
            o["izin"]   = int(_izin_gunu_hesapla(o["kum"]))
            o["onay"]   = 0
            if ozet_repo and o["tc"]:
                try:
                    rec = ozet_repo.get_by_pk((o["tc"], o["ad"], str(ay), str(yil)))
                    o["onay"] = int((rec or {}).get("RksOnay", 0))
                except Exception:
                    pass

        self._tablo_rows = sorted(ozet.values(), key=lambda x: x["ad"])
        self._ay = ay; self._yil = yil
        self._fill_table()
        self._update_kartlar()
        self.btn_ozet.setEnabled(bool(self._tablo_rows))
        self.btn_pdf.setEnabled(bool(self._tablo_rows))
        self.btn_excel.setEnabled(bool(self._tablo_rows))
        self.lbl_durum.setText(
            f"{AY_ADLARI[ay-1]} {yil}  |  {len(self._tablo_rows)} personel"
        )

    # ── Tablo ──────────────────────────────────────────────────

    def _fill_table(self):
        self.tablo.setRowCount(0)
        self.tablo.setRowCount(len(self._tablo_rows))
        for i, o in enumerate(self._tablo_rows):
            self.tablo.setRowHeight(i, 25)
            onay = o["onay"]
            bg   = QColor("#1B3A1B") if onay else None

            def _it(text, center=False, _bg=bg):
                it = QTableWidgetItem(str(text) if text is not None else "")
                if _bg:
                    it.setBackground(QBrush(_bg))
                if center:
                    it.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                return it

            self.tablo.setItem(i, 0, _it(o["tc"],            center=True))
            self.tablo.setItem(i, 1, _it(o["ad"]))
            self.tablo.setItem(i, 2, _it(o["birim"]))
            self.tablo.setItem(i, 3, _it(o["vaka"],          center=True))
            self.tablo.setItem(i, 4, _it(f"{o['saat']:.2f}", center=True))
            self.tablo.setItem(i, 5, _it(f"{o['kum']:.2f}",  center=True))

            iz = o["izin"]
            it_iz = _it(str(iz), center=True)
            if iz >= 10:
                it_iz.setForeground(QBrush(QColor("#FFD600")))
            elif iz >= 5:
                it_iz.setForeground(QBrush(QColor("#81C784")))
            self.tablo.setItem(i, 6, it_iz)

            it_on = _it("✓ Onaylı" if onay else "—", center=True)
            if onay:
                it_on.setForeground(QBrush(QColor("#81C784")))
            self.tablo.setItem(i, 7, it_on)

    def _update_kartlar(self):
        rows = self._tablo_rows
        if not rows:
            for k in [self.k_personel, self.k_vaka, self.k_saat,
                      self.k_ort, self.k_izin, self.k_onay]:
                self._kset(k, "—")
            return
        self._kset(self.k_personel, len(rows))
        self._kset(self.k_vaka,     sum(o["vaka"] for o in rows))
        toplam_saat = sum(o["saat"] for o in rows)
        self._kset(self.k_saat,     f"{toplam_saat:.1f}")
        self._kset(self.k_ort,      f"{toplam_saat/len(rows):.1f}")
        self._kset(self.k_izin,     max(o["izin"] for o in rows))
        self._kset(self.k_onay,     f"{sum(1 for o in rows if o['onay'])}/{len(rows)}")

    def _secim_degisti(self):
        idx = self.tablo.currentRow()
        if 0 <= idx < len(self._tablo_rows):
            self.btn_onayla.setEnabled(not self._tablo_rows[idx]["onay"])
        else:
            self.btn_onayla.setEnabled(False)

    # ── Dönem özeti kaydet ─────────────────────────────────────

    def _ozet_kaydet(self):
        if not self._tablo_rows or not self._svc:
            return
        ay, yil = self._ay, self._yil
        ok = err = 0
        hatalar = []
        for o in self._tablo_rows:
            sonuc = self._svc.ozet_hesapla_ve_kaydet(o["tc"], o["ad"], ay, yil)
            if sonuc.basarili:
                ok += 1
            else:
                err += 1
                hatalar.append(f"- {o['ad']}: {sonuc.mesaj}")

        msg = f"{ok} kişi için dönem özeti kaydedildi."
        if err:
            msg += f"\n{err} kayıt atlandı (onaylı veya hatalı)."
        bilgi_goster(self, msg)
        self._load_data()

    # ── RKS Onayla ─────────────────────────────────────────────

    def _rks_onayla(self):
        if not self._svc:
            QMessageBox.critical(self, "Hata", "Veritabanı hizmeti bulunamadı.")
            return

        idx = self.tablo.currentRow()
        if not (0 <= idx < len(self._tablo_rows)):
            return
        o   = self._tablo_rows[idx]
        ay, yil = self._ay, self._yil
        if not soru_sor(self,
            f"<b>{o['ad']}</b> — {AY_ADLARI[ay-1]} {yil}<br><br>"
            f"Aylık saat: {o['saat']:.2f}  |  "
            f"Kümülatif: {o['kum']:.2f}  |  "
            f"İzin hakkı: {o['izin']} gün<br><br>"
            "Onaylanan kayıt değiştirilemez. Devam etmek istiyor musunuz?",
            baslik="RKS Onayı"):
            return

        ozet_sonuc = self._svc.get_ozet(o["tc"], ay, yil).veri or []
        if ozet_sonuc.basarili and not ozet_sonuc.data:
            self._svc.ozet_hesapla_ve_kaydet(o["tc"], o["ad"], ay, yil)

        onay_sonuc = self._svc.ozet_onayla(o["tc"], ay, yil)
        if onay_sonuc.basarili:
            bilgi_goster(self, f"{o['ad']} — {AY_ADLARI[ay-1]} {yil} onaylandı.")
            self._load_data()
        else:
            hata_goster(self, f"Onay işlemi başarısız: {onay_sonuc.mesaj}")

    # ── PDF ────────────────────────────────────────────────────

    def _pdf(self):
        if not self._tablo_rows:
            return
        ay, yil = self._ay, self._yil
        ana   = self.cmb_ana.currentText()
        birim = self.cmb_bir.currentText()

        path, _ = QFileDialog.getSaveFileName(
            self, "PDF Kaydet",
            f"puantaj_{AY_ADLARI[ay-1]}_{yil}.pdf",
            "PDF (*.pdf)"
        )
        if not path:
            return

        try:
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.lib.styles import ParagraphStyle
            from reportlab.lib.units import mm
            from reportlab.lib import colors
            from reportlab.platypus import (
                SimpleDocTemplate, Paragraph, Spacer,
                Table, TableStyle, HRFlowable
            )
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont
            try:
                pdfmetrics.registerFont(TTFont("Arial","arial.ttf"))
                pdfmetrics.registerFont(TTFont("ArialBold","arialbd.ttf"))
                fn, fnb = "Arial", "ArialBold"
            except Exception:
                fn, fnb = "Helvetica", "Helvetica-Bold"
        except ImportError:
            QMessageBox.critical(self, "Eksik Paket",
                                 "reportlab kurulu değil.\npip install reportlab")
            return

        doc = SimpleDocTemplate(
            path, pagesize=landscape(A4),
            leftMargin=15*mm, rightMargin=15*mm,
            topMargin=15*mm, bottomMargin=15*mm
        )

        h1 = ParagraphStyle("H1", fontName=fnb, fontSize=14, leading=18, spaceAfter=3)
        h2 = ParagraphStyle("H2", fontName=fnb, fontSize=10, leading=13,
                             textColor=colors.HexColor("#1D3557"))
        sm = ParagraphStyle("S",  fontName=fn,  fontSize=8,  leading=11,
                             textColor=colors.HexColor("#555555"))
        nm = ParagraphStyle("N",  fontName=fn,  fontSize=9,  leading=13)

        rows = self._tablo_rows
        toplam_personel = len(rows)
        toplam_vaka     = sum(o["vaka"] for o in rows)
        toplam_saat     = sum(o["saat"] for o in rows)
        onay_sayisi     = sum(1 for o in rows if o["onay"])

        elements = []

        # Başlık
        elements.append(Paragraph(
            "DIŞ ALAN RADYASYON — AYLIK PUANTAJ RAPORU", h1
        ))
        elements.append(Paragraph(
            f"Dönem: {AY_ADLARI[ay-1]} {yil}  |  "
            f"Anabilim Dalı: {ana}  |  Birim: {birim}  |  "
            f"Tarih: {datetime.now().strftime('%d.%m.%Y %H:%M')}",
            sm
        ))
        elements.append(Spacer(1, 3*mm))
        elements.append(HRFlowable(
            width="100%", thickness=0.5, color=colors.HexColor("#1D3557")
        ))
        elements.append(Spacer(1, 3*mm))

        # Özet satırı
        ozet_data = [
            ["Personel", "Toplam Vaka", "Toplam Saat",
             "Kişi Başı Ort. Saat", "RKS Onaylı"],
            [str(toplam_personel),
             str(toplam_vaka),
             f"{toplam_saat:.2f}",
             f"{toplam_saat/toplam_personel:.2f}" if toplam_personel else "—",
             f"{onay_sayisi}/{toplam_personel}"],
        ]
        ozet_tbl = Table(ozet_data, colWidths=[40*mm]*5)
        ozet_tbl.setStyle(TableStyle([
            ("FONTNAME",  (0,0),(-1,-1), fn),
            ("FONTSIZE",  (0,0),(-1,-1), 9),
            ("FONTNAME",  (0,0),(-1,0),  fnb),
            ("BACKGROUND",(0,0),(-1,0),  colors.HexColor("#1D3557")),
            ("TEXTCOLOR", (0,0),(-1,0),  colors.white),
            ("ALIGN",     (0,0),(-1,-1), "CENTER"),
            ("GRID",      (0,0),(-1,-1), 0.3, colors.HexColor("#CCCCCC")),
            ("BACKGROUND",(0,1),(-1,1),  colors.HexColor("#F0F4F8")),
        ]))
        elements.append(ozet_tbl)
        elements.append(Spacer(1, 4*mm))

        # Ana tablo
        elements.append(Paragraph("Kişi Bazlı Dönem Detayı", h2))
        elements.append(Spacer(1, 2*mm))

        tbl_data = [
            ["#", "TC Kimlik", "Ad Soyad", "Birim",
             "Vaka", "Aylık Saat", "Küm. Saat", "İzin Günü", "Onay"]
        ]
        for idx, o in enumerate(rows, 1):
            tbl_data.append([
                str(idx),
                o["tc"] or "—",
                o["ad"],
                o["birim"],
                str(o["vaka"]),
                f"{o['saat']:.2f}",
                f"{o['kum']:.2f}",
                str(o["izin"]),
                "✓ Onaylı" if o["onay"] else "—",
            ])

        # Toplam satırı
        tbl_data.append([
            "", "", "TOPLAM", "",
            str(toplam_vaka),
            f"{toplam_saat:.2f}",
            "—", "—", ""
        ])

        col_w = [10*mm, 28*mm, 55*mm, 35*mm,
                 14*mm, 22*mm, 22*mm, 20*mm, 20*mm]
        ana_tbl = Table(tbl_data, colWidths=col_w, repeatRows=1)

        row_styles = []
        for i, o in enumerate(rows, 1):
            if o["onay"]:
                row_styles.append(
                    ("BACKGROUND", (0,i), (-1,i), colors.HexColor("#E8F5E9"))
                )
            elif i % 2 == 0:
                row_styles.append(
                    ("BACKGROUND", (0,i), (-1,i), colors.HexColor("#F5F5F5"))
                )

        ana_tbl.setStyle(TableStyle([
            ("FONTNAME",    (0,0), (-1,-1), fn),
            ("FONTSIZE",    (0,0), (-1,-1), 8),
            ("FONTNAME",    (0,0), (-1,0),  fnb),
            ("BACKGROUND",  (0,0), (-1,0),  colors.HexColor("#1D3557")),
            ("TEXTCOLOR",   (0,0), (-1,0),  colors.white),
            ("GRID",        (0,0), (-1,-1), 0.3, colors.HexColor("#CCCCCC")),
            ("ALIGN",       (2,1), (2,-1),  "LEFT"),
            ("ALIGN",       (0,0), (1,-1),  "CENTER"),
            ("ALIGN",       (4,0), (-1,-1), "CENTER"),
            # Son satır (toplam) kalın
            ("FONTNAME",    (0,-1), (-1,-1), fnb),
            ("BACKGROUND",  (0,-1), (-1,-1), colors.HexColor("#E3F2FD")),
        ] + row_styles))
        elements.append(ana_tbl)
        elements.append(Spacer(1, 6*mm))

        # İmza alanı
        imza_data = [
            ["Hazırlayan", "Kontrol Eden", "RKS Sorumlusu"],
            ["\n\n\n__________________",
             "\n\n\n__________________",
             "\n\n\n__________________"],
        ]
        imza_tbl = Table(imza_data, colWidths=[80*mm, 80*mm, 80*mm])
        imza_tbl.setStyle(TableStyle([
            ("FONTNAME",  (0,0),(-1,-1), fn),
            ("FONTSIZE",  (0,0),(-1,-1), 9),
            ("FONTNAME",  (0,0),(-1,0),  fnb),
            ("ALIGN",     (0,0),(-1,-1), "CENTER"),
            ("GRID",      (0,0),(-1,-1), 0.3, colors.HexColor("#CCCCCC")),
        ]))
        elements.append(imza_tbl)

        try:
            doc.build(elements)
            QMessageBox.information(self, "PDF Oluşturuldu",
                                    f"Rapor kaydedildi:\n{path}")
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"PDF oluşturulamadı:\n{e}")

    # ── Excel ──────────────────────────────────────────────────

    def _excel(self):
        if not self._tablo_rows:
            return
        ay, yil = self._ay, self._yil
        path, _ = QFileDialog.getSaveFileName(
            self, "Excel Kaydet",
            f"puantaj_{AY_ADLARI[ay-1]}_{yil}.xlsx",
            "Excel (*.xlsx)"
        )
        if not path:
            return

        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment

            wb = openpyxl.Workbook()
            ws = wb.active
            if ws is None:
                # This should not happen with a new workbook, but as a safeguard
                ws = wb.create_sheet()
            ws.title = f"{AY_ADLARI[ay-1]} {yil}"

            ws.append([
                "TC Kimlik", "Ad Soyad", "Anabilim Dalı", "Birim",
                "Vaka", "Aylık Saat", "Küm. Saat", "İzin Günü", "Onay"
            ])
            for cell in ws[1]:
                cell.font      = Font(bold=True, color="FFFFFF")
                cell.fill      = PatternFill("solid", fgColor="1D3557")
                cell.alignment = Alignment(horizontal="center")

            for o in self._tablo_rows:
                ws.append([
                    o["tc"], o["ad"], o["ana"], o["birim"],
                    o["vaka"], round(o["saat"], 2), round(o["kum"], 2),
                    o["izin"], "Onaylı" if o["onay"] else ""
                ])
                if o["onay"]:
                    for cell in ws[ws.max_row]:
                        cell.fill = PatternFill("solid", fgColor="E8F5E9")

            ws.append([
                "", "TOPLAM", "", "",
                sum(o["vaka"] for o in self._tablo_rows),
                round(sum(o["saat"] for o in self._tablo_rows), 2),
                "", "", ""
            ])
            for cell in ws[ws.max_row]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill("solid", fgColor="E3F2FD")

            from openpyxl.utils import get_column_letter
            for idx, col in enumerate(ws.columns, 1):
                max_len = max(len(str(c.value or "")) for c in col)
                ws.column_dimensions[get_column_letter(idx)].width = max_len + 3

            wb.save(path)
            QMessageBox.information(self, "Kaydedildi", f"Excel kaydedildi:\n{path}")

        except ImportError:
            QMessageBox.critical(self, "Eksik Paket",
                                 "openpyxl kurulu değil.\npip install openpyxl")
        except Exception as e:
            QMessageBox.critical(self, "Hata", str(e))
