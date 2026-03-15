# ui/pages/fhsz/puantaj_rapor_page.py
# -*- coding: utf-8 -*-
"""
Dış Alan Import Denetim ve Karşılaştırma Ekranı

Aynı döneme ait birden fazla Excel import karşılaştırılır:
  - Mükerrer kişiler (aynı kişi birden fazla dosyada)
  - Eksik kişiler (bir dosyada var, diğerinde yok)
  - Kişi sayısı uyarıları
  - Vaka sayısı tutarsızlıkları
  - Fark raporu PDF olarak dışa aktarılabilir
"""
from datetime import datetime

from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton,
    QComboBox, QFrame, QTableWidget, QTableWidgetItem,
    QHeaderView, QAbstractItemView, QMessageBox, QFileDialog,
    QSplitter, QTextEdit, QTabWidget, QSizePolicy
)
from PySide6.QtCore import Qt, QDate
from PySide6.QtGui import QColor, QBrush, QFont

from ui.styles.components import STYLES as S
from ui.styles.icons import IconRenderer
from core.di import get_dis_alan_service
from core.logger import logger

AY_ADLARI = ["Ocak","Şubat","Mart","Nisan","Mayıs","Haziran",
              "Temmuz","Ağustos","Eylül","Ekim","Kasım","Aralık"]

# Renk kodları
RENK_MUKERRER = QColor("#4A2800")   # Turuncu zemin — mükerrer
RENK_EKSIK    = QColor("#2A0A0A")   # Kırmızı zemin — eksik
RENK_FAZLA    = QColor("#1A2A0A")   # Yeşil zemin — fazla / tek dosyada var
RENK_UYARI    = QColor("#2A2A00")   # Sarı zemin — uyarı

YAZ_MUKERRER  = QColor("#FFB300")
YAZ_EKSIK     = QColor("#EF9A9A")
YAZ_FAZLA     = QColor("#A5D6A7")
YAZ_UYARI     = QColor("#FFE082")
YAZ_NORMAL    = QColor("#E0E0E0")


def _int(v):
    try: return int(v)
    except: return 0

def _float(v):
    try: return float(str(v).replace(",", "."))
    except: return 0.0

def _kisi_key(r):
    tc = str(r.get("TCKimlik", "")).strip()
    return tc if tc else str(r.get("AdSoyad", "")).strip().upper()


class _KartFrame(QFrame):
    def __init__(self, parent=None):
        super().__init__(parent)
        self._lv: QLabel | None = None


class DisAlanPuantajRaporPage(QWidget):
    def __init__(self, db=None, parent=None):
        super().__init__(parent)
        self.setStyleSheet(S["page"])
        self._db  = db
        self._svc = get_dis_alan_service(db) if db else None
        self._rows_cache = []
        self._analiz_sonucu = None
        self._setup_ui()
        self._connect_signals()
        self._load_all_rows()

    # =========================================================
    #  UI
    # =========================================================

    def _setup_ui(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(20, 15, 20, 15)
        root.setSpacing(10)

        # ── Filtre paneli ─────────────────────────────────────
        top = QFrame()
        top.setStyleSheet(S["filter_panel"])
        top.setMaximumHeight(56)
        top_lay = QHBoxLayout(top)
        top_lay.setContentsMargins(12, 6, 12, 6)
        top_lay.setSpacing(10)

        lbl = QLabel("Import Denetim & Karşılaştırma")
        lbl.setStyleSheet("font-size:15px; font-weight:bold; color:#1D75FE;")
        top_lay.addWidget(lbl)
        top_lay.addStretch()

        top_lay.addWidget(QLabel("Anabilim Dalı:"))
        self.cmb_anabilim = QComboBox()
        self.cmb_anabilim.setFixedWidth(220)
        top_lay.addWidget(self.cmb_anabilim)

        top_lay.addWidget(QLabel("Birim:"))
        self.cmb_birim = QComboBox()
        self.cmb_birim.setFixedWidth(200)
        top_lay.addWidget(self.cmb_birim)

        top_lay.addWidget(QLabel("Dönem:"))
        self.cmb_ay = QComboBox()
        self.cmb_ay.addItems(AY_ADLARI)
        self.cmb_ay.setCurrentIndex(QDate.currentDate().month() - 1)
        self.cmb_ay.setFixedWidth(110)
        top_lay.addWidget(self.cmb_ay)

        self.cmb_yil = QComboBox()
        self.cmb_yil.addItems([
            str(y) for y in range(QDate.currentDate().year() - 2,
                                   QDate.currentDate().year() + 2)
        ])
        self.cmb_yil.setCurrentText(str(QDate.currentDate().year()))
        self.cmb_yil.setFixedWidth(100)
        top_lay.addWidget(self.cmb_yil)

        self.btn_analiz = QPushButton("Analiz Et")
        self.btn_analiz.setStyleSheet(S["save_btn"])
        self.btn_analiz.setFixedHeight(36)
        self.btn_analiz.setFixedWidth(100)
        IconRenderer.set_button_icon(self.btn_analiz, "search", color="#FFFFFF")
        top_lay.addWidget(self.btn_analiz)

        root.addWidget(top)

        # ── Özet kartlar ──────────────────────────────────────
        kart_lay = QHBoxLayout()
        kart_lay.setSpacing(8)
        self.kart_dosya    = self._kart("Import Dosyası",  "—", "#457B9D")
        self.kart_toplam   = self._kart("Toplam Kişi",     "—", "#457B9D")
        self.kart_mukerrer = self._kart("Mükerrer",        "—", "#E65100")
        self.kart_eksik    = self._kart("Eksik / Fazla",   "—", "#B71C1C")
        self.kart_uyari    = self._kart("Uyarı",           "—", "#F57F17")
        self.kart_temiz    = self._kart("Temiz Kayıt",     "—", "#1B5E20")
        for k in [self.kart_dosya, self.kart_toplam, self.kart_mukerrer,
                  self.kart_eksik, self.kart_uyari, self.kart_temiz]:
            kart_lay.addWidget(k)
        root.addLayout(kart_lay)

        # ── İçerik: tablo + log splitter ─────────────────────
        splitter = QSplitter(Qt.Orientation.Vertical)

        # Detay tablosu
        detay_frame = QFrame()
        detay_lay = QVBoxLayout(detay_frame)
        detay_lay.setContentsMargins(0, 0, 0, 0)

        self.tablo = QTableWidget()
        kolonlar = [
            ("Durum",         80),
            ("TC Kimlik",    120),
            ("Ad Soyad",     200),
            ("Birim",        140),
            ("Dosya / Tutanak", 220),
            ("Vaka",          60),
            ("Saat",          70),
            ("Sorun",        300),
        ]
        self.tablo.setColumnCount(len(kolonlar))
        self.tablo.setHorizontalHeaderLabels([c[0] for c in kolonlar])
        self.tablo.verticalHeader().setVisible(False)
        self.tablo.setStyleSheet(S["table"])
        self.tablo.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.tablo.setSelectionMode(QAbstractItemView.SelectionMode.NoSelection)
        self.tablo.setAlternatingRowColors(False)

        hdr = self.tablo.horizontalHeader()
        for i, (_, w) in enumerate(kolonlar):
            if kolonlar[i][0] in ("Ad Soyad", "Sorun"):
                hdr.setSectionResizeMode(i, QHeaderView.ResizeMode.Stretch)
            else:
                hdr.setSectionResizeMode(i, QHeaderView.ResizeMode.Fixed)
                self.tablo.setColumnWidth(i, w)

        detay_lay.addWidget(self.tablo)
        splitter.addWidget(detay_frame)

        # Log / özet metin paneli
        self.txt_log = QTextEdit()
        self.txt_log.setReadOnly(True)
        self.txt_log.setMaximumHeight(180)
        self.txt_log.setProperty("bg-role", "panel")
        self.txt_log.setProperty("style-role", "log")
        splitter.addWidget(self.txt_log)
        splitter.setSizes([450, 180])

        root.addWidget(splitter)

        # ── Alt buton çubuğu ──────────────────────────────────
        bot = QFrame()
        bot.setProperty("bg-role", "panel")
        bot.setMaximumHeight(50)
        bot_lay = QHBoxLayout(bot)
        bot_lay.setContentsMargins(12, 6, 12, 6)
        bot_lay.setSpacing(10)

        self.btn_pdf = QPushButton("PDF Rapor Oluştur")
        self.btn_pdf.setProperty("style-role", "danger")
        self.btn_pdf.setFixedHeight(34)
        self.btn_pdf.setEnabled(False)
        IconRenderer.set_button_icon(self.btn_pdf, "download", color="#FFFFFF")

        self.btn_excel = QPushButton("Excel'e Aktar")
        self.btn_excel.setProperty("style-role", "secondary")
        self.btn_excel.setFixedHeight(34)
        self.btn_excel.setEnabled(False)

        bot_lay.addWidget(self.btn_pdf)
        bot_lay.addWidget(self.btn_excel)
        bot_lay.addStretch()

        self.lbl_durum = QLabel("")
        self.lbl_durum.setProperty("style-role", "footer")
        bot_lay.addWidget(self.lbl_durum)

        root.addWidget(bot)

    def _kart(self, baslik, deger, renk="#457B9D"):
        f = _KartFrame()
        f.setProperty("bg-role", "panel")
        f.setProperty("border-role", "accent")
        lay = QVBoxLayout(f)
        lay.setContentsMargins(12, 8, 12, 8)
        lay.setSpacing(2)
        lb = QLabel(baslik)
        lb.setProperty("style-role", "stat-label")
        lv = QLabel(deger)
        lv.setProperty("style-role", "stat-value")
        lay.addWidget(lb)
        lay.addWidget(lv)
        f._lv = lv
        return f

    def _kart_set(self, kart, val):
        if kart._lv is not None:
            kart._lv.setText(str(val))

    # =========================================================
    #  Sinyaller
    # =========================================================

    def _connect_signals(self):
        self.cmb_anabilim.currentIndexChanged.connect(self._on_anabilim_changed)
        self.btn_analiz.clicked.connect(self._analiz)
        self.btn_pdf.clicked.connect(self._pdf_rapor)
        self.btn_excel.clicked.connect(self._export_excel)

    # =========================================================
    #  Veri
    # =========================================================

    def _load_all_rows(self):
        if not self._svc:
            return
        try:
            self._rows_cache = self._svc._r.get("Dis_Alan_Calisma").get_all() or []
        except Exception as e:
            logger.error(f"PuantajRapor._load_all_rows: {e}")
            self._rows_cache = []
        self._update_anabilim_combo()

    def _update_anabilim_combo(self):
        anabilimler = sorted({
            str(r.get("AnaBilimDali","")).strip()
            for r in self._rows_cache if r.get("AnaBilimDali")
        })
        self.cmb_anabilim.blockSignals(True)
        self.cmb_anabilim.clear()
        self.cmb_anabilim.addItem("Tümü")
        self.cmb_anabilim.addItems(anabilimler)
        self.cmb_anabilim.blockSignals(False)
        self._on_anabilim_changed()

    def _on_anabilim_changed(self):
        secili = self.cmb_anabilim.currentText()
        birimler = sorted({
            str(r.get("Birim","")).strip()
            for r in self._rows_cache
            if r.get("Birim") and (secili == "Tümü" or r.get("AnaBilimDali","") == secili)
        })
        self.cmb_birim.blockSignals(True)
        self.cmb_birim.clear()
        self.cmb_birim.addItem("Tümü")
        self.cmb_birim.addItems(birimler)
        self.cmb_birim.blockSignals(False)

    # =========================================================
    #  Analiz
    # =========================================================

    def _analiz(self):
        self._load_all_rows()

        ay    = self.cmb_ay.currentIndex() + 1
        yil   = int(self.cmb_yil.currentText())
        ana   = self.cmb_anabilim.currentText()
        birim = self.cmb_birim.currentText()

        # Filtre
        rows = [
            r for r in self._rows_cache
            if _int(r.get("DonemAy")) == ay and _int(r.get("DonemYil")) == yil
        ]
        if ana != "Tümü":
            rows = [r for r in rows if r.get("AnaBilimDali","") == ana]
        if birim != "Tümü":
            rows = [r for r in rows if r.get("Birim","") == birim]

        if not rows:
            self.txt_log.setPlainText(
                f"⚠  {AY_ADLARI[ay-1]} {yil} dönemine ait kayıt bulunamadı."
            )
            self.lbl_durum.setText("Kayıt yok")
            return

        # Import dosyalarını grupla (TutanakNo = bir import işlemi)
        dosyalar: dict[str, list] = {}
        for r in rows:
            tn = str(r.get("TutanakNo","")).strip()
            if tn not in dosyalar:
                dosyalar[tn] = []
            dosyalar[tn].append(r)

        # Kişi → dosya(lar) haritası
        kisi_dosya: dict[str, set] = {}    # key → {TutanakNo, ...}
        kisi_vaka:  dict[str, dict] = {}   # key → {TutanakNo: vaka_sayisi}

        for tn, kayitlar in dosyalar.items():
            for r in kayitlar:
                k = _kisi_key(r)
                if k not in kisi_dosya:
                    kisi_dosya[k] = set()
                    kisi_vaka[k]  = {}
                kisi_dosya[k].add(tn)
                kisi_vaka[k][tn] = kisi_vaka[k].get(tn, 0) + _int(r.get("VakaSayisi",0))

        dosya_listesi = sorted(dosyalar.keys())
        n_dosya = len(dosya_listesi)

        # Analiz sonuçları
        sonuclar = []  # liste: {row_data, durum, sorun}

        mukerrer_set = set()
        eksik_set    = set()
        uyari_set    = set()
        temiz_set    = set()

        for k, tn_set in kisi_dosya.items():
            # Temsili kayıt al
            temsil = next(
                r for r in rows if _kisi_key(r) == k
            )

            if len(tn_set) > 1:
                # Aynı kişi birden fazla dosyada → mükerrer
                mukerrer_set.add(k)
                sorun = f"Mükerrer: {len(tn_set)} dosyada var"
                # Vaka sayıları farklı mı?
                vakalar = list(kisi_vaka[k].values())
                if len(set(vakalar)) > 1:
                    sorun += f" — Vaka sayısı tutarsız: {vakalar}"
                sonuclar.append({
                    "durum": "MÜKERRER",
                    "kisi_key": k,
                    "temsil": temsil,
                    "tn_set": tn_set,
                    "sorun": sorun,
                    "bg": RENK_MUKERRER, "fg": YAZ_MUKERRER,
                })
            elif n_dosya > 1 and len(tn_set) < n_dosya:
                # Birden fazla dosya var ama bu kişi sadece birinde → eksik
                eksik_set.add(k)
                eksik_dosyalar = [tn[:12]+"…" for tn in dosya_listesi if tn not in tn_set]
                sorun = f"Eksik: {len(eksik_dosyalar)} dosyada yok"
                sonuclar.append({
                    "durum": "EKSİK",
                    "kisi_key": k,
                    "temsil": temsil,
                    "tn_set": tn_set,
                    "sorun": sorun,
                    "bg": RENK_EKSIK, "fg": YAZ_EKSIK,
                })
            else:
                # Temiz
                temiz_set.add(k)
                sonuclar.append({
                    "durum": "TAMAM",
                    "kisi_key": k,
                    "temsil": temsil,
                    "tn_set": tn_set,
                    "sorun": "",
                    "bg": None, "fg": YAZ_NORMAL,
                })

        # Dosyalar arası kişi sayısı farkı kontrolü
        kisi_sayilari = {tn: len(set(_kisi_key(r) for r in kayitlar))
                         for tn, kayitlar in dosyalar.items()}

        if n_dosya > 1:
            sayilar = list(kisi_sayilari.values())
            if max(sayilar) - min(sayilar) > 0:
                uyari_set.add("_kisi_sayisi_farki")
                for s in sonuclar:
                    if s["durum"] == "TAMAM":
                        s["durum"] = "UYARI"
                        s["bg"]    = RENK_UYARI
                        s["fg"]    = YAZ_UYARI
                        s["sorun"] = "Dosyalar arası kişi sayısı eşit değil"
                        uyari_set.add(s["kisi_key"])

        # Sıralama: sorunlular önce
        sira = {"MÜKERRER": 0, "EKSİK": 1, "UYARI": 2, "TAMAM": 3}
        sonuclar.sort(key=lambda x: sira.get(x["durum"], 9))

        self._analiz_sonucu = {
            "ay": ay, "yil": yil, "ana": ana, "birim": birim,
            "dosyalar": dosyalar,
            "dosya_listesi": dosya_listesi,
            "kisi_sayilari": kisi_sayilari,
            "sonuclar": sonuclar,
            "mukerrer": mukerrer_set,
            "eksik": eksik_set,
            "uyari": uyari_set,
            "temiz": temiz_set,
        }

        self._tablo_doldur(sonuclar, dosyalar)
        self._kartlar_guncelle()
        self._log_yaz()

        self.btn_pdf.setEnabled(True)
        self.btn_excel.setEnabled(True)
        self.lbl_durum.setText(
            f"{AY_ADLARI[ay-1]} {yil}  |  "
            f"{n_dosya} dosya  |  "
            f"{len(kisi_dosya)} kişi"
        )

    # =========================================================
    #  Tablo doldur
    # =========================================================

    def _tablo_doldur(self, sonuclar, dosyalar):
        self.tablo.setRowCount(0)
        self.tablo.setRowCount(len(sonuclar))

        for i, s in enumerate(sonuclar):
            self.tablo.setRowHeight(i, 24)
            r   = s["temsil"]
            bg  = s["bg"]
            fg  = s["fg"]

            def _it(text, center=False):
                it = QTableWidgetItem(str(text) if text is not None else "")
                if bg:
                    it.setBackground(QBrush(bg))
                it.setForeground(QBrush(fg))
                if center:
                    it.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                return it

            # Emoji yerine plain text, ikon render UI'da yapılmalı (ör: QLabel/QTableWidgetItem'da ikon desteği varsa eklenebilir)
            # Burada sadece metin gösteriyoruz, ikon için ileride IconRenderer ile ekleme yapılabilir.
            durum_labels = {
                "MÜKERRER": "Mükerrer",
                "EKSİK":    "Eksik",
                "UYARI":    "Uyarı",
                "TAMAM":    "Tamam",
            }
            self.tablo.setItem(i, 0, _it(durum_labels.get(s["durum"], s["durum"]), center=True))
            self.tablo.setItem(i, 1, _it(r.get("TCKimlik",""), center=True))
            self.tablo.setItem(i, 2, _it(r.get("AdSoyad","")))
            self.tablo.setItem(i, 3, _it(r.get("Birim","")))

            # Dosya kısa adı
            tn_str = " | ".join(tn[:16]+"…" for tn in sorted(s["tn_set"]))
            self.tablo.setItem(i, 4, _it(tn_str))

            # Toplam vaka (tüm dosyalardan)
            toplam_vaka = sum(
                _int(rec.get("VakaSayisi",0))
                for tn in s["tn_set"]
                for rec in dosyalar.get(tn,[])
                if _kisi_key(rec) == s["kisi_key"]
            )
            toplam_saat = sum(
                _float(rec.get("HesaplananSaat",0))
                for tn in s["tn_set"]
                for rec in dosyalar.get(tn,[])
                if _kisi_key(rec) == s["kisi_key"]
            )
            self.tablo.setItem(i, 5, _it(str(toplam_vaka), center=True))
            self.tablo.setItem(i, 6, _it(f"{toplam_saat:.2f}", center=True))
            self.tablo.setItem(i, 7, _it(s["sorun"]))

    # =========================================================
    #  Kartlar & log
    # =========================================================

    def _kartlar_guncelle(self):
        a = self._analiz_sonucu
        if not a:
            return
        self._kart_set(self.kart_dosya,    len(a["dosya_listesi"]))
        self._kart_set(self.kart_toplam,   len(set(s["kisi_key"] for s in a["sonuclar"])))
        self._kart_set(self.kart_mukerrer, len(a["mukerrer"]))
        self._kart_set(self.kart_eksik,    len(a["eksik"]))
        self._kart_set(self.kart_uyari,    len(a["uyari"]))
        self._kart_set(self.kart_temiz,    len(a["temiz"]))

    def _log_yaz(self):
        a = self._analiz_sonucu
        if not a:
            return
        ay, yil = a["ay"], a["yil"]
        lines = []
        lines.append(f"{'='*60}")
        lines.append(f"  {AY_ADLARI[ay-1]} {yil} — Import Analizi")
        lines.append(f"  Anabilim Dalı: {a['ana']}  |  Birim: {a['birim']}")
        lines.append(f"{'='*60}")
        lines.append(f"  Import dosyası sayısı : {len(a['dosya_listesi'])}")
        for tn, kayitlar in a["dosyalar"].items():
            kisi_n = len(set(_kisi_key(r) for r in kayitlar))
            lines.append(f"  • Tutanak {tn[:20]}…  →  {kisi_n} kişi")
        lines.append(f"")
        lines.append(f"  Mükerrer kişi : {len(a['mukerrer'])}")
        lines.append(f"  Eksik kişi    : {len(a['eksik'])}")
        lines.append(f"  Uyarı         : {len(a['uyari'])}")
        lines.append(f"  Temiz kayıt   : {len(a['temiz'])}")


        if a["mukerrer"]:
            lines.append("")
            lines.append("  [MÜKERRER] KAYITLAR:")
            for s in a["sonuclar"]:
                if s["durum"] == "MÜKERRER":
                    r = s["temsil"]
                    lines.append(f"    - {r.get('TCKimlik','')} {r.get('AdSoyad','')}: {s['sorun']}")

        if a["eksik"]:
            lines.append("")
            lines.append("  [EKSİK] KAYITLAR:")
            for s in a["sonuclar"]:
                if s["durum"] == "EKSİK":
                    r = s["temsil"]
                    lines.append(f"    - {r.get('TCKimlik','')} {r.get('AdSoyad','')}: {s['sorun']}")

        self.txt_log.setPlainText("\n".join(lines))

    # =========================================================
    #  PDF Rapor
    # =========================================================

    def _pdf_rapor(self):
        if not self._analiz_sonucu:
            return

        path, _ = QFileDialog.getSaveFileName(
            self, "PDF Kaydet",
            f"dis_alan_denetim_{self._analiz_sonucu['ay']}_{self._analiz_sonucu['yil']}.pdf",
            "PDF (*.pdf)"
        )
        if not path:
            return

        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import mm
            from reportlab.lib import colors
            from reportlab.platypus import (
                SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
            )
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont
            try:
                pdfmetrics.registerFont(TTFont("Arial", "arial.ttf"))
                pdfmetrics.registerFont(TTFont("ArialBold", "arialbd.ttf"))
                fn, fnb = "Arial", "ArialBold"
            except Exception:
                fn, fnb = "Helvetica", "Helvetica-Bold"
        except ImportError:
            QMessageBox.critical(
                self, "Eksik Paket",
                "reportlab kurulu değil.\nKurulum: pip install reportlab"
            )
            return

        a = self._analiz_sonucu
        if not a:
            return
        ay = a["ay"]
        yil = a["yil"]
        doc = SimpleDocTemplate(
            path, pagesize=A4,
            leftMargin=18*mm, rightMargin=18*mm,
            topMargin=18*mm, bottomMargin=18*mm
        )

        normal = ParagraphStyle("N", fontName=fn,  fontSize=9,  leading=13)
        bold   = ParagraphStyle("B", fontName=fnb, fontSize=9,  leading=13)
        h1     = ParagraphStyle("H1",fontName=fnb, fontSize=14, leading=18, spaceAfter=4)
        h2     = ParagraphStyle("H2",fontName=fnb, fontSize=11, leading=14, spaceAfter=2,
                                textColor=colors.HexColor("#1D3557"))
        small  = ParagraphStyle("S", fontName=fn,  fontSize=8,  leading=11,
                                textColor=colors.HexColor("#555555"))
        err    = ParagraphStyle("E", fontName=fnb, fontSize=9,  leading=13,
                                textColor=colors.HexColor("#C62828"))
        warn   = ParagraphStyle("W", fontName=fnb, fontSize=9,  leading=13,
                                textColor=colors.HexColor("#E65100"))
        ok     = ParagraphStyle("OK",fontName=fn,  fontSize=9,  leading=13,
                                textColor=colors.HexColor("#1B5E20"))

        elements = []

        # Başlık
        elements.append(Paragraph("DIŞ ALAN RADYASYON — IMPORT DENETİM RAPORU", h1))
        elements.append(Paragraph(
            f"Dönem: {AY_ADLARI[ay-1]} {yil}  |  "
            f"Anabilim Dalı: {a['ana']}  |  Birim: {a['birim']}  |  "
            f"Rapor: {datetime.now().strftime('%d.%m.%Y %H:%M')}",
            small
        ))
        elements.append(Spacer(1, 6*mm))
        elements.append(HRFlowable(width="100%", thickness=0.5,
                                   color=colors.HexColor("#1D3557")))
        elements.append(Spacer(1, 4*mm))

        # Özet kutusu
        elements.append(Paragraph("Özet", h2))
        ozet_data = [
            ["Import Dosyası Sayısı", str(len(a["dosya_listesi"]))],
            ["Toplam Kişi",           str(len({s["kisi_key"] for s in a["sonuclar"]}))],
            ["Mükerrer Kayıt",        str(len(a["mukerrer"]))],
            ["Eksik Kayıt",           str(len(a["eksik"]))],
            ["Uyarı",                 str(len(a["uyari"]))],
            ["Temiz Kayıt",           str(len(a["temiz"]))],
        ]
        ozet_tbl = Table(ozet_data, colWidths=[80*mm, 40*mm])
        ozet_tbl.setStyle(TableStyle([
            ("FONTNAME",  (0,0), (-1,-1), fn),
            ("FONTSIZE",  (0,0), (-1,-1), 9),
            ("FONTNAME",  (0,0), (0,-1),  fnb),
            ("GRID",      (0,0), (-1,-1), 0.3, colors.HexColor("#CCCCCC")),
            ("BACKGROUND",(0,0), (-1,0),  colors.HexColor("#1D3557")),
            ("TEXTCOLOR", (0,0), (-1,0),  colors.white),
            ("ROWBACKGROUNDS", (0,1), (-1,-1),
             [colors.HexColor("#F5F5F5"), colors.white]),
        ]))
        elements.append(ozet_tbl)
        elements.append(Spacer(1, 5*mm))

        # Import dosyaları detayı
        elements.append(Paragraph("Import Dosyaları", h2))
        for tn, kayitlar in a["dosyalar"].items():
            kisi_n = len({_kisi_key(r) for r in kayitlar})
            vaka_n = sum(_int(r.get("VakaSayisi",0)) for r in kayitlar)
            elements.append(Paragraph(
                f"Tutanak: {tn[:36]}…  |  {kisi_n} kişi  |  {vaka_n} vaka",
                small
            ))
        elements.append(Spacer(1, 5*mm))

        # Sorun tablosu
        sorunlu = [s for s in a["sonuclar"] if s["durum"] != "TAMAM"]
        if sorunlu:
            elements.append(Paragraph("Sorunlu Kayıtlar", h2))
            tbl_data = [["Durum","TC Kimlik","Ad Soyad","Vaka","Sorun"]]
            for s in sorunlu:
                r   = s["temsil"]
                vaka = sum(
                    _int(rec.get("VakaSayisi",0))
                    for tn in s["tn_set"]
                    for rec in a["dosyalar"].get(tn,[])
                    if _kisi_key(rec) == s["kisi_key"]
                )
                tbl_data.append([
                    s["durum"],
                    str(r.get("TCKimlik","")),
                    str(r.get("AdSoyad","")),
                    str(vaka),
                    s["sorun"][:60],
                ])

            sorun_tbl = Table(
                tbl_data,
                colWidths=[22*mm, 28*mm, 50*mm, 14*mm, None]
            )
            sorun_tbl.setStyle(TableStyle([
                ("FONTNAME",  (0,0), (-1,-1), fn),
                ("FONTSIZE",  (0,0), (-1,-1), 8),
                ("FONTNAME",  (0,0), (-1,0),  fnb),
                ("BACKGROUND",(0,0), (-1,0),  colors.HexColor("#37474F")),
                ("TEXTCOLOR", (0,0), (-1,0),  colors.white),
                ("GRID",      (0,0), (-1,-1), 0.3, colors.HexColor("#CCCCCC")),
                ("ROWBACKGROUNDS", (0,1), (-1,-1),
                 [colors.HexColor("#FFF3E0"), colors.white]),
                ("VALIGN",    (0,0), (-1,-1), "MIDDLE"),
            ]))
            elements.append(sorun_tbl)
            elements.append(Spacer(1, 5*mm))

        # Tüm kayıtlar tablosu
        elements.append(Paragraph("Tüm Kayıtlar", h2))
        tum_data = [["TC Kimlik","Ad Soyad","Birim","Vaka","Saat","Durum"]]
        for s in a["sonuclar"]:
            r = s["temsil"]
            vaka = sum(
                _int(rec.get("VakaSayisi",0))
                for tn in s["tn_set"]
                for rec in a["dosyalar"].get(tn,[])
                if _kisi_key(rec) == s["kisi_key"]
            )
            saat = sum(
                _float(rec.get("HesaplananSaat",0))
                for tn in s["tn_set"]
                for rec in a["dosyalar"].get(tn,[])
                if _kisi_key(rec) == s["kisi_key"]
            )
            tum_data.append([
                str(r.get("TCKimlik","")),
                str(r.get("AdSoyad","")),
                str(r.get("Birim","")),
                str(vaka),
                f"{saat:.2f}",
                s["durum"],
            ])

        tum_tbl = Table(
            tum_data,
            colWidths=[28*mm, 50*mm, 35*mm, 15*mm, 18*mm, 22*mm]
        )
        row_bg = []
        for i, s in enumerate([None] + a["sonuclar"]):
            if s is None:
                continue
            row_i = i
            if s["durum"] == "MÜKERRER":
                row_bg.append(("BACKGROUND", (0,row_i), (-1,row_i), colors.HexColor("#FFF3E0")))
            elif s["durum"] == "EKSİK":
                row_bg.append(("BACKGROUND", (0,row_i), (-1,row_i), colors.HexColor("#FFEBEE")))

        tum_tbl.setStyle(TableStyle([
            ("FONTNAME",  (0,0), (-1,-1), fn),
            ("FONTSIZE",  (0,0), (-1,-1), 8),
            ("FONTNAME",  (0,0), (-1,0),  fnb),
            ("BACKGROUND",(0,0), (-1,0),  colors.HexColor("#1D3557")),
            ("TEXTCOLOR", (0,0), (-1,0),  colors.white),
            ("GRID",      (0,0), (-1,-1), 0.3, colors.HexColor("#CCCCCC")),
            ("ROWBACKGROUNDS", (0,1), (-1,-1),
             [colors.HexColor("#F5F5F5"), colors.white]),
            ("VALIGN",    (0,0), (-1,-1), "MIDDLE"),
        ] + row_bg))
        elements.append(tum_tbl)

        try:
            doc.build(elements)
            QMessageBox.information(self, "PDF Oluşturuldu",
                                    f"Rapor kaydedildi:\n{path}")
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"PDF oluşturulamadı:\n{e}")

    # =========================================================
    #  Excel'e aktar
    # =========================================================

    def _export_excel(self):
        if not self._analiz_sonucu:
            return

        a  = self._analiz_sonucu
        ay = a["ay"]; yil = a["yil"]
        path, _ = QFileDialog.getSaveFileName(
            self, "Excel Kaydet",
            f"dis_alan_denetim_{ay}_{yil}.xlsx",
            "Excel (*.xlsx)"
        )
        if not path:
            return

        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            QMessageBox.critical(self, "Eksik Paket",
                                 "openpyxl kurulu değil.\npip install openpyxl")
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        if ws is None:
            ws = wb.create_sheet()
        ws.title = f"Denetim {AY_ADLARI[ay-1]} {yil}"

        renkler = {
            "MÜKERRER": "FFE0B2",
            "EKSİK":    "FFCDD2",
            "UYARI":    "FFF9C4",
            "TAMAM":    "F1F8E9",
        }

        basliklar = ["Durum","TC Kimlik","Ad Soyad","Birim","Tutanak No","Vaka","Saat","Sorun"]
        ws.append(basliklar)
        for cell in ws[1]:
            cell.font  = Font(bold=True, color="FFFFFF")
            cell.fill  = PatternFill("solid", fgColor="1D3557")
            cell.alignment = Alignment(horizontal="center")

        for s in a["sonuclar"]:
            r = s["temsil"]
            vaka = sum(
                _int(rec.get("VakaSayisi",0))
                for tn in s["tn_set"]
                for rec in a["dosyalar"].get(tn,[])
                if _kisi_key(rec) == s["kisi_key"]
            )
            saat = sum(
                _float(rec.get("HesaplananSaat",0))
                for tn in s["tn_set"]
                for rec in a["dosyalar"].get(tn,[])
                if _kisi_key(rec) == s["kisi_key"]
            )
            tn_str = ", ".join(sorted(s["tn_set"]))
            row = [
                s["durum"],
                r.get("TCKimlik",""),
                r.get("AdSoyad",""),
                r.get("Birim",""),
                tn_str,
                vaka,
                round(saat, 2),
                s["sorun"],
            ]
            ws.append(row)
            bg = renkler.get(s["durum"], "FFFFFF")
            for cell in ws[ws.max_row]:
                cell.fill = PatternFill("solid", fgColor=bg)

        from openpyxl.utils import get_column_letter
        for col_idx, col in enumerate(ws.iter_cols(1, ws.max_column), start=1):
            maxlen = max(len(str(c.value or "")) for c in col)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(maxlen + 3, 50)

        try:
            wb.save(path)
            QMessageBox.information(self, "Kaydedildi", f"Excel kaydedildi:\n{path}")
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Excel kaydedilemedi:\n{e}")
