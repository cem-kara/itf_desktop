# ui/pages/fhsz/dis_alan_kurulum_page.py
# -*- coding: utf-8 -*-
"""
Dış Alan Birim Kurulum Sihirbazı

Kullanıcı yalnızca gerçek operasyonel verileri girer:
  - Hangi birim
  - Hangi yılın verisi
  - Kaç ameliyat/işlem yapıldı
  - Kaçında C-kollu kullanıldı
  - C-kollu ne kadar süre kullanıldı
  - Kaç personel var

Sistem otomatik olarak:
  - C-kollu oranını hesaplar
  - Katsayıyı hesaplar (saat/vaka)
  - Katsayı protokolünü oluşturur
"""
from datetime import date

from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton,
    QFrame, QFormLayout, QSpinBox, QComboBox,
    QScrollArea
)
from PySide6.QtCore import Qt
from PySide6.QtGui import QFont

from core.hata_yonetici import bilgi_goster, hata_goster, uyari_goster
from core.logger import logger
from core.di import get_dis_alan_katsayi_service


# ─────────────────────────────────────────────────────────────
#  Yardımcı widget: bilgi kartı
# ─────────────────────────────────────────────────────────────

class _InfoKart(QFrame):
    def __init__(self, baslik: str, deger: str = "", aciklama: str = "", vurgu=False):
        super().__init__()
        renk = "#1D75FE" if vurgu else "#457B9D"
        self.setStyleSheet(
            f"QFrame {{ background:#1A2535; border:1px solid {renk}; "
            f"border-radius:8px; padding:0px; }}"
        )
        lay = QVBoxLayout(self)
        lay.setContentsMargins(14, 10, 14, 10)
        lay.setSpacing(2)

        lbl_baslik = QLabel(baslik)
        lbl_baslik.setProperty("color-role", "primary")
        lay.addWidget(lbl_baslik)

        self.lbl_deger = QLabel(deger)
        font = QFont()
        font.setPointSize(16 if vurgu else 13)
        font.setBold(vurgu)
        self.lbl_deger.setFont(font)
        self.lbl_deger.setStyleSheet(
            f"color:{'#FFD600' if vurgu else '#E0E0E0'}; border:none;"
        )
        lay.addWidget(self.lbl_deger)

        if aciklama:
            lbl_ac = QLabel(aciklama)
            lbl_ac.setProperty("color-role", "primary")
            lbl_ac.setWordWrap(True)
            lay.addWidget(lbl_ac)

    def set_deger(self, deger: str):
        self.lbl_deger.setText(deger)


# ─────────────────────────────────────────────────────────────
#  Ana sayfa
# ─────────────────────────────────────────────────────────────

class DisAlanKurulumPage(QWidget):
    def __init__(self, db=None, parent=None):
        super().__init__(parent)
        self.setProperty("bg-role", "page")
        self._db = db
        self._kat_svc = get_dis_alan_katsayi_service(db) if db else None
        self._setup_ui()
        self._connect_signals()

    # =========================================================
    #  UI
    # =========================================================

    def _setup_ui(self):
        # Kaydırılabilir alan
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.Shape.NoFrame)
        scroll.setProperty("bg-role", "panel")

        inner = QWidget()
        inner.setProperty("bg-role", "panel")
        main = QVBoxLayout(inner)
        main.setContentsMargins(30, 20, 30, 30)
        main.setSpacing(20)

        scroll.setWidget(inner)
        outer = QVBoxLayout(self)
        outer.setContentsMargins(0, 0, 0, 0)
        outer.addWidget(scroll)

        # ── Üst bilgi ─────────────────────────────────────────
        baslik = QLabel("Birim Kurulum Sihirbazı")
        baslik.setStyleSheet(
            "font-size:18px; font-weight:bold; color:#1D75FE;"
        )
        main.addWidget(baslik)

        aciklama = QLabel(
            "Biriminizin geçen yılki operasyonel verilerini girin. "
            "Sistem katsayıyı otomatik hesaplar ve önümüzdeki yıl için "
            "tüm ayarları tek tıkla kaydeder."
        )
        aciklama.setProperty("color-role", "primary")
        aciklama.setWordWrap(True)
        main.addWidget(aciklama)

        # ── Bölüm 1: Birim ve Yıl ─────────────────────────────
        main.addWidget(self._bolum_baslik("1  —  Birim ve Protokol Yılı"))

        form1 = QFormLayout()
        form1.setSpacing(10)
        form1.setLabelAlignment(Qt.AlignmentFlag.AlignRight)
        form1.setContentsMargins(0, 0, 0, 0)

        self.cmb_anabilim = QComboBox()
        self.cmb_anabilim.setEditable(True)
        self._sabitler_doldur(self.cmb_anabilim, "AnaBilimDali")
        form1.addRow("Anabilim Dalı:", self.cmb_anabilim)

        self.cmb_birim = QComboBox()
        self.cmb_birim.setEditable(True)
        self._sabitler_doldur(self.cmb_birim, "Birim")
        form1.addRow("Birim:", self.cmb_birim)

        self.spn_veri_yil = QSpinBox()
        self.spn_veri_yil.setRange(2020, date.today().year)
        self.spn_veri_yil.setValue(date.today().year)
        self.spn_veri_yil.setToolTip(
            "Hangi yılın verilerini giriyorsunuz?\n"
            "Genellikle geçen yıl — bu veriden önümüzdeki yıl için katsayı hesaplanır."
        )
        form1.addRow("Protokol geçerlilik yılı:", self.spn_veri_yil)

        main.addLayout(form1)

        # ── Bölüm 2: İşlem verileri ───────────────────────────
        main.addWidget(self._bolum_baslik("2  —  İşlem Verileri"))

        form2 = QFormLayout()
        form2.setSpacing(10)
        form2.setLabelAlignment(Qt.AlignmentFlag.AlignRight)
        form2.setContentsMargins(0, 0, 0, 0)

        self.spn_ckollu_gun = QSpinBox()
        self.spn_ckollu_gun.setRange(0, 999)
        self.spn_ckollu_gun.setValue(14)
        self.spn_ckollu_gun.setSuffix("  işlem/gün")
        self.spn_ckollu_gun.setToolTip(
            "Tipik bir iş gününde C-kollu (floroskopi) kullanılan işlem sayısı.\n"
            "Örnek: Günde 22 ameliyattan 14'ünde C-kollu kullanılıyorsa → 14"
        )
        form2.addRow("Günlük C-kollu kullanan işlem:", self.spn_ckollu_gun)

        self.spn_toplam_gun = QSpinBox()
        self.spn_toplam_gun.setRange(1, 999)
        self.spn_toplam_gun.setValue(22)
        self.spn_toplam_gun.setSuffix("  işlem/gün")
        self.spn_toplam_gun.setToolTip(
            "Tipik bir iş gününde toplam kaç işlem yapılıyor.\n"
            "C-kollu oran = C-kollu işlem / Toplam işlem"
        )
        form2.addRow("Günlük toplam işlem:", self.spn_toplam_gun)

        self.spn_sure = QSpinBox()
        self.spn_sure.setRange(1, 480)
        self.spn_sure.setValue(20)
        self.spn_sure.setSuffix("  dakika")
        self.spn_sure.setToolTip(
            "Her C-kollu kullanımında radyasyona maruz kalınan ortalama süre.\n"
            "Ameliyatın tamamı değil — sadece C-kollu açık olan süre.\n"
            "Örnek: 2 saatlik ameliyatta C-kollu 20 dk kullanılıyorsa → 20"
        )
        form2.addRow("C-kollu kullanım süresi (dk):", self.spn_sure)

        main.addLayout(form2)

        # ── Bölüm 3: Canlı hesap ──────────────────────────────
        main.addWidget(self._bolum_baslik("3  —  Sistem Hesabı (Otomatik)"))

        kart_grid = QHBoxLayout()
        kart_grid.setSpacing(10)

        self.kart_oran    = _InfoKart("C-kollu oran", "—", "C-kollu işlem / Toplam işlem")
        self.kart_katsayi = _InfoKart("Katsayı (saat/vaka)", "—",
                                      "C-kollu süre × oran / 60", vurgu=True)

        for k in [self.kart_oran, self.kart_katsayi]:
            kart_grid.addWidget(k)

        main.addLayout(kart_grid)

        # ── Bölüm 4: Özet ve Kaydet ───────────────────────────
        main.addWidget(self._bolum_baslik("4  —  Kaydet"))

        ozet_frame = QFrame()
        ozet_frame.setStyleSheet(
            "QFrame { background:#0D1A2A; border:1px solid #1D75FE; "
            "border-radius:10px; }"
        )
        ozet_lay = QVBoxLayout(ozet_frame)
        ozet_lay.setContentsMargins(16, 14, 16, 14)
        ozet_lay.setSpacing(6)

        self.lbl_ozet = QLabel("Verileri girdikten sonra özet burada görünecek.")
        self.lbl_ozet.setProperty("color-role", "primary")
        self.lbl_ozet.setWordWrap(True)
        ozet_lay.addWidget(self.lbl_ozet)

        self.lbl_uyari = QLabel("")
        self.lbl_uyari.setProperty("color-role", "primary")
        self.lbl_uyari.setWordWrap(True)
        self.lbl_uyari.setVisible(False)
        ozet_lay.addWidget(self.lbl_uyari)

        main.addWidget(ozet_frame)

        # Kaydet butonu
        btn_lay = QHBoxLayout()
        btn_lay.addStretch()

        self.btn_hesapla = QPushButton("Hesapla ve Önizle")
        self.btn_hesapla.setProperty("style-role", "secondary")
        self.btn_hesapla.setFixedHeight(40)
        self.btn_hesapla.setFixedWidth(160)

        self.btn_sablon = QPushButton("📥  Şablon İndir")
        self.btn_sablon.setStyleSheet(
            "QPushButton { background:#1565C0; color:#fff; border-radius:8px; "
            "font-size:12px; font-weight:bold; padding:0 18px; }"
            "QPushButton:hover { background:#0D47A1; }"
            "QPushButton:disabled { background:#1A1A2A; color:#556; }"
        )
        self.btn_sablon.setFixedHeight(40)
        self.btn_sablon.setEnabled(False)
        self.btn_sablon.setToolTip(
            "Birim bilgileri dolu, kullanıma hazır Excel şablonu indir.\n"
            "Önce 'Kaydet ve Uygula' ile birimi kaydedin."
        )

        self.btn_kaydet = QPushButton("✓  Kaydet ve Uygula")
        self.btn_kaydet.setStyleSheet(
            "QPushButton { background:#1D75FE; color:#fff; border-radius:8px; "
            "font-size:13px; font-weight:bold; padding:0 24px; }"
            "QPushButton:hover { background:#1558C0; }"
            "QPushButton:disabled { background:#2A3A4A; color:#556; }"
        )
        self.btn_kaydet.setFixedHeight(40)
        self.btn_kaydet.setEnabled(False)

        btn_lay.addWidget(self.btn_hesapla)
        btn_lay.addWidget(self.btn_sablon)
        btn_lay.addWidget(self.btn_kaydet)
        main.addLayout(btn_lay)

        main.addStretch()

    # =========================================================
    #  Bölüm başlık yardımcısı
    # =========================================================

    def _bolum_baslik(self, metin: str) -> QLabel:
        lbl = QLabel(metin)
        lbl.setStyleSheet(
            "font-size:12px; font-weight:bold; color:#457B9D; "
            "padding:6px 0 2px 0; border-bottom:1px solid #1A2535;"
        )
        return lbl

    # =========================================================
    #  Sabitler doldurucu
    # =========================================================

    def _sabitler_doldur(self, cmb: QComboBox, kod: str):
        cmb.addItem("")
        if not self._db:
            return
        try:
            from core.di import get_registry
            rows = get_registry(self._db).get("Sabitler").get_all() or []
            for r in sorted(rows, key=lambda x: x.get("MenuEleman", "")):
                if str(r.get("Kod", "")).strip() == kod:
                    val = str(r.get("MenuEleman", "")).strip()
                    if val:
                        cmb.addItem(val)
        except Exception as e:
            logger.error(f"DisAlanKurulumPage._sabitler_doldur({kod}): {e}")

    # =========================================================
    #  Sinyaller
    # =========================================================

    def _connect_signals(self):
        for w in [self.spn_ckollu_gun, self.spn_toplam_gun, self.spn_sure]:
            w.valueChanged.connect(self._canli_hesapla)

        self.btn_hesapla.clicked.connect(self._onizle)
        self.btn_sablon.clicked.connect(self._sablon_indir)
        self.btn_kaydet.clicked.connect(self._kaydet)

    # =========================================================
    #  Canlı hesaplama (kartları günceller)
    # =========================================================

    def _canli_hesapla(self):
        gun_ckollu = self.spn_ckollu_gun.value()
        gun_toplam = self.spn_toplam_gun.value()
        sure       = self.spn_sure.value()

        if gun_toplam == 0:
            return

        oran    = gun_ckollu / gun_toplam
        katsayi = (sure * oran) / 60.0

        self.kart_oran.set_deger(f"{oran:.3f}  ({gun_ckollu}/{gun_toplam})")
        self.kart_katsayi.set_deger(f"{katsayi:.4f}")
        self.btn_kaydet.setEnabled(False)

    # =========================================================
    #  Önizle
    # =========================================================

    def _onizle(self):
        ana   = self.cmb_anabilim.currentText().strip()
        birim = self.cmb_birim.currentText().strip()
        yil   = self.spn_veri_yil.value()

        if not ana:
            uyari_goster(self, "Anabilim Dalı boş olamaz.")
            return
        if not birim:
            uyari_goster(self, "Birim boş olamaz.")
            return

        gun_ckollu = self.spn_ckollu_gun.value()
        gun_toplam = self.spn_toplam_gun.value()
        sure       = self.spn_sure.value()

        if gun_toplam == 0:
            uyari_goster(self, "Günlük toplam işlem 0 olamaz.")
            return

        oran    = gun_ckollu / gun_toplam
        katsayi = (sure * oran) / 60.0

        # Mevcut protokol kontrolü
        uyari_metni = ""
        if self._kat_svc:
            try:
                mevcut = self._kat_svc.get_aktif_katsayi(ana, birim)
                if mevcut:
                    uyari_metni = (
                        f"⚠  Bu birim için zaten aktif bir protokol var "
                        f"(Katsayı: {mevcut.get('Katsayi')}, "
                        f"Başlangıç: {mevcut.get('GecerlilikBaslangic')}). "
                        f"Önce Katsayı Protokolleri ekranından mevcut protokolü "
                        f"pasife alın, sonra burada kaydedin."
                    )
            except Exception:
                pass

        ozet = (
            f"<b>Birim:</b> {ana} / {birim}<br>"
            f"<b>Protokol geçerlilik yılı:</b> {yil}<br><br>"
            f"<b>Oluşturulacak Katsayı Protokolü</b><br>"
            f"  • Katsayı: <b>{katsayi:.4f}</b> saat/vaka<br>"
            f"  • Geçerlilik: {yil}-01-01<br>"
            f"  • Formül: {sure} dk × {oran:.3f} C-kollu oran / 60"
        )

        self.lbl_ozet.setText(ozet)
        self.lbl_uyari.setText(uyari_metni)
        self.lbl_uyari.setVisible(bool(uyari_metni))
        self.btn_kaydet.setEnabled(True)

        self._onizle_veri = {
            "ana": ana, "birim": birim, "yil": yil,
            "oran": oran, "katsayi": katsayi, "sure": sure,
            "gun_ckollu": gun_ckollu, "gun_toplam": gun_toplam,
        }

    # =========================================================
    #  Kaydet
    # =========================================================

    def _kaydet(self):
        if not hasattr(self, "_onizle_veri"):
            uyari_goster(self, "Lütfen önce 'Hesapla ve Önizle' butonuna tıklayın.")
            return
        if not self._db:
            hata_goster(self, "Veritabanı bağlantısı yok.")
            return

        v = self._onizle_veri

        try:
            kat_svc = self._kat_svc
            protokol = {
                "AnaBilimDali":        v["ana"],
                "Birim":               v["birim"],
                "GecerlilikBaslangic": f"{v['yil']}-01-01",
                "Katsayi":             round(v["katsayi"], 4),
                "OrtSureDk":           v["sure"],
                "AlanTipAciklama":     f"{v['birim']} — Kurulum sihirbazı",
                "AciklamaFormul":      (
                    f"{v['sure']} dk × {v['oran']:.3f} C-kollu / 60 = {v['katsayi']:.4f}"
                ),
                "ProtokolRef":         (
                    f"Sihirbaz | {v['gun_ckollu']}/{v['gun_toplam']} C-kollu/gün, "
                    f"{v['sure']} dk"
                ),
                "Aktif":               1,
                "KaydedenKullanici":   "Kurulum Sihirbazı",
            }
            ok = kat_svc.protokol_ekle(protokol)
        except Exception as e:
            hata_goster(self, str(e))
            return

        if ok:
            bilgi_goster(
                self, "Kurulum Tamamlandı",
                f"{v['ana']} / {v['birim']} için katsayı protokolü oluşturuldu.\n\n"
                f"Katsayı : {v['katsayi']:.4f} saat/vaka\n"
                f"Geçerlilik : {v['yil']}-01-01 tarihinden itibaren\n\n"
                f"Artık bu birimden Excel import yapabilirsiniz."
            )
            self.btn_kaydet.setEnabled(False)
            self.btn_sablon.setEnabled(True)
            del self._onizle_veri
        else:
            uyari_goster(
                self, "Eklenemedi",
                f"{v['yil']}-01-01 başlangıç tarihli bir protokol zaten mevcut.\n"
                "Önce Katsayı Protokolleri ekranından mevcut protokolü pasife alın."
            )

    def _sablon_indir(self):
        """
        Birim bilgileri (B3/B4) ve dönem yılı önceden dolu
        Excel şablonu oluşturup kaydeder.
        """
        from PySide6.QtWidgets import QFileDialog

        ana   = self.cmb_anabilim.currentText().strip()
        birim = self.cmb_birim.currentText().strip()
        yil   = self.spn_veri_yil.value()

        dosya_adi = (
            f"Dis_Alan_Sablonu_{birim.replace(' ','_')[:20]}_{yil}.xlsx"
        )
        path, _ = QFileDialog.getSaveFileName(
            self, "Şablonu Kaydet", dosya_adi, "Excel (*.xlsx)"
        )
        if not path:
            return

        try:
            import openpyxl
            from openpyxl.styles import (
                Font, PatternFill, Alignment, Border, Side
            )

            wb = openpyxl.Workbook()
            ws = wb.active
            if ws is None:
                # This should not happen with a new workbook, but as a safeguard
                ws = wb.create_sheet()
            ws.title = "Bildirim"

            mavi   = "1D3557"
            sari   = "FFD600"
            gri_bg = "F0F4F8"

            veri_f = Font(name="Calibri", size=10)
            ince_border = Border(
                left=Side(style="thin", color="CCCCCC"),
                right=Side(style="thin", color="CCCCCC"),
                top=Side(style="thin", color="CCCCCC"),
                bottom=Side(style="thin", color="CCCCCC"),
            )

            # Kolon genişlikleri
            ws.column_dimensions["A"].width = 16
            ws.column_dimensions["B"].width = 22
            ws.column_dimensions["C"].width = 22
            ws.column_dimensions["D"].width = 14
            ws.row_dimensions[1].height = 36
            ws.row_dimensions[3].height = 24
            ws.row_dimensions[4].height = 24
            ws.row_dimensions[6].height = 24

            # Satır 1 — Başlık
            ws.merge_cells("A1:D1")
            c = ws["A1"]
            c.value = "DIŞ ALAN RADYASYON ÇALIŞMA BİLDİRİM ŞABLONU"
            c.font  = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
            c.fill  = PatternFill("solid", fgColor=mavi)
            c.alignment = Alignment(horizontal="center", vertical="center")

            # Satır 3 — Anabilim Dalı + Dönem Ay
            ws["A3"] = "Anabilim Dalı"
            ws["A3"].font = Font(bold=True, size=10)
            ws["B3"] = ana
            ws["B3"].font = Font(size=10)
            ws["C3"] = "Dönem Ay"
            ws["C3"].font = Font(bold=True, size=10, color="FFFFFF")
            ws["C3"].fill = PatternFill("solid", fgColor=mavi)
            ws["C3"].alignment = Alignment(horizontal="center")
            ws["D3"].font = Font(bold=True, size=11, color=sari)
            ws["D3"].fill = PatternFill("solid", fgColor="1A2535")
            ws["D3"].alignment = Alignment(horizontal="center")

            # Satır 4 — Birim + Dönem Yıl
            ws["A4"] = "Birim"
            ws["A4"].font = Font(bold=True, size=10)
            ws["B4"] = birim
            ws["B4"].font = Font(size=10)
            ws["C4"] = "Dönem Yıl"
            ws["C4"].font = Font(bold=True, size=10, color="FFFFFF")
            ws["C4"].fill = PatternFill("solid", fgColor=mavi)
            ws["C4"].alignment = Alignment(horizontal="center")
            ws["D4"] = yil
            ws["D4"].font = Font(bold=True, size=11)
            ws["D4"].alignment = Alignment(horizontal="center")

            # Satır 6 — Sütun başlıkları
            for col_i, bas in enumerate(
                ["TC Kimlik No", "Ad Soyad *", "Çalışılan Alan *", "Vaka Sayısı *"], 1
            ):
                cell = ws.cell(row=6, column=col_i, value=bas)
                cell.font  = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
                cell.fill  = PatternFill("solid", fgColor=mavi)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = ince_border

            # Satır 7-36 — 30 boş veri satırı
            for row_i in range(7, 37):
                ws.row_dimensions[row_i].height = 18
                for col_i in range(1, 5):
                    cell = ws.cell(row=row_i, column=col_i)
                    cell.font   = veri_f
                    cell.border = ince_border
                    if row_i % 2 == 0:
                        cell.fill = PatternFill("solid", fgColor=gri_bg)

            # Kılavuz sayfası
            ws2 = wb.create_sheet("Kılavuz")
            ws2.column_dimensions["A"].width = 60
            kilavuz = [
                "DOLDURMA KILAVUZU", "",
                "• TC Kimlik No: 11 haneli TC kimlik numarası (opsiyonel)",
                "• Ad Soyad *: Zorunlu alan — tam isim yazın",
                f"• Çalışılan Alan *: Bu birim için geçerli alan: {birim}",
                "• Vaka Sayısı *: Bu dönemde yapılan toplam işlem sayısı",
                "", "NOTLAR:",
                "• D3 hücresine ay adı veya numarası yazın (örn: Ocak veya 1)",
                "• D4 hücresine 4 haneli yılı yazın (örn: 2026)",
                "• * işaretli alanlar zorunludur",
                "• Katsayı ve saat hesabı sistem tarafından otomatik yapılır",
                "", f"Bu şablon {ana} / {birim} birimi için hazırlanmıştır.",
                f"Protokol yılı: {yil}",
            ]
            for i, sat in enumerate(kilavuz, 1):
                cell = ws2.cell(row=i, column=1, value=sat)
                if i == 1:
                    cell.font = Font(bold=True, size=12, color="FFFFFF")
                    cell.fill = PatternFill("solid", fgColor=mavi)

            wb.save(path)
            bilgi_goster(
                self, "Şablon Hazır",
                f"Şablon kaydedildi:\n{path}\n\n"
                f"Anabilim Dalı ve Birim bilgileri dolu.\n"
                f"Dönem ayını D3 hücresine, yılı D4'e yazın."
            )

        except ImportError:
            hata_goster(self, "openpyxl kurulu değil.\npip install openpyxl")
        except Exception as e:
            hata_goster(self, f"Şablon oluşturulamadı:\n{e}")
