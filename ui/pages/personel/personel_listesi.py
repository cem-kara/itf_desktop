# -*- coding: utf-8 -*-
"""
Personel Listesi — v3 (Tema Entegrasyonu)
──────────────────────────────────────────
Tüm renkler merkezi ThemeManager / DarkTheme / ComponentStyles üzerinden gelir.
Hardcoded renk yok.
"""
from PySide6.QtCore import (
    Qt, QSortFilterProxyModel,
    Signal, QRect, QPoint, QSize, QTimer, QThread,
)
from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel,
    QFrame, QProgressBar, QPushButton, QTableView,
    QComboBox, QLineEdit, QMenu, QStyledItemDelegate,
    QStyle, QToolTip,
)
from PySide6.QtGui import (
    QColor, QCursor, QPainter, QBrush, QPen, QFont, QFontMetrics, QPixmap,
)

from core.logger import logger
from core.hata_yonetici import bilgi_goster, hata_goster, uyari_goster, soru_sor
from core.di import get_personel_service, get_izin_service
from ui.components.base_table_model import BaseTableModel
from ui.styles import DarkTheme
from ui.styles.icons import IconRenderer
from ui.theme_manager import ThemeManager

C = DarkTheme  # kısayol

# ── Durum renk yardımcıları (ComponentStyles yerine) ─────────────────────────
_STATUS_COLORS = {
    "Aktif":   (46,  201, 142, 38),   # yeşil  rgba(46,201,142,0.15)
    "Pasif":   (232, 85,  85,  38),   # kırmızı
    "İzinli":  (232, 160, 48,  38),   # turuncu
}
_STATUS_TEXT = {
    "Aktif":   "#2ec98e",
    "Pasif":   "#e85555",
    "İzinli":  "#e8a030",
}

def _get_status_color(durum: str) -> tuple:
    return _STATUS_COLORS.get(durum, (143, 163, 184, 25))

def _get_status_text_color(durum: str) -> str:
    return _STATUS_TEXT.get(durum, "#8fa3b8")


# ─── Sütun tanımları ──────────────────────────────────────────
COLUMNS = [
    ("_avatar",     "",               60),
    ("AdSoyad",     "Ad Soyad",      100),
    ("_tc_sicil",   "TC / Sicil",    100),
    ("_birim",      "Birim · Ünvan", 100),
    ("CepTelefonu", "Telefon",       120),
    ("_izin_bar",   "İzin Bakiye",   160),
    ("Durum",       "Durum",          320),
]
COL_IDX = {c[0]: i for i, c in enumerate(COLUMNS)}


# ─── Avatar Downloader Worker ──────────────────────────────
class AvatarDownloaderWorker(QThread):
    """Personel avatarlarını Drive'dan indir ve cache'le."""
    avatar_ready = Signal(str, QPixmap)  # (tc, pixmap)

    def __init__(self, image_url: str, tc: str, parent=None):
        super().__init__(parent)
        self._url = image_url
        self._tc = tc

    def run(self):
        try:
            if not self._url or self._url.strip().startswith("http") is False:
                return
            
            # Basit: URL'den indir (timeout 5s)
            import urllib.request
            import urllib.error
            try:
                req = urllib.request.Request(
                    self._url,
                    headers={'User-Agent': 'Mozilla/5.0'}
                )
                with urllib.request.urlopen(req, timeout=5) as response:
                    data = response.read()
                    pixmap = QPixmap()
                    if pixmap.loadFromData(data):
                        self.avatar_ready.emit(self._tc, pixmap)
            except (urllib.error.URLError, urllib.error.HTTPError, Exception) as e:
                logger.debug(f"Avatar download hatası ({self._tc}): {e}")
        except Exception as e:
            logger.debug(f"Avatar worker hatası: {e}")


# ═══════════════════════════════════════════════════════════
#  TABLO MODELİ
# ═══════════════════════════════════════════════════════════

class PersonelTableModel(BaseTableModel):

    RAW_ROW_ROLE  = Qt.ItemDataRole.UserRole + 1
    IZIN_PCT_ROLE = Qt.ItemDataRole.UserRole + 2   # float 0–1  (-1 = veri yok)
    IZIN_TXT_ROLE = Qt.ItemDataRole.UserRole + 3   # "13 / 20"

    def __init__(self, data=None, parent=None):
        super().__init__(COLUMNS, data, parent)
        self._izin_map: dict[str, dict] = {}

    def headerData(self, section, orientation, role=Qt.ItemDataRole.DisplayRole):
        if role == Qt.ItemDataRole.SizeHintRole and orientation == Qt.Orientation.Horizontal:
            return QSize(COLUMNS[section][2], 28)
        return super().headerData(section, orientation, role)

    def data(self, index, role=Qt.ItemDataRole.DisplayRole):
        if not index.isValid():
            return None
        row = self._data[index.row()]
        col = self._keys[index.column()]

        if role == self.RAW_ROW_ROLE:
            return row
        if role == self.IZIN_PCT_ROLE and col == "_izin_bar":
            return self._izin_pct(row)
        if role == self.IZIN_TXT_ROLE and col == "_izin_bar":
            return self._izin_txt(row)
        if role == Qt.ItemDataRole.DisplayRole:
            if col in ("_avatar", "_izin_bar", "_actions"): return ""
            if col == "AdSoyad":      return str(row.get("AdSoyad", ""))
            if col == "_tc_sicil":   return str(row.get("KimlikNo", ""))
            if col == "_birim":      return str(row.get("GorevYeri", ""))
            if col == "CepTelefonu": return str(row.get("CepTelefonu", "") or "—")
            if col == "Durum":       return str(row.get("Durum", ""))
            return str(row.get(col, ""))
        if role == Qt.ItemDataRole.TextAlignmentRole:
            if col in ("Durum", "_avatar"):
                return Qt.AlignmentFlag.AlignCenter
            return Qt.AlignmentFlag.AlignVCenter | Qt.AlignmentFlag.AlignLeft
        return None

    def get_row(self, idx: int):
        return self._data[idx] if 0 <= idx < len(self._data) else None

    def set_data(self, data: list):
        super().set_data(data)

    def set_izin_map(self, m: dict):
        """Geriye dönük uyumluluk için tutuldu. Artık izin bilgileri row içinde."""
        self._izin_map = m or {}
        if self._data:
            col = COL_IDX["_izin_bar"]
            self.dataChanged.emit(
                self.index(0, col),
                self.index(len(self._data) - 1, col),
            )

    def _izin_pct(self, row: dict) -> float:
        """İzin yüzdesi hesapla. Artık izin bilgileri doğrudan row içinde."""
        # Eski yöntem: _izin_map'ten al
        # tc = str(row.get("KimlikNo", "")).strip()
        # bilgi = self._izin_map.get(tc, {})
        
        # Yeni yöntem: row içinden oku (JOIN ile geldi)
        toplam = float(row.get("YillikToplamHak", 0) or 0)
        kalan  = float(row.get("YillikKalan", 0) or 0)
        if toplam <= 0:
            return -1.0
        return max(0.0, min(1.0, kalan / toplam))

    def _izin_txt(self, row: dict) -> str:
        """İzin metni. Artık izin bilgileri doğrudan row içinde."""
        # Eski yöntem: _izin_map'ten al
        # tc = str(row.get("KimlikNo", "")).strip()
        # bilgi = self._izin_map.get(tc, {})
        
        # Yeni yöntem: row içinden oku (JOIN ile geldi)
        kalan = row.get('YillikKalan', '—')
        toplam = row.get('YillikToplamHak', '—')
        return f"{kalan} / {toplam}"


# ═══════════════════════════════════════════════════════════
#  DELEGATE
# ═══════════════════════════════════════════════════════════

class PersonelDelegate(QStyledItemDelegate):
    """
    Özel hücre çizimi. Tüm renkler DarkTheme ve ComponentStyles'dan alınır.
    Avatar caching + fallback monogram destekli.
    """

    BTN_W, BTN_H, BTN_GAP = 52, 26, 6

    def __init__(self, parent=None):
        super().__init__(parent)
        self._hover_row  = -1
        self._avatar_cache: dict[str, QPixmap] = {}  # TC -> QPixmap
        self._avatar_loading: set[str] = set()       # Yüklenme sırasında olan TC'ler

    def set_hover_row(self, row: int):
        self._hover_row = row

    def set_avatar_pixmap(self, tc: str, pixmap: QPixmap):
        """Dış kaynaktan avatar cache'e pixmap ekle — daire şeklinde crop."""
        if pixmap and not pixmap.isNull():
            from PySide6.QtGui import QPainterPath
            
            # Daire şeklinde kırp
            size = 26
            tm = QPixmap(size, size)
            tm.fill(Qt.GlobalColor.transparent)
            
            pp = QPainter(tm)
            pp.setRenderHint(QPainter.RenderHint.Antialiasing)
            
            # Daire mask path'i oluştur
            path = QPainterPath()
            path.addEllipse(0, 0, size, size)
            pp.setClipPath(path)
            
            # Pixmap'i scaled versiyonu daire içine çiz
            scaled = pixmap.scaledToWidth(size, Qt.TransformationMode.SmoothTransformation)
            pp.drawPixmap(0, 0, scaled)
            pp.end()
            
            self._avatar_cache[tc] = tm
            self._avatar_loading.discard(tc)

    def sizeHint(self, option, index):
        return QSize(COLUMNS[index.column()][2], 40)

    def paint(self, painter: QPainter, option, index):
        painter.save()
        painter.setRenderHint(QPainter.RenderHint.Antialiasing)

        row  = index.row()
        col  = index.column()
        key  = COLUMNS[col][0]
        rect = option.rect
        is_sel   = bool(option.state & QStyle.StateFlag.State_Selected)
        is_hover = (row == self._hover_row)

        # ── Zemin (temadan) ──
        if is_sel:
            c = QColor(C.BTN_PRIMARY_BG)
            c.setAlpha(60)
            painter.fillRect(rect, c)
        elif is_hover:
            c = QColor(C.TEXT_PRIMARY)
            c.setAlpha(10)
            painter.fillRect(rect, c)

        raw = index.model().data(index, PersonelTableModel.RAW_ROW_ROLE)
        if raw is None:
            painter.restore()
            return

        if key == "_avatar":
            self._draw_avatar(painter, rect, raw)
        elif key == "AdSoyad":
            self._draw_primary(painter, rect, str(raw.get("AdSoyad", "")))
        elif key == "_tc_sicil":
            self._draw_two(painter, rect,
                           str(raw.get("KimlikNo", "")),
                           str(raw.get("KurumSicilNo", "") or ""),
                           mono_top=True)
        elif key == "_birim":
            self._draw_two(painter, rect,
                           str(raw.get("GorevYeri", "") or "—"),
                           str(raw.get("KadroUnvani", "") or ""))
        elif key == "CepTelefonu":
            self._draw_mono(painter, rect, str(raw.get("CepTelefonu", "") or "—"))
        elif key == "_izin_bar":
            pct = index.model().data(index, PersonelTableModel.IZIN_PCT_ROLE)
            txt = index.model().data(index, PersonelTableModel.IZIN_TXT_ROLE)
            self._draw_izin_bar(painter, rect, pct, txt)
        elif key == "Durum":
            self._draw_status_pill(painter, rect, raw)

        painter.restore()

    # ── Çizim yardımcıları ───────────────────────────────

    def _draw_avatar(self, p, rect, row):
        """Avatar çizimi: Fotoğraf cache'den varsa göster, yoksa monogram."""
        tc = str(row.get("KimlikNo", "")).strip()
        
        # Cache'te varsa fotoğraf göster
        if tc and tc in self._avatar_cache:
            pixmap = self._avatar_cache[tc]
            if pixmap and not pixmap.isNull():
                cx, cy = rect.center().x(), rect.center().y()
                px = int(cx - pixmap.width() / 2)
                py = int(cy - pixmap.height() / 2)
                p.drawPixmap(px, py, pixmap)
                return
        
        # Fallback: Monogram dairesi — renk addan türetilir
        ad = str(row.get("AdSoyad", "")).strip()
        initials = "".join(w[0].upper() for w in ad.split()[:2]) if ad else "?"
        cx, cy, r = rect.center().x(), rect.center().y(), 13
        hue = (sum(ord(c) for c in ad) * 37) % 360
        p.setBrush(QBrush(QColor.fromHsl(hue, 100, 55)))
        p.setPen(Qt.PenStyle.NoPen)
        p.drawEllipse(QPoint(cx, cy), r, r)
        p.setFont(QFont("", 8, QFont.Weight.Bold))
        p.setPen(QColor(C.TEXT_PRIMARY))
        p.drawText(rect, Qt.AlignmentFlag.AlignCenter, initials)

    def _draw_primary(self, p, rect, text):
        p.setFont(QFont("", 9, QFont.Weight.Medium))
        p.setPen(QColor(C.TEXT_PRIMARY))
        r = QRect(rect.x() + 8, rect.y(), rect.width() - 16, rect.height())
        p.drawText(r, Qt.AlignmentFlag.AlignVCenter | Qt.AlignmentFlag.AlignLeft,
                   p.fontMetrics().elidedText(text, Qt.TextElideMode.ElideRight, r.width()))

    def _draw_two(self, p, rect, top, bottom, mono_top=False):
        pad = 8
        r1 = QRect(rect.x() + pad, rect.y() + 4,  rect.width() - pad*2, 17)
        r2 = QRect(rect.x() + pad, rect.y() + 21, rect.width() - pad*2, 14)
        # Üst satır
        p.setFont(QFont("Courier New", 8) if mono_top else QFont("", 9, QFont.Weight.Medium))
        p.setPen(QColor(C.TEXT_SECONDARY))
        p.drawText(r1, Qt.AlignmentFlag.AlignVCenter | Qt.AlignmentFlag.AlignLeft,
                   p.fontMetrics().elidedText(top, Qt.TextElideMode.ElideRight, r1.width()))
        # Alt satır (muted)
        p.setFont(QFont("", 8))
        p.setPen(QColor(C.TEXT_MUTED))
        p.drawText(r2, Qt.AlignmentFlag.AlignVCenter | Qt.AlignmentFlag.AlignLeft,
                   p.fontMetrics().elidedText(bottom, Qt.TextElideMode.ElideRight, r2.width()))

    def _draw_mono(self, p, rect, text):
        p.setFont(QFont("Courier New", 8))
        p.setPen(QColor(C.TEXT_MUTED))
        r = QRect(rect.x() + 8, rect.y(), rect.width() - 16, rect.height())
        p.drawText(r, Qt.AlignmentFlag.AlignVCenter | Qt.AlignmentFlag.AlignLeft, text)

    def _draw_izin_bar(self, p, rect, pct, txt):
        """Üstte kalan/toplam metni, altta ince progress bar."""
        pad = 8
        bw  = rect.width() - pad * 2
        bx  = rect.x() + pad
        bh  = 4
        # Bar dikey ortalama: metin üstte (y+5..18), bar altta (y+24..28)
        ty  = rect.y() + 5
        by  = rect.y() + rect.height() - 10

        # Metin
        p.setFont(QFont("Courier New", 8))
        p.setPen(QColor(C.TEXT_MUTED))
        p.drawText(QRect(bx, ty, bw, 14), Qt.AlignmentFlag.AlignVCenter | Qt.AlignmentFlag.AlignLeft, txt or "—")

        # Arka plan (BG_TERTIARY)
        p.setBrush(QBrush(QColor(C.BG_TERTIARY)))
        p.setPen(Qt.PenStyle.NoPen)
        p.drawRoundedRect(bx, by, bw, bh, 2, 2)

        # Dolu kısım (pct'ye göre renk)
        if pct is not None and pct >= 0:
            if pct >= 0.7:
                fc = QColor(_get_status_text_color("Aktif"))   # yeşil
            elif pct >= 0.3:
                fc = QColor(C.BTN_PRIMARY_BORDER)                             # mavi
            else:
                fc = QColor(_get_status_text_color("İzinli"))  # sarı
            fw = max(3, int(bw * pct))
            p.setBrush(QBrush(fc))
            p.drawRoundedRect(bx, by, fw, bh, 2, 2)

    def _draw_status_pill(self, p, rect, row: dict):
        """Tema renklerini kullanan durum pill. Varsa DurumDetay metnini gösterir."""
        durum = str(row.get("Durum", "")).strip()
        text = str(row.get("DurumDetay", "")).strip() or durum

        r_val, g_val, b_val, a_val = _get_status_color(durum)
        fg_hex = _get_status_text_color(durum)

        font = QFont("", 8, QFont.Weight.Medium)
        p.setFont(font)
        fm   = QFontMetrics(font)
        max_pw = max(60, rect.width() - 12)
        elided = fm.elidedText(text, Qt.TextElideMode.ElideRight, max_pw - 20)
        tw = fm.horizontalAdvance(elided)
        pw, ph = min(max_pw, tw + 20), fm.height() + 8
        px = rect.x() + 6
        py   = rect.center().y() - ph // 2

        p.setBrush(QBrush(QColor(r_val, g_val, b_val, a_val)))
        p.setPen(QPen(QColor(r_val, g_val, b_val, min(a_val + 80, 255)), 1))
        p.drawRoundedRect(px, py, pw, ph, 4, 4)
        p.setPen(QColor(fg_hex))
        p.drawText(
            QRect(px + 8, py, pw - 12, ph),
            Qt.AlignmentFlag.AlignVCenter | Qt.AlignmentFlag.AlignLeft,
            elided,
        )


# ═══════════════════════════════════════════════════════════
#  PERSONEL LİSTESİ SAYFASI
# ═══════════════════════════════════════════════════════════

class PersonelListesiPage(QWidget):

    detay_requested = Signal(dict)
    izin_requested  = Signal(dict)
    yeni_requested  = Signal()
    close_requested = Signal()

    def __init__(self, db=None, action_guard=None, parent=None):
        super().__init__(parent)
        self.setProperty("bg-role", "page")
        self._db             = db
        self._svc            = get_personel_service(db) if db else None
        self._izin_svc       = get_izin_service(db) if db else None
        self._action_guard   = action_guard
        self._all_data       = []
        self._izin_map       = {}
        self._active_filter  = "Aktif"
        self._filter_btns    = {}
        self._izinli_bugun   = {}
        self._izinli_bugun_date = None
        self._last_tooltip_tc = None
        self._filtered_data   = []  # Son filtrelenmiş veri (Excel export için)
        
        # Arama debounce timer (300ms)
        self._search_timer = QTimer()
        self._search_timer.setInterval(300)
        self._search_timer.setSingleShot(True)
        self._search_timer.timeout.connect(self._execute_search)
        self._last_search_text = ""
        
        # Avatar download workers
        self._avatar_workers: list[AvatarDownloaderWorker] = []
        
        # Lazy-loading pagination state
        self._page_size = 100  # Her sayfada kaç kayıt
        self._current_page = 1  # Şu anki sayfa
        self._total_count = 0  # Toplam kayıt sayısı
        self._is_loading = False  # Yükleme durumu (spinner göstermek için)
        
        self._setup_ui()
        self._connect_signals()
        self._theme_manager = ThemeManager.instance()
        self._theme_manager.theme_changed.connect(self._on_theme_changed)
        self._apply_runtime_styles()

    # ─── UI Kurulum ──────────────────────────────────────

    def _setup_ui(self):
        main = QVBoxLayout(self)
        main.setContentsMargins(0, 0, 0, 0)
        main.setSpacing(0)
        main.addWidget(self._build_toolbar())
        main.addWidget(self._build_subtoolbar())
        main.addWidget(self._build_table(), 1)
        main.addWidget(self._build_footer())

    def _build_toolbar(self) -> QFrame:
        frame = QFrame()
        self._toolbar_frame = frame
        frame.setFixedHeight(60)
        lay = QHBoxLayout(frame)
        lay.setContentsMargins(16, 0, 16, 0)
        lay.setSpacing(8)

        title = QLabel("Personel Listesi")
        title.setProperty("color-role", "primary")
        title.setProperty("bg-role", "panel")
        lay.addWidget(title)
        lay.addWidget(self._sep())

        # Durum filtre pill butonları — her durum için özel renk
        filter_styles = {
            "Aktif": {
                "bg": _get_status_color("Aktif"),
                "text": _get_status_text_color("Aktif")
            },
            "Pasif": {
                "bg": _get_status_color("Pasif"),
                "text": _get_status_text_color("Pasif")
            },
            "İzinli": {
                "bg": _get_status_color("İzinli"),
                "text": _get_status_text_color("İzinli")
            },
            "Tümü": {
                "bg": (30, 180, 216, 35),  # Accent color
                "text": C.ACCENT
            }
        }
        
        for lbl in ("Aktif", "Pasif", "İzinli", "Tümü"):
            btn = QPushButton(lbl)
            btn.setCheckable(True)
            btn.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
            btn.setFixedHeight(28)
            btn.setMinimumWidth(90)  # Sayıların sığması için sabit genişlik
            
            # Her buton için özel stil
            style_info = filter_styles[lbl]
            r, g, b, a = style_info["bg"]
            text_color = style_info["text"]
            
            btn_style = f"""
                QPushButton {{
                    background: rgba({r}, {g}, {b}, {a});
                    color: {text_color};
                    border: 1px solid rgba({r}, {g}, {b}, {min(a + 80, 255)});
                    border-radius: 4px;
                    padding: 4px 12px;
                    font-size: 11px;
                    font-weight: 500;
                }}
                QPushButton:hover {{
                    background: rgba({r}, {g}, {b}, {min(a + 30, 255)});
                    border: 1px solid rgba({r}, {g}, {b}, {min(a + 100, 255)});
                }}
                QPushButton:checked {{
                    background: rgba({r}, {g}, {b}, {min(a + 60, 255)});
                    border: 1px solid {text_color};
                    font-weight: 600;
                }}
            """
            btn.setStyleSheet(btn_style)
            
            if lbl == self._active_filter:
                btn.setChecked(True)
            self._filter_btns[lbl] = btn
            lay.addWidget(btn)

        lay.addWidget(self._sep())

        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("Ad, TC, birim ara…")
        self.search_input.setClearButtonEnabled(True)
        self.search_input.setFixedWidth(200)
        # setStyleSheet kaldırıldı: search — global QSS kuralı geçerli
        lay.addWidget(self.search_input)

        lay.addStretch()

        self.btn_yenile = QPushButton()
        self.btn_yenile.setToolTip("Yenile")
        self.btn_yenile.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self.btn_yenile.setFixedSize(32, 28)
        self.btn_yenile.setProperty("style-role", "refresh")
        IconRenderer.set_button_icon(self.btn_yenile, "refresh", color=C.TEXT_SECONDARY, size=16)
        lay.addWidget(self.btn_yenile)

        self.btn_yeni = QPushButton(" Yeni Personel")
        self.btn_yeni.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self.btn_yeni.setProperty("style-role", "action")
        IconRenderer.set_button_icon(self.btn_yeni, "user_add", color=C.BTN_PRIMARY_TEXT, size=18)
        self.btn_yeni.setIconSize(QSize(18, 18))
        
        # IP-06: Aksiyon bazlı yetki kontrolü
        if self._action_guard:
            self._action_guard.disable_if_unauthorized(self.btn_yeni, "personel.write")
        
        lay.addWidget(self.btn_yeni)
        
        # Kapat butonu (en sağ köşe)
        self.btn_close = QPushButton()
        self.btn_close.setFixedSize(28, 28)
        self.btn_close.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self.btn_close.setToolTip("Kapat")
        self.btn_close.setProperty("style-role", "close")
        IconRenderer.set_button_icon(self.btn_close, "x", color=C.TEXT_SECONDARY, size=16)
        self.btn_close.setIconSize(QSize(16, 16))
        lay.addWidget(self.btn_close)
        return frame

    def _build_subtoolbar(self) -> QFrame:
        frame = QFrame()
        self._subtoolbar_frame = frame
        frame.setFixedHeight(48)
        lay = QHBoxLayout(frame)
        lay.setContentsMargins(16, 0, 16, 0)
        lay.setSpacing(8)

        lbl = QLabel("FİLTRE:")
        lbl.setProperty("color-role", "disabled")
        lbl.setProperty("bg-role", "panel")
        lay.addWidget(lbl)

        self.cmb_gorev_yeri = QComboBox()
        self.cmb_gorev_yeri.addItem("Tüm Birimler")
        self.cmb_gorev_yeri.setFixedWidth(250)
        # setStyleSheet kaldırıldı: combo — global QSS kuralı geçerli
        lay.addWidget(self.cmb_gorev_yeri)

        self.cmb_hizmet = QComboBox()
        self.cmb_hizmet.addItem("Tüm Sınıflar")
        self.cmb_hizmet.setFixedWidth(300)
        # setStyleSheet kaldırıldı: combo — global QSS kuralı geçerli
        lay.addWidget(self.cmb_hizmet)

        lay.addStretch()

        self.btn_excel = QPushButton(" Excel")
        self.btn_excel.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self.btn_excel.setProperty("style-role", "success")
        IconRenderer.set_button_icon(self.btn_excel, "download", color=C.TEXT_SECONDARY, size=16)
        self.btn_excel.setIconSize(QSize(16, 16))
        lay.addWidget(self.btn_excel)
        return frame

    def _build_table(self) -> QTableView:
        self._model = PersonelTableModel()
        self._proxy = QSortFilterProxyModel()
        self._proxy.setSourceModel(self._model)
        self._proxy.setFilterCaseSensitivity(Qt.CaseSensitivity.CaseInsensitive)
        self._proxy.setFilterKeyColumn(-1)

        self.table = QTableView()
        self.table.setModel(self._proxy)
        # setStyleSheet kaldırıldı: table — global QSS kuralı geçerli
        self.table.setAlternatingRowColors(False)
        self.table.setSelectionBehavior(QTableView.SelectionBehavior.SelectRows)
        self.table.setSelectionMode(QTableView.SelectionMode.SingleSelection)
        self.table.setSortingEnabled(True)
        self.table.verticalHeader().setVisible(False)
        self.table.setShowGrid(False)
        self.table.setMouseTracking(True)
        self.table.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.table.verticalHeader().setDefaultSectionSize(40)

        self._delegate = PersonelDelegate(self.table)
        self.table.setItemDelegate(self._delegate)

        self._model.setup_columns(
            self.table,
            stretch_keys=["AdSoyad", "_birim"],
        )
        return self.table

    def _build_footer(self) -> QFrame:
        frame = QFrame()
        self._footer_frame = frame
        frame.setFixedHeight(40)
        lay = QHBoxLayout(frame)
        lay.setContentsMargins(16, 0, 16, 0)
        lay.setSpacing(16)

        self.lbl_info = QLabel("0 kayıt")
        self.lbl_info.setProperty("style-role", "footer")
        lay.addWidget(self.lbl_info)

        self.lbl_detail = QLabel("")
        self.lbl_detail.setProperty("style-role", "footer")
        lay.addWidget(self.lbl_detail)

        lay.addStretch()

        self.progress = QProgressBar()
        self.progress.setFixedSize(140, 4)
        self.progress.setVisible(False)
        self.progress.setTextVisible(False)
        # setStyleSheet kaldırıldı: progress — global QSS kuralı geçerli
        lay.addWidget(self.progress)

        # Lazy-loading: "Daha fazla yükle" butonu
        self.btn_load_more = QPushButton("Daha Fazla Yükle")
        self.btn_load_more.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self.btn_load_more.setFixedHeight(28)
        self.btn_load_more.setProperty("style-role", "action")
        self.btn_load_more.setVisible(False)  # İlk başta gizle
        lay.addWidget(self.btn_load_more)
        
        return frame

    # ─── Sinyaller ───────────────────────────────────────

    def _connect_signals(self):
        for text, btn in self._filter_btns.items():
            btn.clicked.connect(lambda _, t=text: self._on_filter_click(t))
        self.search_input.textChanged.connect(self._on_search)
        self.cmb_gorev_yeri.currentTextChanged.connect(lambda _: self._apply_filters())
        self.cmb_hizmet.currentTextChanged.connect(lambda _: self._apply_filters())
        self.btn_close.clicked.connect(self.close_requested.emit)
        self.btn_yenile.clicked.connect(self.load_data)
        self.btn_yeni.clicked.connect(self.yeni_requested.emit)
        self.btn_load_more.clicked.connect(self._load_more_data)
        self.btn_excel.clicked.connect(self._export_excel)
        self.table.doubleClicked.connect(self._on_double_click)
        self.table.customContextMenuRequested.connect(self._show_context_menu)
        self.table.mouseMoveEvent  = self._tbl_mouse_move
        self.table.mousePressEvent = self._tbl_mouse_press

    def _on_theme_changed(self, _theme_name: str):
        """Tema değiştiğinde runtime stilleri yeniden uygula."""
        self._apply_runtime_styles()
        self.table.viewport().update()

    def _apply_runtime_styles(self):
        """Tema token'larına bağlı stilleri yeniden uygular."""
        if hasattr(self, "_toolbar_frame"):
            self._toolbar_frame.setStyleSheet(
                "QFrame { background-color: %s; border-bottom: 1px solid %s; }"
                % (C.BG_SECONDARY, C.BORDER_PRIMARY)
            )
        if hasattr(self, "_subtoolbar_frame"):
            self._subtoolbar_frame.setStyleSheet(
                "QFrame { background-color: %s; border-bottom: 1px solid %s; }"
                % (C.BG_PRIMARY, C.BORDER_PRIMARY)
            )
        if hasattr(self, "_footer_frame"):
            self._footer_frame.setStyleSheet(
                "QFrame { background-color: %s; border-top: 1px solid %s; }"
                % (C.BG_SECONDARY, C.BORDER_PRIMARY)
            )

        # Durum filtre butonlarını aktif tema renkleriyle yeniden boya.
        filter_styles = {
            "Aktif": {
                "bg": _get_status_color("Aktif"),
                "text": _get_status_text_color("Aktif")
            },
            "Pasif": {
                "bg": _get_status_color("Pasif"),
                "text": _get_status_text_color("Pasif")
            },
            "İzinli": {
                "bg": _get_status_color("İzinli"),
                "text": _get_status_text_color("İzinli")
            },
            "Tümü": {
                "bg": (30, 180, 216, 35),
                "text": C.ACCENT
            }
        }
        for lbl, btn in self._filter_btns.items():
            style_info = filter_styles.get(lbl)
            if not style_info:
                continue
            r, g, b, a = style_info["bg"]
            text_color = style_info["text"]
            btn.setStyleSheet(
                f"""
                QPushButton {{
                    background: rgba({r}, {g}, {b}, {a});
                    color: {text_color};
                    border: 1px solid rgba({r}, {g}, {b}, {min(a + 80, 255)});
                    border-radius: 4px;
                    padding: 4px 12px;
                    font-size: 11px;
                    font-weight: 500;
                }}
                QPushButton:hover {{
                    background: rgba({r}, {g}, {b}, {min(a + 30, 255)});
                    border: 1px solid rgba({r}, {g}, {b}, {min(a + 100, 255)});
                }}
                QPushButton:checked {{
                    background: rgba({r}, {g}, {b}, {min(a + 60, 255)});
                    border: 1px solid {text_color};
                    font-weight: 600;
                }}
                """
            )

        if hasattr(self, "btn_close"):
            self.btn_close.setProperty("style-role", "close")

    # ─── Veri Yükleme ────────────────────────────────────

    def load_data(self):
        """İlk kez veri yükle (Sayfa 1, Pagination ile)"""
        if not self._db or not self._svc:
            logger.warning("Personel listesi: DB yok")
            return
        try:
            # Lazy-loading: Başlangıç state'i ayarla
            self._current_page = 1
            self._total_count = 0
            self._all_data = []
            
            personel_repo_sonuc = self._svc.get_personel_repo()
            if not personel_repo_sonuc.basarili or not personel_repo_sonuc.veri:
                logger.warning("Personel repository alınamadı")
                return
            personel_repo = personel_repo_sonuc.veri
            
            # ✅ LAZY-LOADING: İlk sayfayı yükle (sayfa boyutu: 100 kayıt)
            page_data, total_count = personel_repo.get_paginated_with_bakiye(
                page=self._current_page,
                page_size=self._page_size
            )
            
            self._all_data = page_data
            self._total_count = total_count
            self._izin_map = {}
            
            # Model'e veriyi set et
            self._model.set_data(self._all_data)
            self._model.set_izin_map({})
            
            self._populate_combos()
            self._refresh_izinli_bugun()
            self._apply_filters()
            
            # Avatar download başlat (arka planda)
            self._start_avatar_downloads()
            
            # "Daha fazla yükle" butonu kontrol et
            self._update_load_more_button()
            
        except Exception as e:
            logger.error(f"Personel yükleme: {e}")

    def _load_more_data(self):
        """Sonraki sayfayı yükle ve tabloya ekle"""
        if self._is_loading or not self._db or not self._svc:
            return
        
        try:
            self._is_loading = True
            self.btn_load_more.setEnabled(False)
            self.progress.setVisible(True)
            
            personel_repo_sonuc = self._svc.get_personel_repo()
            if not personel_repo_sonuc.basarili or not personel_repo_sonuc.veri:
                logger.warning("Personel repository alınamadı")
                return
            personel_repo = personel_repo_sonuc.veri
            
            # Sonraki sayfa
            self._current_page += 1
            page_data, _ = personel_repo.get_paginated_with_bakiye(
                page=self._current_page,
                page_size=self._page_size
            )
            
            if not page_data:
                # Daha kayıt yok
                self._current_page -= 1
                logger.info("Tüm personel yüklendi")
            else:
                # Mevcut verilere ekle
                self._all_data.extend(page_data)
                self._model.set_data(self._all_data)
                self._apply_filters()
                self._start_avatar_downloads()
                logger.debug(f"Sayfa {self._current_page} yüklendi ({len(page_data)} kayıt)")
            
            # Button kontrolü
            self._update_load_more_button()
            
        except Exception as e:
            logger.error(f"Daha fazla yükleme: {e}")
        finally:
            self._is_loading = False
            self.btn_load_more.setEnabled(True)
            self.progress.setVisible(False)

    def _update_load_more_button(self):
        """'Daha fazla yükle' butonunun görünürlüğünü ayarla"""
        loaded_count = len(self._all_data)
        has_more = loaded_count < self._total_count
        self.btn_load_more.setVisible(has_more)
        if has_more:
            self.btn_load_more.setText(
                f"Daha Fazla Yükle ({loaded_count}/{self._total_count})"
            )
    
    def _start_avatar_downloads(self):
        """Tüm personelin avatarlarını arka panda indir."""
        for row in self._all_data:
            tc = str(row.get("KimlikNo", "")).strip()
            resim_url = str(row.get("Resim", "")).strip()
            
            # URL varsa ve http(s) ile başlıyorsa download başlat
            if tc and resim_url and resim_url.startswith("http"):
                worker = AvatarDownloaderWorker(resim_url, tc, self)
                worker.avatar_ready.connect(self._on_avatar_ready)
                self._avatar_workers.append(worker)
                worker.start()
    
    def _on_avatar_ready(self, tc: str, pixmap: QPixmap):
        """Avatar indirme tamamlandı — delegate'e pixmap ekle ve table redraw et."""
        if self._delegate:
            self._delegate.set_avatar_pixmap(tc, pixmap)
            # Tüm satırları redraw et (avatar sütunu değişti)
            self.table.viewport().update()

    def _populate_combos(self):
        """Filtreleme combo box'larını doldur."""
        if not self._svc:
            return
        try:
            # Görev yerleri
            gorev_yerleri = self._svc.get_gorev_yerleri().veri or []
            self.cmb_gorev_yeri.blockSignals(True)
            self.cmb_gorev_yeri.clear()
            self.cmb_gorev_yeri.addItem("Tüm Birimler")
            self.cmb_gorev_yeri.addItems(gorev_yerleri)
            self.cmb_gorev_yeri.blockSignals(False)

            # Hizmet sınıfları
            hizmet_siniflari = self._svc.get_hizmet_siniflari().veri or []
            self.cmb_hizmet.blockSignals(True)
            self.cmb_hizmet.clear()
            self.cmb_hizmet.addItem("Tüm Sınıflar")
            self.cmb_hizmet.addItems(hizmet_siniflari)
            self.cmb_hizmet.blockSignals(False)
        except Exception as e:
            logger.error(f"Combo doldurma hatası: {e}")

    # ─── Filtreleme ──────────────────────────────────────

    def _on_filter_click(self, text: str):
        self._active_filter = text
        for t, btn in self._filter_btns.items():
            btn.setChecked(t == text)
        self._apply_filters()

    def _on_search(self, text: str):
        """Arama input'u değiştiğinde debounce timer'ını restart et."""
        self._last_search_text = text
        self._search_timer.stop()
        self._search_timer.start()  # 300ms delay sonra _execute_search çağrılır

    def _execute_search(self):
        """Debounce timeout sonrası gerçek aramaları yap."""
        self._proxy.setFilterFixedString(self._last_search_text)
        self._update_count()

    def _apply_filters(self):
        filtered = self._all_data
        self._refresh_izinli_bugun()

        if self._active_filter != "Tümü":
            if self._active_filter == "İzinli":
                self._refresh_izinli_bugun()
                izinli_tc = set(self._izinli_bugun.keys())
                filtered = [r for r in filtered
                           if str(r.get("KimlikNo", "")).strip() in izinli_tc]
            else:
                # Durum filtresini uygula (sadece Durum sütununa bak)
                filtered = [r for r in filtered
                           if str(r.get("Durum", "")).strip() == self._active_filter]

        birim = self.cmb_gorev_yeri.currentText()
        if birim and birim != "Tüm Birimler":
            filtered = [r for r in filtered
                        if str(r.get("GorevYeri", "")).strip() == birim]

        sinif = self.cmb_hizmet.currentText()
        if sinif and sinif != "Tüm Sınıflar":
            filtered = [r for r in filtered
                        if str(r.get("HizmetSinifi", "")).strip() == sinif]

        # Tooltip metnini Durum sütununda da göster (bugün izinli personeller).
        render_rows = []
        for row in filtered:
            row_view = dict(row)
            tc = str(row_view.get("KimlikNo", "")).strip()
            izinler = self._izinli_bugun.get(tc, [])
            if izinler:
                bitisler = [str(bit).strip() for _, bit in izinler if str(bit).strip()]
                bitis = max(bitisler) if bitisler else ""
                row_view["DurumDetay"] = (
                    f"Personel bugün izinli izin bitiş tarih: {bitis}"
                    if bitis else
                    "Personel bugün izinli"
                )
            render_rows.append(row_view)

        self._model.set_data(render_rows)
        self._update_count()
        self._update_pill_counts()
        self._filtered_data = filtered  # Excel export için sakla
    def _refresh_izinli_bugun(self):
        from datetime import date
        today = date.today()
        if self._izinli_bugun_date != today:
            self._izinli_bugun = self._get_izinli_bugun_detay()
            self._izinli_bugun_date = today

    def _get_izinli_bugun_detay(self) -> dict:
        """Bugün izinli olan personellerin detaylı bilgisini getir."""
        if not self._izin_svc:
            return {}
        try:
            return self._izin_svc.get_izinli_personeller_bugun().veri or {}
        except Exception as e:
            logger.error(f"İzinli sorgu: {e}")
            return {}

    def _update_count(self):
        self.lbl_info.setText(
            f"Gösterilen {self._proxy.rowCount()} / {len(self._all_data)}"
        )

    def _update_pill_counts(self):
        """Buton sayaçlarını güncelle (Durum sütununa göre)"""
        aktif = sum(1 for r in self._all_data 
                   if str(r.get("Durum", "")).strip() == "Aktif")
        pasif = sum(1 for r in self._all_data 
                   if str(r.get("Durum", "")).strip() == "Pasif")
        self._refresh_izinli_bugun()
        izinli_tc = set(self._izinli_bugun.keys())
        izinli = sum(1 for r in self._all_data
                    if str(r.get("KimlikNo", "")).strip() in izinli_tc)
        
        self.lbl_detail.setText(f"Aktif {aktif}  ·  Pasif {pasif}  ·  İzinli {izinli}")
        counts = {"Aktif": aktif, "Pasif": pasif, "İzinli": izinli, "Tümü": len(self._all_data)}
        for t, btn in self._filter_btns.items():
            c = counts.get(t, "")
            if c != "":
                btn.setText(f"{t} ({c})")
            else:
                btn.setText(t)

    # ─── Mouse ───────────────────────────────────────────
    def _tbl_mouse_move(self, event):
        idx = self.table.indexAt(event.pos())
        src_row = self._proxy.mapToSource(idx).row() if idx.isValid() else -1
        self._delegate.set_hover_row(src_row)
        self.table.viewport().update()
        self._show_izin_tooltip(idx, event.globalPos())
        QTableView.mouseMoveEvent(self.table, event)

    def _show_izin_tooltip(self, idx, global_pos):
        if not idx.isValid():
            if self._last_tooltip_tc:
                QToolTip.hideText()
                self._last_tooltip_tc = None
            return

        src = self._proxy.mapToSource(idx)
        row_data = self._model.get_row(src.row())
        tc = str((row_data or {}).get("KimlikNo", "")).strip()
        self._refresh_izinli_bugun()
        if tc and tc in self._izinli_bugun:
            if tc != self._last_tooltip_tc:
                text = self._build_izin_tooltip(tc)
                if text:
                    QToolTip.showText(global_pos, text, self.table)
                    self._last_tooltip_tc = tc
        else:
            if self._last_tooltip_tc:
                QToolTip.hideText()
                self._last_tooltip_tc = None

    def _build_izin_tooltip(self, tc: str) -> str:
        izinler = self._izinli_bugun.get(tc, [])
        if not izinler:
            return ""
        lines = ["Bugun izinli:"]
        for bas, bit in izinler:
            if bas == bit:
                lines.append(f"- {bas}")
            else:
                lines.append(f"- {bas} -> {bit}")
        return "\n".join(lines)
    def _tbl_mouse_press(self, event):
        idx = self.table.indexAt(event.pos())
        if idx.isValid():
            self._proxy.mapToSource(idx)
            # Butonlar sadece double click'te gösterilsin, mouse press'e alet yapma
        QTableView.mousePressEvent(self.table, event)

    def _on_double_click(self, index):
        src = self._proxy.mapToSource(index)
        row = self._model.get_row(src.row())
        if row:
            self.detay_requested.emit(row)

    def get_selected(self):
        idxs = self.table.selectionModel().selectedRows()
        if idxs:
            return self._model.get_row(self._proxy.mapToSource(idxs[0]).row())
        return None

    # ─── Sağ tık ─────────────────────────────────────────

    def _show_context_menu(self, pos):
        idx = self.table.indexAt(pos)
        if not idx.isValid():
            return
        row = self._model.get_row(self._proxy.mapToSource(idx).row())
        if not row:
            return
        ad, tc = row.get("AdSoyad", ""), row.get("KimlikNo", "")
        durum  = str(row.get("Durum", "")).strip()

        menu = QMenu(self)
        menu.addAction("Detay Görüntüle").triggered.connect(
            lambda: self.detay_requested.emit(row))
        menu.addSeparator()
        menu.addAction("İzin Girişi").triggered.connect(
            lambda: self.izin_requested.emit(row))
        menu.addSeparator()
        for d in ("Aktif", "Pasif", "İzinli"):
            if d != durum:
                menu.addAction(f"{d} Yap").triggered.connect(
                    lambda _, dd=d: self._change_durum(tc, ad, dd))
        menu.exec(self.table.viewport().mapToGlobal(pos))

    def _change_durum(self, tc, ad, yeni):
        if not self._svc:
            hata_goster(self, "Servis bağlantısı yok")
            return
        if not soru_sor(
            self,
            f'"{ad}" personelinin durumu "{yeni}" olarak değiştirilsin mi?',
            "Durum Değiştir",
        ):
            return
        try:
            # Service kullanarak güncelle
            success = self._svc.guncelle(tc, {"Durum": yeni})
            if success:
                logger.info(f"Durum: {tc} → {yeni}")
                self.load_data()
            else:
                hata_goster(self, "Durum güncellemesi başarısız")
        except Exception as e:
            logger.error(f"Durum değiştirme: {e}")
            hata_goster(self, f"İşlem başarısız:\n{e}")

    # ─── Excel Export ────────────────────────────────────

    def _get_all_filtered_data(self):
        """Pagination olmadan TÜM personel verilerini çek ve filtrele (Excel export için)."""
        if not self._db or not self._svc:
            return []
        
        try:
            # TÜM personel verilerini çek (pagination olmadan)
            personel_repo_sonuc = self._svc.get_personel_repo()
            if not personel_repo_sonuc.basarili or not personel_repo_sonuc.veri:
                return []
            personel_repo = personel_repo_sonuc.veri
            all_data = personel_repo.get_all_with_bakiye() or []
            
            # Mevcut filtreleri uygula
            filtered = all_data
            
            # Durum filtresi
            if self._active_filter != "Tümü":
                if self._active_filter == "İzinli":
                    self._refresh_izinli_bugun()
                    izinli_tc = set(self._izinli_bugun.keys())
                    filtered = [r for r in filtered
                               if str(r.get("KimlikNo", "")).strip() in izinli_tc]
                else:
                    filtered = [r for r in filtered
                               if str(r.get("Durum", "")).strip() == self._active_filter]
            
            # Birim filtresi
            birim = self.cmb_gorev_yeri.currentText()
            if birim and birim != "Tüm Birimler":
                filtered = [r for r in filtered
                            if str(r.get("GorevYeri", "")).strip() == birim]
            
            # Sınıf filtresi
            sinif = self.cmb_hizmet.currentText()
            if sinif and sinif != "Tüm Sınıflar":
                filtered = [r for r in filtered
                            if str(r.get("HizmetSinifi", "")).strip() == sinif]
            
            # Arama filtresi
            if self._last_search_text:
                search_lower = self._last_search_text.lower()
                filtered = [r for r in filtered
                           if search_lower in str(r.get("AdSoyad", "")).lower()
                           or search_lower in str(r.get("KimlikNo", "")).lower()
                           or search_lower in str(r.get("GorevYeri", "")).lower()
                           or search_lower in str(r.get("KadroUnvani", "")).lower()]
            
            return filtered
        except Exception as e:
            logger.error(f"Tüm veri yükleme hatası: {e}")
            return []

    def _export_excel(self):
        """Filtrelenmiş personel listesini Excel dosyasına dışa aktar."""
        # TÜM filtrelenmiş verileri çek (pagination olmadan)
        export_data = self._get_all_filtered_data()
        
        if not export_data:
            uyari_goster(self, "Dışa aktarılacak veri yok")
            return
        
        try:
            from PySide6.QtWidgets import QFileDialog
            from datetime import datetime
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            # Sheet adını filtrelere göre oluştur
            sheet_name = self._build_export_sheet_name()
            
            # Kaydetme dialogu
            file_path, _ = QFileDialog.getSaveFileName(
                self,
                "Excel Dosyasını Kaydet",
                f"Personel_Listesi_{sheet_name}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                "Excel Files (*.xlsx)"
            )
            
            if not file_path:
                return
            
            # Excel workbook oluştur
            wb = openpyxl.Workbook()
            # Varsayılan sheet'i kaldır ve yeni sheet oluştur
            if len(wb.sheetnames) > 0:
                default_sheet = wb.active
                if default_sheet:
                    wb.remove(default_sheet)
            ws = wb.create_sheet(sheet_name[:31])
            
            # Header'ları yazma
            headers = ["Ad Soyad", "TC Kimlik No", "Sicil No", "Birim", "Ünvan", 
                      "Telefon", "İzin Bakiye", "Durum"]
            
            # Header styling
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Header satırını yaz
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col_idx)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = thin_border
            
            # Veri satırlarını yaz
            data_font = Font(size=10)
            data_alignment = Alignment(horizontal="left", vertical="center")
            
            for row_idx, row_data in enumerate(export_data, start=2):
                values = [
                    str(row_data.get("AdSoyad", "")),
                    str(row_data.get("KimlikNo", "")),
                    str(row_data.get("KurumSicilNo", "") or ""),
                    str(row_data.get("GorevYeri", "") or ""),
                    str(row_data.get("KadroUnvani", "") or ""),
                    str(row_data.get("CepTelefonu", "") or ""),
                    f"{row_data.get('YillikKalan', '—')} / {row_data.get('YillikToplamHak', '—')}",
                    str(row_data.get("Durum", "")),
                ]
                
                for col_idx, value in enumerate(values, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.value = value
                    cell.font = data_font
                    cell.alignment = data_alignment
                    cell.border = thin_border
            
            # Sütun genişlikleri
            ws.column_dimensions['A'].width = 20  # Ad Soyad
            ws.column_dimensions['B'].width = 15  # TC
            ws.column_dimensions['C'].width = 15  # Sicil
            ws.column_dimensions['D'].width = 18  # Birim
            ws.column_dimensions['E'].width = 20  # Ünvan
            ws.column_dimensions['F'].width = 15  # Telefon
            ws.column_dimensions['G'].width = 15  # İzin Bakiye
            ws.column_dimensions['H'].width = 12  # Durum
            
            # İlk satırı dondur
            ws.freeze_panes = "A2"
            
            # Dosyayı kaydet
            wb.save(file_path)
            
            bilgi_goster(
                self,
                f"Excel dosyası başarıyla kaydedildi:\n{file_path}\n\nToplam: {len(export_data)} kayıt",
                "Başarılı",
            )
            logger.info(f"Excel export: {file_path} ({len(export_data)} kayıt)")
            
        except Exception as e:
            logger.error(f"Excel export hatası: {e}")
            hata_goster(self, f"Excel dosyası oluşturulamadı:\n{e}")
    
    def _build_export_sheet_name(self) -> str:
        """Filter durumuna göre sheet adı oluştur."""
        parts = []
        
        # Durum filtresi
        if self._active_filter != "Tümü":
            parts.append(self._active_filter)
        
        # Birim filtresi
        birim = self.cmb_gorev_yeri.currentText()
        if birim and birim != "Tüm Birimler":
            parts.append(birim[:15])  # Uzunluğu sınırla
        
        # Sınıf filtresi
        sinif = self.cmb_hizmet.currentText()
        if sinif and sinif != "Tüm Sınıflar":
            parts.append(sinif[:15])  # Uzunluğu sınırla
        
        # Arama filtresi
        if self._last_search_text:
            parts.append("Arama")
        
        # Default
        if not parts:
            return "Personel_Listesi"
        
        sheet_name = "_".join(parts)
        # Sheet adı max 31 karakter olmalı
        return sheet_name[:31]

    # ─── Yardımcılar ─────────────────────────────────────

    @staticmethod
    def _sep() -> QFrame:
        s = QFrame()
        s.setFixedSize(1, 20)
        s.setProperty("bg-role", "separator")
        return s
