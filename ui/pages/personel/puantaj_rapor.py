# -*- coding: utf-8 -*-
"""
Puantaj Raporlama ve Şua Takip Sistemi

Orijinal form: fshz_puantaj.py → PuantajRaporPenceresi
─────────────────────────────────────────────────────
• FHSZ_Puantaj tablosundan yıl/dönem bazlı rapor
• Kümülatif saat: Aynı yılda, o dönem dahil önceki tüm dönemlerin toplamı
• Hak Edilen Şua: sua_hak_edis_hesapla(kümülatif)
• Excel / PDF dışa aktarım
"""
import os
from datetime import datetime
from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton,
    QComboBox, QFrame, QTableWidget, QTableWidgetItem,
    QHeaderView, QProgressBar,
    QAbstractItemView, QFileDialog, QStyledItemDelegate, QStyle
)
from PySide6.QtCore import Qt, QRectF
from PySide6.QtGui import QColor, QCursor, QFont, QPainter, QBrush, QPen, QPainterPath

from core.logger import logger
from ui.dialogs.mesaj_kutusu import MesajKutusu
from core.hesaplamalar import sua_hak_edis_hesapla
from ui.styles import Colors, DarkTheme
from ui.styles.icons import IconRenderer


AY_ISIMLERI = [
    "Ocak", "Şubat", "Mart", "Nisan", "Mayıs", "Haziran",
    "Temmuz", "Ağustos", "Eylül", "Ekim", "Kasım", "Aralık",
]

# Dönem sıralaması (ay adı → sıra no)
AY_SIRA = {ay: i for i, ay in enumerate(AY_ISIMLERI)}

TABLO_KOLONLARI = [
    "Kimlik No", "Adı Soyadı", "Yıl", "Dönem",
    "Top. Gün", "Top. İzin", "Fiili Saat",
    "Kümülatif Saat", "Hak Edilen Şua (Gün)",
]
C_KIMLIK, C_AD, C_YIL, C_DONEM = 0, 1, 2, 3
C_GUN, C_IZIN, C_SAAT, C_KUM, C_SUA = 4, 5, 6, 7, 8


# ═══════════════════════════════════════════════
#  DELEGATE: Şua Hak Ediş badge (Kolon 8)
# ═══════════════════════════════════════════════

class SuaDelegate(QStyledItemDelegate):
    """Şua hak ediş günlerini renkli badge olarak gösterir."""
    def paint(self, painter, option, index):
        bg = QColor(29, 117, 254, 60) if option.state & QStyle.StateFlag.State_Selected \
            else QColor("transparent")
        painter.fillRect(option.rect, bg)

        try:
            deger = float(index.data(Qt.ItemDataRole.DisplayRole))
        except (ValueError, TypeError):
            deger = 0

        painter.save()
        painter.setRenderHint(QPainter.RenderHint.Antialiasing)
        rect = QRectF(option.rect)
        rect.adjust(8, 4, -8, -4)

        if deger >= 20:
            c_bg, c_border, c_text = QColor(27, 94, 32, 160), QColor("#66bb6a"), QColor("#ffffff")
        elif deger >= 10:
            c_bg, c_border, c_text = QColor(133, 100, 0, 140), QColor(Colors.YELLOW_400), QColor("#ffffff")
        elif deger > 0:
            c_bg, c_border, c_text = QColor(21, 101, 192, 120), QColor("#42a5f5"), QColor("#ffffff")
        else:
            c_bg, c_border, c_text = QColor(62, 62, 62, 80), QColor("#555"), QColor("#aaaaaa")

        path = QPainterPath()
        path.addRoundedRect(rect, 4, 4)
        painter.setBrush(QBrush(c_bg))
        painter.setPen(QPen(c_border, 1))
        painter.drawPath(path)
        painter.setPen(c_text)
        painter.setFont(QFont("Segoe UI", 9, QFont.Weight.Bold))
        painter.drawText(rect, Qt.AlignmentFlag.AlignCenter, f"{deger:.0f}")
        painter.restore()


# ═══════════════════════════════════════════════
#  PUANTAJ RAPOR SAYFASI
# ═══════════════════════════════════════════════

class PuantajRaporPage(QWidget):

    def __init__(self, db=None, parent=None):
        super().__init__(parent)
        self.setProperty("bg-role", "page")
        self.style().unpolish(self)
        self.style().polish(self)
        self._db = db
        self._rapor_data = []     # Tablodaki satırlar (dict list)

        self._setup_ui()
        self._connect_signals()

    # ═══════════════════════════════════════════
    #  UI
    # ═══════════════════════════════════════════

    def _setup_ui(self):
        main = QVBoxLayout(self)
        main.setContentsMargins(20, 12, 20, 12)
        main.setSpacing(10)

        # ── ÜST BAR: Rapor Filtreleri ──
        filter_frame = QFrame()
        filter_frame.setProperty("bg-role", "panel")
        filter_frame.style().unpolish(filter_frame)
        filter_frame.style().polish(filter_frame)
        fp = QHBoxLayout(filter_frame)
        fp.setContentsMargins(12, 8, 12, 8)
        fp.setSpacing(8)

        # Yıl
        fp.addWidget(self._make_label("Rapor Yılı:"))
        self.cmb_yil = QComboBox()
        # self.cmb_yil.setStyleSheet kaldırıldı (global QSS yönetir)
        self.cmb_yil.setFixedWidth(100)
        by = datetime.now().year
        for y in range(by - 5, by + 5):
            self.cmb_yil.addItem(str(y))
        self.cmb_yil.setCurrentText(str(by))
        fp.addWidget(self.cmb_yil)

        # Dönem / Ay
        fp.addWidget(self._make_label("Dönem:"))
        self.cmb_donem = QComboBox()
        # self.cmb_donem.setStyleSheet kaldırıldı (global QSS yönetir)
        self.cmb_donem.setFixedWidth(130)
        self.cmb_donem.addItem("Tümü")
        self.cmb_donem.addItems(AY_ISIMLERI)
        fp.addWidget(self.cmb_donem)

        fp.addStretch()

        self.btn_getir = QPushButton("Raporu Olustur")
        self.btn_getir.setProperty("style-role", "action")
        self.btn_getir.style().unpolish(self.btn_getir)
        self.btn_getir.style().polish(self.btn_getir)
        self.btn_getir.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        IconRenderer.set_button_icon(self.btn_getir, "clipboard_list", color=DarkTheme.TEXT_PRIMARY, size=14)
        fp.addWidget(self.btn_getir)

        main.addWidget(filter_frame)

        # ── BİLGİ LABEL ──
        self.lbl_bilgi = QLabel("Veri bekleniyor...")
        self.lbl_bilgi.setProperty("color-role", "muted")
        self.lbl_bilgi.style().unpolish(self.lbl_bilgi)
        self.lbl_bilgi.style().polish(self.lbl_bilgi)
        self.lbl_bilgi.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
        main.addWidget(self.lbl_bilgi)

        # ── TABLO ──
        self.tablo = QTableWidget()
        self.tablo.setColumnCount(len(TABLO_KOLONLARI))
        self.tablo.setHorizontalHeaderLabels(TABLO_KOLONLARI)
        self.tablo.verticalHeader().setVisible(False)
        self.tablo.setAlternatingRowColors(True)
        self.tablo.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.tablo.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        # self.tablo.setStyleSheet kaldırıldı (global QSS yönetir)

        h = self.tablo.horizontalHeader()
        h.setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        h.setSectionResizeMode(C_KIMLIK, QHeaderView.ResizeMode.ResizeToContents)
        h.setSectionResizeMode(C_SUA, QHeaderView.ResizeMode.Fixed)
        self.tablo.setColumnWidth(C_SUA, 150)

        # Şua delegate
        self.tablo.setItemDelegateForColumn(C_SUA, SuaDelegate(self.tablo))

        main.addWidget(self.tablo, 1)

        # ── ALT BAR ──
        bot_frame = QFrame()
        bot_frame.setProperty("bg-role", "panel")
        bot_frame.style().unpolish(bot_frame)
        bot_frame.style().polish(bot_frame)
        bf = QHBoxLayout(bot_frame)
        bf.setContentsMargins(12, 8, 12, 8)
        bf.setSpacing(12)

        self.lbl_durum = QLabel("Hazır")
        self.lbl_durum.setProperty("color-role", "muted")
        self.lbl_durum.style().unpolish(self.lbl_durum)
        self.lbl_durum.style().polish(self.lbl_durum)
        bf.addWidget(self.lbl_durum)

        bf.addStretch()

        self.progress = QProgressBar()
        self.progress.setVisible(False)
        self.progress.setRange(0, 0)
        self.progress.setFixedWidth(160)
        self.progress.setFixedHeight(14)
        self.progress.setStyleSheet("""
            QProgressBar {{
                background-color: rgba(255,255,255,0.05);
                border: 1px solid rgba(255,255,255,0.08);
                border-radius: 3px; font-size: 10px; color: {};
            }}
            QProgressBar::chunk {{
                background-color: rgba(29,117,254,0.5); border-radius: 2px;
            }}
        """.format(DarkTheme.TEXT_MUTED))
        bf.addWidget(self.progress)


        self.btn_excel = QPushButton("Excel Indir")
        self.btn_excel.setProperty("style-role", "success-filled")
        self.btn_excel.style().unpolish(self.btn_excel)
        self.btn_excel.style().polish(self.btn_excel)
        self.btn_excel.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self.btn_excel.setEnabled(False)
        IconRenderer.set_button_icon(self.btn_excel, "download", color=DarkTheme.TEXT_PRIMARY, size=14)
        bf.addWidget(self.btn_excel)

        self.btn_pdf = QPushButton("PDF Indir")
        self.btn_pdf.setProperty("style-role", "danger")
        self.btn_pdf.style().unpolish(self.btn_pdf)
        self.btn_pdf.style().polish(self.btn_pdf)
        self.btn_pdf.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self.btn_pdf.setEnabled(False)
        IconRenderer.set_button_icon(self.btn_pdf, "save", color=DarkTheme.TEXT_PRIMARY, size=14)
        bf.addWidget(self.btn_pdf)

        main.addWidget(bot_frame)

    # ─── UI yardımcıları ───

    def _make_label(self, text):
        lbl = QLabel(text)
        lbl.setProperty("color-role", "muted")
        lbl.style().unpolish(lbl)
        lbl.style().polish(lbl)
        return lbl

    def _add_sep(self, layout):
        sep = QFrame()
        sep.setFixedWidth(1); sep.setFixedHeight(20)
        sep.setProperty("bg-role", "separator")
        sep.style().unpolish(sep)
        sep.style().polish(sep)
        layout.addWidget(sep)

    # ═══════════════════════════════════════════
    #  SİNYALLER
    # ═══════════════════════════════════════════

    def _connect_signals(self):
        self.btn_getir.clicked.connect(self._rapor_olustur)
        self.btn_excel.clicked.connect(self._excel_indir)
        self.btn_pdf.clicked.connect(self._pdf_indir)

    # ═══════════════════════════════════════════
    #  VERİ YÜKLEME (boş — sayfa ilk açıldığında)
    # ═══════════════════════════════════════════

    def load_data(self):
        """Sayfa açılışında çağrılır — henüz rapor oluşturma yok."""
        pass

    # ═══════════════════════════════════════════
    #  📋 RAPOR OLUŞTUR
    # ═══════════════════════════════════════════

    def _rapor_olustur(self):
        """
        FHSZ_Puantaj tablosundan seçili yıl + dönem aralığı verilerini çek.
        Her personel için kümülatif saat ve şua hak ediş hesapla.
        """
        if not self._db:
            return

        yil_str = self.cmb_yil.currentText()
        donem_str = self.cmb_donem.currentText()   # "Tümü" veya ay adı
        tek_donem = donem_str != "Tümü"

        self.tablo.setRowCount(0)
        self._rapor_data = []
        self.btn_getir.setEnabled(False)
        self.progress.setVisible(True)
        self.lbl_durum.setText("Rapor hazırlanıyor...")

        try:
            from core.di import get_fhsz_service
            fhsz_svc = get_fhsz_service(self._db)
            
            tum = fhsz_svc.get_puantaj_listesi().veri or []

            # Yıla ait kayıtları filtrele
            yil_kayitlar = [
                r for r in tum
                if str(r.get("AitYil", "")).strip() == yil_str
            ]

            if not yil_kayitlar:
                self.lbl_bilgi.setText("Kayıt bulunamadı.")
                self.lbl_durum.setText("Hazır")
                self.progress.setVisible(False)
                self.btn_getir.setEnabled(True)
                self.btn_excel.setEnabled(False)
                self.btn_pdf.setEnabled(False)
                return

            # ── Kümülatif hesaplama ──
            # Personel bazlı dönem sıralı grupla
            personel_map = {}   # {tc: [kayıt, kayıt, ...]}  (dönem sıralı)
            for r in yil_kayitlar:
                tc = str(r.get("Personelid", "")).strip()
                if tc not in personel_map:
                    personel_map[tc] = []
                personel_map[tc].append(r)

            # Her personelin kayıtlarını dönem sırasına göre sırala
            for tc in personel_map:
                personel_map[tc].sort(
                    key=lambda r: AY_SIRA.get(str(r.get("Donem", "")).strip(), 99)
                )

            # ── Tabloya doldur ──
            rows = []
            for tc, kayitlar in sorted(personel_map.items(),
                                        key=lambda x: str(x[1][0].get("AdSoyad", ""))):
                kumulatif = 0
                toplam_gun = 0
                toplam_izin = 0
                toplam_saat = 0
                for r in kayitlar:
                    donem = str(r.get("Donem", "")).strip()
                    try:
                        saat = float(str(r.get("FiiliCalismaSaat", 0)).replace(",", "."))
                    except (ValueError, TypeError):
                        saat = 0
                    kumulatif += saat
                    toplam_saat += saat
                    try:
                        toplam_gun += int(r.get("AylikGun", 0))
                    except (ValueError, TypeError):
                        toplam_gun += 0
                    try:
                        toplam_izin += int(r.get("KullanilanIzin", 0))
                    except (ValueError, TypeError):
                        toplam_izin += 0

                    # Tek dönem filtresi
                    if tek_donem and donem != donem_str:
                        continue

                    sua = sua_hak_edis_hesapla(kumulatif)

                    if not tek_donem:
                        continue

                    row = {
                        "Personelid": tc,
                        "AdSoyad": r.get("AdSoyad", ""),
                        "AitYil": yil_str,
                        "Donem": donem,
                        "AylikGun": r.get("AylikGun", 0),
                        "KullanilanIzin": r.get("KullanilanIzin", 0),
                        "FiiliCalismaSaat": saat,
                        "KumulatifSaat": kumulatif,
                        "SuaHakEdis": sua,
                    }
                    rows.append(row)

                # Tum donem secilirse personel toplam satiri ekle
                if not tek_donem and kayitlar:
                    toplam_sua = sua_hak_edis_hesapla(toplam_saat)
                    rows.append({
                        "Personelid": tc,
                        "AdSoyad": kayitlar[0].get("AdSoyad", ""),
                        "AitYil": yil_str,
                        "Donem": "Toplam",
                        "AylikGun": toplam_gun,
                        "KullanilanIzin": toplam_izin,
                        "FiiliCalismaSaat": toplam_saat,
                        "KumulatifSaat": toplam_saat,
                        "SuaHakEdis": toplam_sua,
                    })



            # Tabloya yaz
            self.tablo.setRowCount(len(rows))
            for i, row in enumerate(rows):
                self._set_item(i, C_KIMLIK, row["Personelid"])
                self._set_item(i, C_AD, row["AdSoyad"])
                self._set_item(i, C_YIL, row["AitYil"])
                self._set_item(i, C_DONEM, row["Donem"])
                self._set_item(i, C_GUN, str(row["AylikGun"]))
                self._set_item(i, C_IZIN, str(row["KullanilanIzin"]))
                self._set_item(i, C_SAAT, f"{row['FiiliCalismaSaat']:.0f}")
                self._set_item(i, C_KUM, f"{row['KumulatifSaat']:.0f}")
                # Şua hak ediş — delegate çizecek
                item_sua = QTableWidgetItem(str(row["SuaHakEdis"]))
                item_sua.setFlags(Qt.ItemFlag.ItemIsSelectable | Qt.ItemFlag.ItemIsEnabled)
                item_sua.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                self.tablo.setItem(i, C_SUA, item_sua)

            self._rapor_data = rows

            # Bilgi label
            personel_sayisi = len(personel_map)
            donem_info = donem_str if tek_donem else "Tüm dönemler"
            self.lbl_bilgi.setText(
                f"{personel_sayisi} personel  •  {len(rows)} kayıt  •  {donem_info}  •  {yil_str}"
            )
            self.lbl_durum.setText(f"Rapor hazir - {len(rows)} satir")
            self.btn_excel.setEnabled(len(rows) > 0)
            self.btn_pdf.setEnabled(len(rows) > 0)

            logger.info(f"Puantaj rapor oluşturuldu: {yil_str}, {len(rows)} satır")

        except Exception as e:
            logger.error(f"Puantaj rapor hatası: {e}")
            MesajKutusu.hata(self, str(e))
            self.lbl_bilgi.setText("Hata oluştu.")

        self.progress.setVisible(False)
        self.btn_getir.setEnabled(True)

    def _set_item(self, row, col, text):
        item = QTableWidgetItem(str(text))
        item.setFlags(Qt.ItemFlag.ItemIsSelectable | Qt.ItemFlag.ItemIsEnabled)
        if col >= C_GUN:
            item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
        self.tablo.setItem(row, col, item)

    # ═══════════════════════════════════════════
    #  📥 EXCEL İNDİR
    # ═══════════════════════════════════════════

    def _excel_indir(self):
        if not self._rapor_data:
            return

        yil = self.cmb_yil.currentText()
        donem = self.cmb_donem.currentText().replace(" ", "_")
        default_name = f"FHSZ_Puantaj_Rapor_{yil}_{donem}.xlsx"

        path, _ = QFileDialog.getSaveFileName(
            self, "Excel Kaydet", default_name,
            "Excel Dosyası (*.xlsx)"
        )
        if not path:
            return

        try:
            import openpyxl
            from openpyxl.utils.cell import get_column_letter
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

            wb = openpyxl.Workbook()
            ws = wb.active
            if ws is None:
                raise RuntimeError("Excel çalışma sayfası oluşturulamadı")
            ws.title = "Puantaj Rapor"

            # Başlık satırı
            header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
            header_fill = PatternFill(start_color="1D75FE", end_color="1D75FE", fill_type="solid")
            header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
            thin_border = Border(
                left=Side(style="thin"), right=Side(style="thin"),
                top=Side(style="thin"), bottom=Side(style="thin")
            )

            for c, header in enumerate(TABLO_KOLONLARI, 1):
                cell = ws.cell(row=1, column=c, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
                cell.border = thin_border

            # Veri satırları
            data_font = Font(name="Arial", size=10)
            center_align = Alignment(horizontal="center", vertical="center")
            left_align = Alignment(horizontal="left", vertical="center")

            # Şua renk dolguları
            sua_fill_high = PatternFill(start_color="1B5E20", end_color="1B5E20", fill_type="solid")
            sua_fill_mid = PatternFill(start_color="F57F17", end_color="F57F17", fill_type="solid")
            sua_fill_low = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
            sua_font_white = Font(name="Arial", bold=True, color="FFFFFF", size=10)

            for r, row in enumerate(self._rapor_data, 2):
                vals = [
                    row["Personelid"], row["AdSoyad"], row["AitYil"], row["Donem"],
                    int(row["AylikGun"]), int(row["KullanilanIzin"]),
                    int(row["FiiliCalismaSaat"]), int(row["KumulatifSaat"]),
                    int(row["SuaHakEdis"]),
                ]
                for c, val in enumerate(vals, 1):
                    cell = ws.cell(row=r, column=c, value=val)
                    cell.font = data_font
                    cell.border = thin_border
                    cell.alignment = center_align if c >= 3 else left_align

                # Şua hücre renklendirme
                sua_cell = ws.cell(row=r, column=9)
                sua_val = int(row["SuaHakEdis"])
                if sua_val >= 20:
                    sua_cell.fill = sua_fill_high
                    sua_cell.font = sua_font_white
                elif sua_val >= 10:
                    sua_cell.fill = sua_fill_mid
                    sua_cell.font = sua_font_white
                elif sua_val > 0:
                    sua_cell.fill = sua_fill_low
                    sua_cell.font = sua_font_white

            # Kolon genişlikleri
            widths = [14, 25, 8, 12, 10, 10, 14, 16, 18]
            for i, w in enumerate(widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = w

            # Filtreler
            ws.auto_filter.ref = ws.dimensions

            wb.save(path)

            self.lbl_durum.setText(f"Excel kaydedildi: {os.path.basename(path)}")
            MesajKutusu.bilgi(self, f"Excel dosyası kaydedildi:\n{path}")
            logger.info(f"Puantaj rapor Excel: {path}")

        except ImportError:
            MesajKutusu.uyari(self, "openpyxl modülü yüklü değil.\npip install openpyxl")
        except Exception as e:
            logger.error(f"Excel kaydetme hatası: {e}")
            MesajKutusu.hata(self, f"Excel kaydedilemedi:\n{e}")

    # ═══════════════════════════════════════════
    #  📄 PDF İNDİR
    # ═══════════════════════════════════════════

    def _pdf_indir(self):
        if not self._rapor_data:
            return

        yil = self.cmb_yil.currentText()
        donem = self.cmb_donem.currentText().replace(" ", "_")
        default_name = f"FHSZ_Puantaj_Rapor_{yil}_{donem}.pdf"

        path, _ = QFileDialog.getSaveFileName(
            self, "PDF Kaydet", default_name,
            "PDF Dosyası (*.pdf)"
        )
        if not path:
            return

        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.lib.styles import getSampleStyleSheet
            from reportlab.lib.units import mm
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont

            # Türkçe font
            try:
                pdfmetrics.registerFont(TTFont("Arial", "arial.ttf"))
                pdfmetrics.registerFont(TTFont("ArialBold", "arialbd.ttf"))
                font_name, font_bold = "Arial", "ArialBold"
            except Exception:
                font_name, font_bold = "Helvetica", "Helvetica-Bold"

            doc = SimpleDocTemplate(
                path, pagesize=landscape(A4),
                leftMargin=15 * mm, rightMargin=15 * mm,
                topMargin=15 * mm, bottomMargin=15 * mm
            )

            elements = []
            styles = getSampleStyleSheet()

            # Başlık
            title_text = f"FHSZ Puantaj Raporu — {yil} ({donem})"
            title_style = styles["Title"]
            title_style.fontName = font_bold
            title_style.fontSize = 14
            elements.append(Paragraph(title_text, title_style))
            elements.append(Spacer(1, 5 * mm))

            # Tablo verileri
            header = TABLO_KOLONLARI[:]
            table_data = [header]

            for row in self._rapor_data:
                table_data.append([
                    str(row["Personelid"]),
                    str(row["AdSoyad"]),
                    str(row["AitYil"]),
                    str(row["Donem"]),
                    str(int(row["AylikGun"])),
                    str(int(row["KullanilanIzin"])),
                    str(int(row["FiiliCalismaSaat"])),
                    str(int(row["KumulatifSaat"])),
                    str(int(row["SuaHakEdis"])),
                ])

            col_widths = [55, 100, 35, 50, 40, 40, 55, 60, 70]
            t = Table(table_data, colWidths=col_widths, repeatRows=1)

            style_cmds = [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1D75FE")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), font_bold),
                ("FONTSIZE", (0, 0), (-1, 0), 8),
                ("FONTNAME", (0, 1), (-1, -1), font_name),
                ("FONTSIZE", (0, 1), (-1, -1), 7),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("ALIGN", (1, 1), (1, -1), "LEFT"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F0F4F8")]),
            ]

            # Şua renklendirme
            for i, row in enumerate(self._rapor_data, 1):
                sua = int(row["SuaHakEdis"])
                if sua >= 20:
                    style_cmds.append(("BACKGROUND", (8, i), (8, i), colors.HexColor("#1B5E20")))
                    style_cmds.append(("TEXTCOLOR", (8, i), (8, i), colors.white))
                elif sua >= 10:
                    style_cmds.append(("BACKGROUND", (8, i), (8, i), colors.HexColor("#F57F17")))
                    style_cmds.append(("TEXTCOLOR", (8, i), (8, i), colors.white))
                elif sua > 0:
                    style_cmds.append(("BACKGROUND", (8, i), (8, i), colors.HexColor("#1565C0")))
                    style_cmds.append(("TEXTCOLOR", (8, i), (8, i), colors.white))

            t.setStyle(TableStyle(style_cmds))
            elements.append(t)

            # Alt bilgi
            elements.append(Spacer(1, 5 * mm))
            info_text = f"Rapor tarihi: {datetime.now().strftime('%d.%m.%Y %H:%M')} — Toplam {len(self._rapor_data)} kayıt"
            info_style = styles["Normal"]
            info_style.fontName = font_name
            info_style.fontSize = 8
            info_style.textColor = colors.grey
            elements.append(Paragraph(info_text, info_style))

            doc.build(elements)

            self.lbl_durum.setText(f"PDF kaydedildi: {os.path.basename(path)}")
            MesajKutusu.bilgi(self, f"PDF dosyası kaydedildi:\n{path}")
            logger.info(f"Puantaj rapor PDF: {path}")

        except ImportError:
            MesajKutusu.uyari(self, "reportlab modülü yüklü değil.\npip install reportlab")
        except Exception as e:
            logger.error(f"PDF kaydetme hatası: {e}")
            MesajKutusu.hata(self, f"PDF kaydedilemedi:\n{e}")


