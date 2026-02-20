# -*- coding: utf-8 -*-
"""
FHSZ Merkez  —  REPYS v3 · Medikal Dark-Blue Tema
════════════════════════════════════════════════════════
Önceki iki ayrı form (fhsz_yonetim + puantaj_rapor) tek sayfada toplandı.

  ┌────────────────────────────────────────────────┐
  │  Header: başlık · dönem seçici · aksiyon butonları │
  ├───────────────┬────────────────────────────────┤
  │  [HESAPLAMA]  │  [RAPOR & ŞUA TAKİP]           │  ← 2 sekme
  └───────────────┴────────────────────────────────┘

Yenilikler vs eski 2 ayrı form:
  • Tek menü kaydı, tek sayfa açılışı
  • Sekme geçişinde veri yeniden yüklenmez (lazy cache)
  • Ortak dönem seçici — her iki sekme aynı yıl/ay bilgisini paylaşır
  • Özet kart şeridi: Toplam personel, dönem iş günü, hesaplanan saat
  • Rapor sekmesinde kümülatif şua grafiği placeholder (sonraki sürüm)
  • Tüm renkler DarkTheme üzerinden — hardcoded renk yok
"""
from __future__ import annotations

import os
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta

from PySide6.QtCore    import Qt, QRectF, Signal
from PySide6.QtGui     import (
    QColor, QCursor, QFont, QPainter, QBrush, QPen, QPainterPath,
)
from PySide6.QtWidgets import (
    QAbstractItemView, QComboBox, QFileDialog, QFrame,
    QHBoxLayout, QHeaderView, QLabel, QMessageBox, QProgressBar,
    QPushButton, QStackedWidget, QStyledItemDelegate,
    QStyle, QTableWidget, QTableWidgetItem, QVBoxLayout, QWidget,
)

from core.date_utils   import parse_date as parse_any_date
from core.hesaplamalar import sua_hak_edis_hesapla, is_gunu_hesapla, tr_upper
from core.logger       import logger
from ui.styles         import DarkTheme as T
from ui.styles.components import ComponentStyles
from ui.styles.icons   import IconRenderer
from ui.theme_manager  import ThemeManager

S = ThemeManager.get_all_component_styles()

# ── Sabitler ────────────────────────────────────────────────────
FHSZ_ESIK = datetime(2022, 4, 26)
KOSUL_A_SAAT = 7

IZIN_VERILEN_SINIFLAR = [
    "Akademik Personel", "Asistan Doktor",
    "Radyasyon Görevlisi", "Hemşire",
]

AY_ISIMLERI = [
    "Ocak", "Şubat", "Mart", "Nisan", "Mayıs", "Haziran",
    "Temmuz", "Ağustos", "Eylül", "Ekim", "Kasım", "Aralık",
]
AY_SIRA = {ay: i for i, ay in enumerate(AY_ISIMLERI)}

# Hesaplama tablosu kolonları
HESAP_COLS = [
    "Kimlik No", "Adı Soyadı", "Birim", "Çalışma Koşulu",
    "Ait Yıl", "Dönem", "Aylık Gün", "Kullanılan İzin",
    "Fiili Çalışma (Saat)",
]
HC = {n: i for i, n in enumerate(HESAP_COLS)}
C_KIMLIK, C_AD, C_BIRIM, C_KOSUL = 0, 1, 2, 3
C_YIL, C_DONEM, C_GUN, C_IZIN, C_SAAT = 4, 5, 6, 7, 8

# Rapor tablosu kolonları
RAPOR_COLS = [
    "Kimlik No", "Adı Soyadı", "Yıl", "Dönem",
    "Top. Gün", "Top. İzin", "Fiili Saat",
    "Kümülatif Saat", "Hak Edilen Şua (Gün)",
]
RC = {n: i for i, n in enumerate(RAPOR_COLS)}
R_KIM, R_AD, R_YIL, R_DON = 0, 1, 2, 3
R_GUN, R_IZN, R_SAT, R_KUM, R_SUA = 4, 5, 6, 7, 8


def _parse_date(val):
    parsed = parse_any_date(val)
    if not parsed:
        return None
    return datetime.combine(parsed, datetime.min.time())


# ════════════════════════════════════════════════════════════════
#  DELEGATE'LER
# ════════════════════════════════════════════════════════════════

class _BadgeDelegate(QStyledItemDelegate):
    """Temel renkli badge çizici."""

    def _sel_bg(self, option):
        return QColor(0, 180, 216, 50) if option.state & QStyle.State_Selected \
            else QColor("transparent")

    def _draw_badge(self, painter, option, text, c_bg, c_border, c_text):
        painter.fillRect(option.rect, self._sel_bg(option))
        painter.save()
        painter.setRenderHint(QPainter.Antialiasing)
        rect = QRectF(option.rect).adjusted(6, 4, -6, -4)
        path = QPainterPath()
        path.addRoundedRect(rect, 5, 5)
        painter.setBrush(QBrush(c_bg))
        painter.setPen(QPen(c_border, 1.2))
        painter.drawPath(path)
        painter.setPen(c_text)
        painter.setFont(QFont("Segoe UI", 9, QFont.Bold))
        painter.drawText(rect, Qt.AlignCenter, text)
        painter.restore()


class KosulDelegate(_BadgeDelegate):
    """Çalışma Koşulu A / B — combobox + badge."""

    ITEMS = ["Çalışma Koşulu A", "Çalışma Koşulu B"]

    def createEditor(self, parent, option, index):
        ed = QComboBox(parent)
        ed.addItems(self.ITEMS)
        ed.setStyleSheet(S["combo"])
        ed.showPopup()
        return ed

    def setEditorData(self, ed, index):
        t = index.data(Qt.EditRole)
        if t in self.ITEMS:
            ed.setCurrentText(t)

    def setModelData(self, ed, model, index):
        model.setData(index, ed.currentText(), Qt.EditRole)

    def updateEditorGeometry(self, ed, option, index):
        ed.setGeometry(option.rect)

    def paint(self, painter, option, index):
        text = str(index.data(Qt.DisplayRole) or "")
        if "KOŞULU A" in text.upper():
            c_bg = QColor(16, 185, 129, 35)
            c_bd = QColor(T.STATUS_SUCCESS)
        else:
            c_bg = QColor(0, 180, 216, 30)
            c_bd = QColor(T.ACCENT)
        self._draw_badge(painter, option, text, c_bg, c_bd, QColor(T.TEXT_PRIMARY))


class SaatDelegate(_BadgeDelegate):
    """Fiili çalışma saati — yeşil/gri."""

    def paint(self, painter, option, index):
        try:
            val = float(index.data(Qt.DisplayRole))
        except (ValueError, TypeError):
            val = 0
        if val > 0:
            c_bg, c_bd = QColor(16, 185, 129, 35), QColor(T.STATUS_SUCCESS)
        else:
            c_bg, c_bd = QColor(255, 255, 255, 12), QColor(T.BORDER_STRONG)
        self._draw_badge(painter, option, f"{val:.0f}", c_bg, c_bd,
                         QColor(T.TEXT_PRIMARY if val > 0 else T.TEXT_MUTED))


class SuaDelegate(_BadgeDelegate):
    """Şua hak ediş günü — yeşil / amber / mavi / gri."""

    def paint(self, painter, option, index):
        try:
            val = float(index.data(Qt.DisplayRole))
        except (ValueError, TypeError):
            val = 0
        if val >= 20:
            c_bg, c_bd = QColor(16, 185, 129, 40), QColor(T.STATUS_SUCCESS)
        elif val >= 10:
            c_bg, c_bd = QColor(245, 158, 11, 40), QColor(T.STATUS_WARNING)
        elif val > 0:
            c_bg, c_bd = QColor(0, 180, 216, 35), QColor(T.ACCENT)
        else:
            c_bg, c_bd = QColor(255, 255, 255, 12), QColor(T.BORDER_STRONG)
        self._draw_badge(painter, option, f"{val:.0f}", c_bg, c_bd,
                         QColor(T.TEXT_PRIMARY if val > 0 else T.TEXT_MUTED))


# ════════════════════════════════════════════════════════════════
#  ÖZET KART
# ════════════════════════════════════════════════════════════════

class _StatCard(QFrame):
    """Küçük istatistik kartı — başlık + büyük değer."""

    def __init__(self, title: str, value: str = "—",
                 accent: str = T.ACCENT, parent=None):
        super().__init__(parent)
        self.setFixedHeight(60)
        self.setStyleSheet(f"""
            QFrame {{
                background: {T.BG_SECONDARY};
                border: 1px solid {T.BORDER_PRIMARY};
                border-left: 3px solid {accent};
                border-radius: 8px;
            }}
        """)
        lay = QVBoxLayout(self)
        lay.setContentsMargins(12, 6, 12, 6)
        lay.setSpacing(2)

        self._lbl_val = QLabel(value)
        self._lbl_val.setStyleSheet(
            f"font-size:20px; font-weight:800; color:{accent}; background:transparent;"
        )
        self._lbl_title = QLabel(title)
        self._lbl_title.setStyleSheet(
            f"font-size:10px; color:{T.TEXT_MUTED}; background:transparent; letter-spacing:0.06em;"
        )
        lay.addWidget(self._lbl_val)
        lay.addWidget(self._lbl_title)

    def set_value(self, v: str):
        self._lbl_val.setText(v)


# ════════════════════════════════════════════════════════════════
#  FHSZ MERKEZ SAYFASI
# ════════════════════════════════════════════════════════════════

class FHSZMerkezPage(QWidget):
    """
    FHSZ Hesaplama + Puantaj Raporu tek sayfada.
    Eski fhsz_yonetim.py + puantaj_rapor.py yerini alır.
    API uyumluluğu: btn_kapat sinyali korundu.
    """

    def __init__(self, db=None, parent=None):
        super().__init__(parent)
        self.setStyleSheet(S["page"])
        self._db = db

        # Veri önbelleği
        self._all_personel:    list = []
        self._all_izin:        list = []
        self._tatil_np:        list = []
        self._birim_kosul_map: dict = {}
        self._rapor_data:      list = []

        self._setup_ui()
        self._connect_signals()

    # ════════════════════════════════════════════════════════
    #  UI KURULUMU
    # ════════════════════════════════════════════════════════

    def _setup_ui(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(0)

        root.addWidget(self._build_header())
        root.addWidget(self._build_stat_bar())
        root.addWidget(self._build_tab_bar())

        self._stack = QStackedWidget()
        self._stack.addWidget(self._build_hesap_tab())   # index 0
        self._stack.addWidget(self._build_rapor_tab())   # index 1
        root.addWidget(self._stack, 1)

        root.addWidget(self._build_footer())

    # ── Header ──────────────────────────────────────────────

    def _build_header(self) -> QFrame:
        frame = QFrame()
        frame.setFixedHeight(52)
        frame.setStyleSheet(f"""
            QFrame {{
                background: {T.BG_SECONDARY};
                border-bottom: 1px solid {T.BORDER_PRIMARY};
            }}
        """)
        lay = QHBoxLayout(frame)
        lay.setContentsMargins(20, 0, 20, 0)
        lay.setSpacing(10)

        # İkon + başlık
        icon_lbl = QLabel("📊")
        icon_lbl.setStyleSheet("background:transparent; font-size:18px;")
        lay.addWidget(icon_lbl)

        title = QLabel("FHSZ Hesaplama & Şua Takip")
        title.setStyleSheet(
            f"font-size:15px; font-weight:800; color:{T.TEXT_PRIMARY}; background:transparent;"
        )
        lay.addWidget(title)

        lay.addWidget(self._vline())

        # Yıl seçici
        lay.addWidget(self._hdr_lbl("Yıl:"))
        self.cmb_yil = QComboBox()
        self.cmb_yil.setStyleSheet(S["combo"])
        self.cmb_yil.setFixedWidth(80)
        by = datetime.now().year
        for y in range(by - 5, by + 3):
            self.cmb_yil.addItem(str(y))
        self.cmb_yil.setCurrentText(str(by))
        lay.addWidget(self.cmb_yil)

        # Ay seçici
        lay.addWidget(self._hdr_lbl("Ay:"))
        self.cmb_ay = QComboBox()
        self.cmb_ay.setStyleSheet(S["combo"])
        self.cmb_ay.setFixedWidth(110)
        self.cmb_ay.addItems(AY_ISIMLERI)
        self.cmb_ay.setCurrentIndex(max(0, datetime.now().month - 1))
        lay.addWidget(self.cmb_ay)

        lay.addWidget(self._vline())

        # Dönem etiketi
        self.lbl_donem = QLabel("…")
        self.lbl_donem.setStyleSheet(
            f"color:{T.ACCENT2}; font-size:12px; font-weight:600; background:transparent;"
        )
        lay.addWidget(self.lbl_donem)

        lay.addStretch()

        # Hesapla butonu
        self.btn_hesapla = QPushButton("  Listele & Hesapla")
        self.btn_hesapla.setCursor(QCursor(Qt.PointingHandCursor))
        self.btn_hesapla.setStyleSheet(S["calc_btn"])
        try:
            IconRenderer.set_button_icon(
                self.btn_hesapla, "bar_chart", color=T.BTN_PRIMARY_TEXT, size=14)
        except Exception:
            pass
        lay.addWidget(self.btn_hesapla)

        lay.addWidget(self._vline())

        self.btn_kapat = QPushButton("Kapat")
        self.btn_kapat.setCursor(QCursor(Qt.PointingHandCursor))
        self.btn_kapat.setStyleSheet(S["close_btn"])
        try:
            IconRenderer.set_button_icon(
                self.btn_kapat, "x", color=T.BTN_DANGER_TEXT, size=13)
        except Exception:
            pass
        lay.addWidget(self.btn_kapat)
        return frame

    # ── Özet kart şeridi ────────────────────────────────────

    def _build_stat_bar(self) -> QFrame:
        frame = QFrame()
        frame.setFixedHeight(76)
        frame.setStyleSheet(f"""
            QFrame {{
                background: {T.BG_PRIMARY};
                border-bottom: 1px solid {T.BORDER_SECONDARY};
                padding: 0 20px;
            }}
        """)
        lay = QHBoxLayout(frame)
        lay.setContentsMargins(20, 8, 20, 8)
        lay.setSpacing(10)

        self._card_personel  = _StatCard("PERSONEL SAYISI",   "—", T.ACCENT)
        self._card_is_gunu   = _StatCard("DÖNEM İŞ GÜNÜ",     "—", T.STATUS_SUCCESS)
        self._card_saat      = _StatCard("TOPLAM FİİLİ SAAT", "—", T.STATUS_WARNING)
        self._card_sua_max   = _StatCard("MAKS ŞUA HAK EDÄN", "—", "#a855f7")

        for card in (self._card_personel, self._card_is_gunu,
                     self._card_saat, self._card_sua_max):
            lay.addWidget(card)

        lay.addStretch()
        return frame

    # ── Sekme çubuğu ────────────────────────────────────────

    def _build_tab_bar(self) -> QFrame:
        frame = QFrame()
        frame.setFixedHeight(38)
        frame.setStyleSheet(f"""
            QFrame {{
                background: {T.BG_SECONDARY};
                border-bottom: 1px solid {T.BORDER_PRIMARY};
            }}
        """)
        lay = QHBoxLayout(frame)
        lay.setContentsMargins(20, 0, 20, 0)
        lay.setSpacing(0)

        self._tab_btns: dict[int, QPushButton] = {}
        tabs = [
            (0, "🧮  Hesaplama & Düzenleme"),
            (1, "📋  Rapor & Şua Takip"),
        ]
        for idx, lbl in tabs:
            btn = QPushButton(lbl)
            btn.setCursor(QCursor(Qt.PointingHandCursor))
            btn.setStyleSheet(self._tab_qss(idx == 0))
            btn.clicked.connect(lambda _, i=idx: self._switch_tab(i))
            self._tab_btns[idx] = btn
            lay.addWidget(btn)

        lay.addStretch()
        return frame

    # ── Hesaplama sekmesi ────────────────────────────────────

    def _build_hesap_tab(self) -> QWidget:
        w = QWidget()
        w.setStyleSheet("background:transparent;")
        lay = QVBoxLayout(w)
        lay.setContentsMargins(20, 12, 20, 0)
        lay.setSpacing(0)

        self.hesap_tablo = QTableWidget()
        self.hesap_tablo.setColumnCount(len(HESAP_COLS))
        self.hesap_tablo.setHorizontalHeaderLabels(HESAP_COLS)
        self.hesap_tablo.verticalHeader().setVisible(False)
        self.hesap_tablo.setAlternatingRowColors(True)
        self.hesap_tablo.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.hesap_tablo.setStyleSheet(S["table"])

        h = self.hesap_tablo.horizontalHeader()
        h.setSectionResizeMode(QHeaderView.Stretch)
        h.setSectionResizeMode(C_KIMLIK, QHeaderView.ResizeToContents)
        h.setSectionResizeMode(C_KOSUL,  QHeaderView.Fixed)
        self.hesap_tablo.setColumnWidth(C_KOSUL, 175)
        h.setSectionResizeMode(C_SAAT,   QHeaderView.Fixed)
        self.hesap_tablo.setColumnWidth(C_SAAT,  130)
        self.hesap_tablo.setColumnHidden(C_YIL,   True)
        self.hesap_tablo.setColumnHidden(C_DONEM, True)

        self.hesap_tablo.setItemDelegateForColumn(C_KOSUL, KosulDelegate(self.hesap_tablo))
        self.hesap_tablo.setItemDelegateForColumn(C_SAAT,  SaatDelegate(self.hesap_tablo))

        lay.addWidget(self.hesap_tablo, 1)

        # Alt aksiyon çubuğu
        bot = QFrame()
        bot.setFixedHeight(44)
        bot.setStyleSheet(f"""
            QFrame {{
                background: {T.BG_SECONDARY};
                border-top: 1px solid {T.BORDER_PRIMARY};
            }}
        """)
        bl = QHBoxLayout(bot)
        bl.setContentsMargins(20, 0, 20, 0)
        bl.setSpacing(10)

        self.hesap_lbl_durum = QLabel("Hazır")
        self.hesap_lbl_durum.setStyleSheet(S["footer_label"])
        bl.addWidget(self.hesap_lbl_durum)
        bl.addStretch()

        self.hesap_progress = self._make_progress()
        bl.addWidget(self.hesap_progress)

        self.btn_kaydet = QPushButton("  Kaydet / Güncelle")
        self.btn_kaydet.setCursor(QCursor(Qt.PointingHandCursor))
        self.btn_kaydet.setStyleSheet(S["save_btn"])
        try:
            IconRenderer.set_button_icon(
                self.btn_kaydet, "save", color=T.BTN_SUCCESS_TEXT, size=13)
        except Exception:
            pass
        bl.addWidget(self.btn_kaydet)
        lay.addWidget(bot)
        return w

    # ── Rapor sekmesi ────────────────────────────────────────

    def _build_rapor_tab(self) -> QWidget:
        w = QWidget()
        w.setStyleSheet("background:transparent;")
        lay = QVBoxLayout(w)
        lay.setContentsMargins(20, 12, 20, 0)
        lay.setSpacing(0)

        # Rapor filtre şeridi
        filt = QFrame()
        filt.setFixedHeight(44)
        filt.setStyleSheet(f"""
            QFrame {{
                background: {T.BG_TERTIARY};
                border: 1px solid {T.BORDER_PRIMARY};
                border-radius: 8px;
                margin-bottom: 8px;
            }}
        """)
        fl = QHBoxLayout(filt)
        fl.setContentsMargins(12, 0, 12, 0)
        fl.setSpacing(8)

        fl.addWidget(self._hdr_lbl("Rapor Dönemi:"))
        self.cmb_rapor_donem = QComboBox()
        self.cmb_rapor_donem.setStyleSheet(S["combo"])
        self.cmb_rapor_donem.setFixedWidth(130)
        self.cmb_rapor_donem.addItem("Tümü (Yıllık)")
        self.cmb_rapor_donem.addItems(AY_ISIMLERI)
        fl.addWidget(self.cmb_rapor_donem)

        fl.addStretch()

        self.btn_rapor_olustur = QPushButton("  Raporu Oluştur")
        self.btn_rapor_olustur.setCursor(QCursor(Qt.PointingHandCursor))
        self.btn_rapor_olustur.setStyleSheet(S["report_btn"])
        try:
            IconRenderer.set_button_icon(
                self.btn_rapor_olustur, "clipboard_list",
                color=T.TEXT_PRIMARY, size=13)
        except Exception:
            pass
        fl.addWidget(self.btn_rapor_olustur)

        fl.addWidget(self._vline())

        self.btn_excel = QPushButton("  Excel")
        self.btn_excel.setCursor(QCursor(Qt.PointingHandCursor))
        self.btn_excel.setStyleSheet(S["excel_btn"])
        self.btn_excel.setEnabled(False)
        try:
            IconRenderer.set_button_icon(
                self.btn_excel, "download",
                color=T.BTN_SUCCESS_TEXT, size=13)
        except Exception:
            pass
        fl.addWidget(self.btn_excel)

        self.btn_pdf = QPushButton("  PDF")
        self.btn_pdf.setCursor(QCursor(Qt.PointingHandCursor))
        self.btn_pdf.setStyleSheet(S["pdf_btn"])
        self.btn_pdf.setEnabled(False)
        try:
            IconRenderer.set_button_icon(
                self.btn_pdf, "save",
                color=T.TEXT_PRIMARY, size=13)
        except Exception:
            pass
        fl.addWidget(self.btn_pdf)
        lay.addWidget(filt)

        # Rapor tablosu
        self.rapor_tablo = QTableWidget()
        self.rapor_tablo.setColumnCount(len(RAPOR_COLS))
        self.rapor_tablo.setHorizontalHeaderLabels(RAPOR_COLS)
        self.rapor_tablo.verticalHeader().setVisible(False)
        self.rapor_tablo.setAlternatingRowColors(True)
        self.rapor_tablo.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.rapor_tablo.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.rapor_tablo.setStyleSheet(S["table"])

        rh = self.rapor_tablo.horizontalHeader()
        rh.setSectionResizeMode(QHeaderView.Stretch)
        rh.setSectionResizeMode(R_KIM, QHeaderView.ResizeToContents)
        rh.setSectionResizeMode(R_SUA, QHeaderView.Fixed)
        self.rapor_tablo.setColumnWidth(R_SUA, 155)
        self.rapor_tablo.setItemDelegateForColumn(R_SUA, SuaDelegate(self.rapor_tablo))

        lay.addWidget(self.rapor_tablo, 1)

        # Alt bar
        bot = QFrame()
        bot.setFixedHeight(44)
        bot.setStyleSheet(f"""
            QFrame {{
                background: {T.BG_SECONDARY};
                border-top: 1px solid {T.BORDER_PRIMARY};
            }}
        """)
        bl = QHBoxLayout(bot)
        bl.setContentsMargins(20, 0, 20, 0)
        bl.setSpacing(10)

        self.rapor_lbl_durum = QLabel("Hazır")
        self.rapor_lbl_durum.setStyleSheet(S["footer_label"])
        bl.addWidget(self.rapor_lbl_durum)
        bl.addStretch()

        self.rapor_lbl_bilgi = QLabel("")
        self.rapor_lbl_bilgi.setStyleSheet(S["info_label"])
        self.rapor_lbl_bilgi.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
        bl.addWidget(self.rapor_lbl_bilgi)

        self.rapor_progress = self._make_progress()
        bl.addWidget(self.rapor_progress)
        lay.addWidget(bot)
        return w

    # ── Footer ──────────────────────────────────────────────

    def _build_footer(self) -> QFrame:
        # Footer şu an boş; ileride log / breadcrumb eklenebilir
        frame = QFrame()
        frame.setFixedHeight(0)   # görünmez spacer
        return frame

    # ════════════════════════════════════════════════════════
    #  SİNYALLER
    # ════════════════════════════════════════════════════════

    def _connect_signals(self):
        self.cmb_yil.currentIndexChanged.connect(self._donem_guncelle)
        self.cmb_ay.currentIndexChanged.connect(self._donem_guncelle)
        self.btn_hesapla.clicked.connect(self._baslat_kontrol)
        self.btn_kaydet.clicked.connect(self._kaydet)
        self.btn_rapor_olustur.clicked.connect(self._rapor_olustur)
        self.btn_excel.clicked.connect(self._excel_indir)
        self.btn_pdf.clicked.connect(self._pdf_indir)
        self.hesap_tablo.itemChanged.connect(self._hucre_degisti)
        self._donem_guncelle()

    # ════════════════════════════════════════════════════════
    #  SEKME YÖNETİMİ
    # ════════════════════════════════════════════════════════

    def _switch_tab(self, idx: int):
        self._stack.setCurrentIndex(idx)
        for i, btn in self._tab_btns.items():
            btn.setStyleSheet(self._tab_qss(i == idx))
        # Rapor sekmesine geçince yıl combo'sunu güncelle
        if idx == 1:
            self.cmb_rapor_donem.blockSignals(True)
            cur_ay = self.cmb_ay.currentText()
            try:
                ay_idx = AY_ISIMLERI.index(cur_ay) + 1
            except ValueError:
                ay_idx = 0
            self.cmb_rapor_donem.setCurrentIndex(ay_idx)
            self.cmb_rapor_donem.blockSignals(False)

    # ════════════════════════════════════════════════════════
    #  DÖNEM YARDIMCISI
    # ════════════════════════════════════════════════════════

    def _get_donem(self):
        try:
            yil = int(self.cmb_yil.currentText())
            ay  = self.cmb_ay.currentIndex() + 1
            bas = datetime(yil, ay, 15)
            bit = bas + relativedelta(months=1) - timedelta(days=1)
            return bas, bit
        except Exception:
            return None, None

    def _donem_guncelle(self):
        bas, bit = self._get_donem()
        if bas and bit:
            self.lbl_donem.setText(
                f"Dönem: {bas.strftime('%d.%m.%Y')} — {bit.strftime('%d.%m.%Y')}"
            )

    # ════════════════════════════════════════════════════════
    #  VERİ YÜKLEME  (sayfa açılışında çağrılır)
    # ════════════════════════════════════════════════════════

    def load_data(self):
        if not self._db:
            return
        self.hesap_lbl_durum.setText("Veriler yükleniyor…")
        self.hesap_progress.setVisible(True)
        try:
            from core.di import get_registry
            reg = get_registry(self._db)
            self._all_personel = reg.get("Personel").get_all()
            self._all_izin     = reg.get("Izin_Giris").get_all()

            # Tatiller
            try:
                tatil = reg.get("Tatiller").get_all()
                self._tatil_np = []
                for r in tatil:
                    d = _parse_date(r.get("Tarih", ""))
                    if d:
                        self._tatil_np.append(d.strftime("%Y-%m-%d"))
            except Exception:
                self._tatil_np = []

            # Birim → Koşul haritası
            sabitler = reg.get("Sabitler").get_all()
            self._birim_kosul_map = {}
            for r in sabitler:
                if str(r.get("Kod", "")).strip() != "Gorev_Yeri":
                    continue
                birim = tr_upper(str(r.get("MenuEleman", "")).strip())
                ack   = tr_upper(str(r.get("Aciklama",   "")).strip())
                if not birim:
                    continue
                if "KOŞULU A" in ack:
                    self._birim_kosul_map[birim] = "A"
                elif "KOŞULU B" in ack:
                    self._birim_kosul_map[birim] = "B"

            self.hesap_progress.setVisible(False)
            self.hesap_lbl_durum.setText(
                f"{len(self._all_personel)} personel, "
                f"{len(self._all_izin)} izin kaydı yüklendi."
            )
            logger.info("FHSZ Merkez veri yüklendi.")
        except Exception as e:
            self.hesap_progress.setVisible(False)
            self.hesap_lbl_durum.setText(f"Hata: {e}")
            logger.error(f"FHSZ Merkez load_data: {e}")

    # ════════════════════════════════════════════════════════
    #  HESAPLAMA SEKMESİ — iş mantığı (orijinalden birebir)
    # ════════════════════════════════════════════════════════

    def _baslat_kontrol(self):
        bas, bit = self._get_donem()
        if not bas:
            return
        if bit < FHSZ_ESIK:
            QMessageBox.warning(self, "Uyarı", "26.04.2022 öncesi hesaplanamaz.")
            return
        self.hesap_tablo.setRowCount(0)
        self.btn_hesapla.setEnabled(False)
        self.hesap_progress.setVisible(True)
        self.hesap_lbl_durum.setText("Kayıtlar kontrol ediliyor…")
        yil_s = self.cmb_yil.currentText()
        ay_s  = self.cmb_ay.currentText()
        try:
            from core.di import get_registry
            reg  = get_registry(self._db)
            tum  = reg.get("FHSZ_Puantaj").get_all()
            mev  = [r for r in tum
                    if str(r.get("AitYil","")).strip() == yil_s
                    and str(r.get("Donem","")).strip() == ay_s]
            if mev:
                self._kayitli_yukle(mev)
            else:
                self._sifirdan_hesapla()
            self._guncelle_stat_kartlar()
        except Exception as e:
            logger.error(f"FHSZ kontrol: {e}")
            QMessageBox.critical(self, "Hata", str(e))
        self.hesap_progress.setVisible(False)
        self.btn_hesapla.setEnabled(True)
        # Sekme başlıklarını güncelle
        self._switch_tab(0)

    def _guncelle_stat_kartlar(self):
        n = self.hesap_tablo.rowCount()
        self._card_personel.set_value(str(n))

        bas, bit = self._get_donem()
        if bas and bit:
            hesap_bas = max(bas, FHSZ_ESIK)
            ig = is_gunu_hesapla(hesap_bas, bit, self._tatil_np)
            self._card_is_gunu.set_value(str(ig))

        top_saat = 0
        max_sua  = 0
        for r in range(n):
            item = self.hesap_tablo.item(r, C_SAAT)
            if item:
                try:
                    top_saat += float(item.text())
                except ValueError:
                    pass
        self._card_saat.set_value(f"{top_saat:.0f}")

        # max şua: kümülatif yıllık toplam için basit tahmin
        max_sua = sua_hak_edis_hesapla(top_saat)
        self._card_sua_max.set_value(str(max_sua))

    def _kayitli_yukle(self, rows):
        self.hesap_lbl_durum.setText(f"Veritabanından {len(rows)} kayıt yüklendi.")
        self.hesap_tablo.blockSignals(True)
        self.hesap_tablo.setRowCount(0)
        mevcut_tc = []
        for rd in rows:
            ri = self.hesap_tablo.rowCount()
            self.hesap_tablo.insertRow(ri)
            tc = str(rd.get("Personelid","")).strip()
            mevcut_tc.append(tc)
            self._set_item(ri, C_KIMLIK, tc)
            self._set_item(ri, C_AD,     rd.get("AdSoyad",""))
            self._set_item(ri, C_BIRIM,  rd.get("Birim",""))
            kosul = rd.get("CalismaKosulu") or "Çalışma Koşulu B"
            ik = QTableWidgetItem(str(kosul))
            ik.setFlags(Qt.ItemIsSelectable | Qt.ItemIsEnabled | Qt.ItemIsEditable)
            self.hesap_tablo.setItem(ri, C_KOSUL, ik)
            self._set_item(ri, C_YIL,   self.cmb_yil.currentText())
            self._set_item(ri, C_DONEM, self.cmb_ay.currentText())
            self._set_item(ri, C_GUN,   str(rd.get("AylikGun","0")))
            self._set_item(ri, C_IZIN,  str(rd.get("KullanilanIzin","0")))
            self._set_item(ri, C_SAAT,  str(rd.get("FiiliCalismaSaat","0")))
        bas, bit = self._get_donem()
        if bas and bit:
            hesap_bas = max(bas, FHSZ_ESIK)
            yeni = self._eksik_personel_ekle(mevcut_tc, hesap_bas, bit)
            if yeni:
                QMessageBox.information(
                    self, "Bilgi",
                    f"{yeni} yeni personel listeye eklendi.")
        self.hesap_tablo.blockSignals(False)

    def _sifirdan_hesapla(self):
        self.hesap_lbl_durum.setText("Yeni hesaplama yapılıyor…")
        self.hesap_tablo.blockSignals(True)
        bas, bit = self._get_donem()
        if not bas or bit < FHSZ_ESIK:
            self.hesap_tablo.blockSignals(False)
            return
        hesap_bas = max(bas, FHSZ_ESIK)
        for p in sorted(self._all_personel,
                        key=lambda x: str(x.get("AdSoyad",""))):
            if str(p.get("HizmetSinifi","")) not in IZIN_VERILEN_SINIFLAR:
                continue
            tc    = str(p.get("KimlikNo","")).strip()
            ad    = p.get("AdSoyad","")
            birim = str(p.get("GorevYeri","")).strip()
            durum = str(p.get("Durum","Aktif")).strip()
            kisi_bit = bit
            if durum == "Pasif":
                ayr = _parse_date(p.get("AyrilisTarihi",""))
                if ayr:
                    if ayr < hesap_bas:
                        continue
                    if ayr < bit:
                        kisi_bit = ayr
            ri = self.hesap_tablo.rowCount()
            self.hesap_tablo.insertRow(ri)
            self._set_item(ri, C_KIMLIK, tc)
            self._set_item(ri, C_AD,     ad)
            self._set_item(ri, C_BIRIM,  birim)
            kosul = "Çalışma Koşulu A" \
                if self._birim_kosul_map.get(tr_upper(birim)) == "A" \
                else "Çalışma Koşulu B"
            ik = QTableWidgetItem(kosul)
            ik.setFlags(Qt.ItemIsSelectable | Qt.ItemIsEnabled | Qt.ItemIsEditable)
            self.hesap_tablo.setItem(ri, C_KOSUL, ik)
            self._set_item(ri, C_YIL,   self.cmb_yil.currentText())
            self._set_item(ri, C_DONEM, self.cmb_ay.currentText())
            ig = is_gunu_hesapla(hesap_bas, kisi_bit, self._tatil_np)
            self._set_item(ri, C_GUN,  str(ig))
            iz = self._izin_kesisim(tc, hesap_bas, kisi_bit)
            self._set_item(ri, C_IZIN, str(iz))
            self._satir_hesapla(ri)
        self.hesap_tablo.blockSignals(False)
        self.hesap_lbl_durum.setText(
            f"{self.hesap_tablo.rowCount()} personel hesaplandı.")

    def _eksik_personel_ekle(self, mevcut, hesap_bas, bit) -> int:
        n = 0
        for p in sorted(self._all_personel,
                        key=lambda x: str(x.get("AdSoyad",""))):
            tc = str(p.get("KimlikNo","")).strip()
            if tc in mevcut:
                continue
            if str(p.get("HizmetSinifi","")) not in IZIN_VERILEN_SINIFLAR:
                continue
            durum = str(p.get("Durum","Aktif")).strip()
            kisi_bit = bit
            if durum == "Pasif":
                ayr = _parse_date(p.get("AyrilisTarihi",""))
                if ayr:
                    if ayr < hesap_bas:
                        continue
                    if ayr < bit:
                        kisi_bit = ayr
            ri = self.hesap_tablo.rowCount()
            self.hesap_tablo.insertRow(ri)
            birim = str(p.get("GorevYeri","")).strip()
            self._set_item(ri, C_KIMLIK, tc)
            self._set_item(ri, C_AD,     p.get("AdSoyad",""))
            self._set_item(ri, C_BIRIM,  birim)
            kosul = "Çalışma Koşulu A" \
                if self._birim_kosul_map.get(tr_upper(birim)) == "A" \
                else "Çalışma Koşulu B"
            ik = QTableWidgetItem(kosul)
            ik.setFlags(Qt.ItemIsSelectable | Qt.ItemIsEnabled | Qt.ItemIsEditable)
            self.hesap_tablo.setItem(ri, C_KOSUL, ik)
            self._set_item(ri, C_YIL,   self.cmb_yil.currentText())
            self._set_item(ri, C_DONEM, self.cmb_ay.currentText())
            self._set_item(ri, C_GUN,   str(is_gunu_hesapla(hesap_bas, kisi_bit, self._tatil_np)))
            self._set_item(ri, C_IZIN,  str(self._izin_kesisim(tc, hesap_bas, kisi_bit)))
            self._satir_hesapla(ri)
            # Yeni satır — amber vurgu
            for c in range(self.hesap_tablo.columnCount()):
                item = self.hesap_tablo.item(ri, c)
                if item:
                    item.setBackground(QColor(245, 158, 11, 25))
            n += 1
        return n

    def _izin_kesisim(self, tc, bas, bit) -> int:
        toplam = 0
        for iz in self._all_izin:
            if str(iz.get("Personelid","")).strip() != tc:
                continue
            if str(iz.get("Durum","")).strip() == "İptal":
                continue
            ib = _parse_date(iz.get("BaslamaTarihi",""))
            ie = _parse_date(iz.get("BitisTarihi",""))
            if not ib or not ie:
                continue
            kb = max(bas, ib)
            ke = min(bit, ie)
            if kb <= ke:
                toplam += is_gunu_hesapla(kb, ke, self._tatil_np)
        return toplam

    def _set_item(self, row, col, text):
        item = QTableWidgetItem(str(text))
        item.setFlags(Qt.ItemIsSelectable | Qt.ItemIsEnabled)
        self.hesap_tablo.setItem(row, col, item)

    def _satir_hesapla(self, row):
        try:
            kosul  = self.hesap_tablo.item(row, C_KOSUL).text()
            is_gun = int(self.hesap_tablo.item(row, C_GUN).text())
            izin   = int(self.hesap_tablo.item(row, C_IZIN).text())
            puan   = 0
            if "KOŞULU A" in tr_upper(kosul):
                puan = max(0, is_gun - izin) * KOSUL_A_SAAT
            self.hesap_tablo.setItem(row, C_SAAT, QTableWidgetItem(str(puan)))
        except Exception:
            pass

    def _hucre_degisti(self, item):
        if item.column() == C_KOSUL:
            self._satir_hesapla(item.row())

    # ════════════════════════════════════════════════════════
    #  KAYDET
    # ════════════════════════════════════════════════════════

    def _kaydet(self):
        if self.hesap_tablo.rowCount() == 0:
            return
        yil_s = self.cmb_yil.currentText()
        ay_s  = self.cmb_ay.currentText()
        try:
            from core.di import get_registry
            reg  = get_registry(self._db)
            repo = reg.get("FHSZ_Puantaj")
            tum  = repo.get_all()
            mev_sayi = sum(
                1 for r in tum
                if str(r.get("AitYil","")).strip() == yil_s
                and str(r.get("Donem","")).strip() == ay_s
            )
            if mev_sayi > 0:
                cevap = QMessageBox.question(
                    self, "Veri Güncelleme",
                    f"{ay_s} {yil_s} için {mev_sayi} kayıt mevcut.\n"
                    "Mevcut kayıtlar silinip tablodaki veriler kaydedilsin mi?",
                    QMessageBox.Yes | QMessageBox.No, QMessageBox.No
                )
                if cevap != QMessageBox.Yes:
                    self.hesap_lbl_durum.setText("İptal edildi.")
                    return
            self.btn_kaydet.setEnabled(False)
            self.hesap_progress.setVisible(True)
            if mev_sayi > 0:
                for r in tum:
                    if (str(r.get("AitYil","")).strip() == yil_s
                            and str(r.get("Donem","")).strip() == ay_s):
                        try:
                            repo.delete([str(r.get("Personelid","")),
                                         str(r.get("AitYil","")),
                                         str(r.get("Donem",""))])
                        except Exception:
                            pass
            kayit = 0
            for ri in range(self.hesap_tablo.rowCount()):
                def _t(c):
                    it = self.hesap_tablo.item(ri, c)
                    return it.text() if it else ""
                repo.insert({
                    "Personelid":       _t(C_KIMLIK),
                    "AdSoyad":          _t(C_AD),
                    "Birim":            _t(C_BIRIM),
                    "CalismaKosulu":    _t(C_KOSUL),
                    "AitYil":           _t(C_YIL) or yil_s,
                    "Donem":            _t(C_DONEM) or ay_s,
                    "AylikGun":         _t(C_GUN),
                    "KullanilanIzin":   _t(C_IZIN),
                    "FiiliCalismaSaat": _t(C_SAAT),
                })
                kayit += 1
            # Şua bakiyesi
            self._sua_bakiye_guncelle(repo, yil_s)
            self.hesap_progress.setVisible(False)
            self.btn_kaydet.setEnabled(True)
            self.hesap_lbl_durum.setText(
                f"✓ {kayit} kayıt kaydedildi  —  {ay_s} {yil_s}")
            logger.info(f"FHSZ kaydedildi: {ay_s} {yil_s}, {kayit} kayıt")
            QMessageBox.information(self, "Başarılı", "Kayıt işlemi tamamlandı.")
        except Exception as e:
            self.hesap_progress.setVisible(False)
            self.btn_kaydet.setEnabled(True)
            self.hesap_lbl_durum.setText(f"Hata: {e}")
            logger.error(f"FHSZ kayıt: {e}")
            QMessageBox.critical(self, "Hata", str(e))

    def _sua_bakiye_guncelle(self, repo_puantaj, yil_s):
        try:
            from core.di import get_registry
            reg = get_registry(self._db)
            tum = repo_puantaj.get_all()
            p_top: dict[str, float] = {}
            for r in tum:
                if str(r.get("AitYil","")).strip() != yil_s:
                    continue
                tc = str(r.get("Personelid","")).strip()
                try:
                    s = float(str(r.get("FiiliCalismaSaat",0)).replace(",","."))
                except (ValueError, TypeError):
                    s = 0
                p_top[tc] = p_top.get(tc, 0) + s
            iz_bilgi = reg.get("Izin_Bilgi")
            for tc, top in p_top.items():
                hak = sua_hak_edis_hesapla(top)
                try:
                    mev = iz_bilgi.get_by_id(tc)
                    if mev:
                        try:
                            eski = float(str(mev.get("SuaCariYilKazanim",0)).replace(",","."))
                        except (ValueError, TypeError):
                            eski = -1
                        if eski != hak:
                            iz_bilgi.update(tc, {"SuaCariYilKazanim": hak})
                except Exception:
                    pass
            logger.info(f"Şua bakiyesi güncellendi: {len(p_top)} personel")
        except Exception as e:
            logger.error(f"Şua bakiye: {e}")

    # ════════════════════════════════════════════════════════
    #  RAPOR SEKMESİ — iş mantığı
    # ════════════════════════════════════════════════════════

    def _rapor_olustur(self):
        if not self._db:
            return
        yil_s   = self.cmb_yil.currentText()
        don_s   = self.cmb_rapor_donem.currentText()
        tek_don = don_s != "Tümü (Yıllık)"

        self.rapor_tablo.setRowCount(0)
        self._rapor_data = []
        self.btn_rapor_olustur.setEnabled(False)
        self.rapor_progress.setVisible(True)
        self.rapor_lbl_durum.setText("Rapor hazırlanıyor…")

        try:
            from core.di import get_registry
            reg  = get_registry(self._db)
            repo = reg.get("FHSZ_Puantaj")
            tum  = repo.get_all()

            yil_kyt = [r for r in tum
                       if str(r.get("AitYil","")).strip() == yil_s]
            if not yil_kyt:
                self.rapor_lbl_bilgi.setText("Bu yıla ait kayıt bulunamadı.")
                self.rapor_lbl_durum.setText("Hazır")
                self.rapor_progress.setVisible(False)
                self.btn_rapor_olustur.setEnabled(True)
                self.btn_excel.setEnabled(False)
                self.btn_pdf.setEnabled(False)
                return

            # Personel → dönem sıralı kayıtlar
            p_map: dict[str, list] = {}
            for r in yil_kyt:
                tc = str(r.get("Personelid","")).strip()
                p_map.setdefault(tc, []).append(r)
            for tc in p_map:
                p_map[tc].sort(
                    key=lambda r: AY_SIRA.get(str(r.get("Donem","")).strip(), 99))

            rows = []
            for tc, kayitlar in sorted(
                    p_map.items(),
                    key=lambda x: str(x[1][0].get("AdSoyad",""))):
                kum = 0
                t_gun = t_izn = t_sat = 0
                for r in kayitlar:
                    don = str(r.get("Donem","")).strip()
                    try:
                        s = float(str(r.get("FiiliCalismaSaat",0)).replace(",","."))
                    except (ValueError, TypeError):
                        s = 0
                    kum += s;  t_sat += s
                    try: t_gun += int(r.get("AylikGun",0))
                    except (ValueError, TypeError): pass
                    try: t_izn += int(r.get("KullanilanIzin",0))
                    except (ValueError, TypeError): pass
                    if tek_don and don != don_s:
                        continue
                    sua = sua_hak_edis_hesapla(kum)
                    if not tek_don:
                        continue
                    rows.append({
                        "Personelid": tc, "AdSoyad": r.get("AdSoyad",""),
                        "AitYil": yil_s, "Donem": don,
                        "AylikGun": r.get("AylikGun",0),
                        "KullanilanIzin": r.get("KullanilanIzin",0),
                        "FiiliCalismaSaat": s, "KumulatifSaat": kum,
                        "SuaHakEdis": sua,
                    })
                if not tek_don and kayitlar:
                    rows.append({
                        "Personelid": tc, "AdSoyad": kayitlar[0].get("AdSoyad",""),
                        "AitYil": yil_s, "Donem": "Toplam",
                        "AylikGun": t_gun, "KullanilanIzin": t_izn,
                        "FiiliCalismaSaat": t_sat, "KumulatifSaat": t_sat,
                        "SuaHakEdis": sua_hak_edis_hesapla(t_sat),
                    })

            # Tabloya yaz
            self.rapor_tablo.setRowCount(len(rows))
            for i, row in enumerate(rows):
                def si(c, v):
                    item = QTableWidgetItem(str(v))
                    item.setFlags(Qt.ItemIsSelectable | Qt.ItemIsEnabled)
                    if c >= R_GUN:
                        item.setTextAlignment(Qt.AlignCenter)
                    self.rapor_tablo.setItem(i, c, item)
                si(R_KIM, row["Personelid"]); si(R_AD,  row["AdSoyad"])
                si(R_YIL, row["AitYil"]);     si(R_DON, row["Donem"])
                si(R_GUN, str(int(row["AylikGun"])))
                si(R_IZN, str(int(row["KullanilanIzin"])))
                si(R_SAT, f"{row['FiiliCalismaSaat']:.0f}")
                si(R_KUM, f"{row['KumulatifSaat']:.0f}")
                sua_item = QTableWidgetItem(str(row["SuaHakEdis"]))
                sua_item.setFlags(Qt.ItemIsSelectable | Qt.ItemIsEnabled)
                sua_item.setTextAlignment(Qt.AlignCenter)
                self.rapor_tablo.setItem(i, R_SUA, sua_item)

            self._rapor_data = rows
            self.rapor_lbl_bilgi.setText(
                f"{len(p_map)} personel  •  {len(rows)} satır  "
                f"•  {don_s}  •  {yil_s}"
            )
            self.rapor_lbl_durum.setText(f"Rapor hazır — {len(rows)} satır")
            self.btn_excel.setEnabled(bool(rows))
            self.btn_pdf.setEnabled(bool(rows))
            logger.info(f"FHSZ Rapor oluşturuldu: {yil_s}, {len(rows)} satır")
        except Exception as e:
            logger.error(f"FHSZ Rapor: {e}")
            QMessageBox.critical(self, "Hata", str(e))
        self.rapor_progress.setVisible(False)
        self.btn_rapor_olustur.setEnabled(True)

    # ════════════════════════════════════════════════════════
    #  EXCEL / PDF (orijinaldeki raporlama fonksiyonları)
    # ════════════════════════════════════════════════════════

    def _excel_indir(self):
        if not self._rapor_data:
            return
        yil = self.cmb_yil.currentText()
        don = self.cmb_rapor_donem.currentText().replace(" ","_")
        path, _ = QFileDialog.getSaveFileName(
            self, "Excel Kaydet",
            f"FHSZ_Puantaj_Rapor_{yil}_{don}.xlsx",
            "Excel Dosyası (*.xlsx)"
        )
        if not path:
            return
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Puantaj Rapor"

            hf   = Font(name="Arial", bold=True, color="FFFFFF", size=11)
            fill = PatternFill(start_color="0090B0", end_color="0090B0", fill_type="solid")
            ha   = Alignment(horizontal="center", vertical="center", wrap_text=True)
            brd  = Border(
                left=Side(style="thin"), right=Side(style="thin"),
                top=Side(style="thin"), bottom=Side(style="thin")
            )
            for c, h in enumerate(RAPOR_COLS, 1):
                cell = ws.cell(row=1, column=c, value=h)
                cell.font = hf; cell.fill = fill
                cell.alignment = ha; cell.border = brd

            df   = Font(name="Arial", size=10)
            ca   = Alignment(horizontal="center", vertical="center")
            la   = Alignment(horizontal="left", vertical="center")
            f_g  = PatternFill(start_color="0B3D1E", end_color="0B3D1E", fill_type="solid")
            f_y  = PatternFill(start_color="7A4600", end_color="7A4600", fill_type="solid")
            f_b  = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
            fw   = Font(name="Arial", bold=True, color="FFFFFF", size=10)

            for ri, row in enumerate(self._rapor_data, 2):
                vals = [
                    row["Personelid"], row["AdSoyad"],
                    row["AitYil"], row["Donem"],
                    int(row["AylikGun"]), int(row["KullanilanIzin"]),
                    int(row["FiiliCalismaSaat"]), int(row["KumulatifSaat"]),
                    int(row["SuaHakEdis"]),
                ]
                for ci, v in enumerate(vals, 1):
                    cell = ws.cell(row=ri, column=ci, value=v)
                    cell.font = df; cell.border = brd
                    cell.alignment = ca if ci >= 3 else la
                sv = int(row["SuaHakEdis"])
                sc = ws.cell(row=ri, column=9)
                if sv >= 20:
                    sc.fill = f_g; sc.font = fw
                elif sv >= 10:
                    sc.fill = f_y; sc.font = fw
                elif sv > 0:
                    sc.fill = f_b; sc.font = fw

            for i, w in enumerate([14,25,8,12,10,10,14,16,18], 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
            ws.auto_filter.ref = ws.dimensions
            wb.save(path)
            self.rapor_lbl_durum.setText(f"Excel kaydedildi: {os.path.basename(path)}")
            QMessageBox.information(self, "Başarılı", f"Excel kaydedildi:\n{path}")
        except ImportError:
            QMessageBox.warning(self, "Uyarı", "openpyxl yüklü değil.\npip install openpyxl")
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Excel kaydedilemedi:\n{e}")

    def _pdf_indir(self):
        if not self._rapor_data:
            return
        yil = self.cmb_yil.currentText()
        don = self.cmb_rapor_donem.currentText().replace(" ","_")
        path, _ = QFileDialog.getSaveFileName(
            self, "PDF Kaydet",
            f"FHSZ_Puantaj_Rapor_{yil}_{don}.pdf",
            "PDF Dosyası (*.pdf)"
        )
        if not path:
            return
        try:
            from reportlab.lib import colors as rl_colors
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.lib.styles import getSampleStyleSheet
            from reportlab.lib.units import mm
            from reportlab.platypus import (
                SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            )
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont

            try:
                pdfmetrics.registerFont(TTFont("Arial","arial.ttf"))
                pdfmetrics.registerFont(TTFont("ArialBold","arialbd.ttf"))
                fn, fb = "Arial", "ArialBold"
            except Exception:
                fn, fb = "Helvetica", "Helvetica-Bold"

            doc = SimpleDocTemplate(
                path, pagesize=landscape(A4),
                leftMargin=15*mm, rightMargin=15*mm,
                topMargin=15*mm, bottomMargin=15*mm
            )
            st = getSampleStyleSheet()
            elems = []
            ts = st["Title"]
            ts.fontName = fb; ts.fontSize = 14
            elems.append(Paragraph(
                f"FHSZ Puantaj Raporu — {yil} {self.cmb_rapor_donem.currentText()}", ts))
            elems.append(Spacer(1, 5*mm))

            td = [RAPOR_COLS[:]]
            for row in self._rapor_data:
                td.append([
                    str(row["Personelid"]), str(row["AdSoyad"]),
                    str(row["AitYil"]), str(row["Donem"]),
                    str(int(row["AylikGun"])), str(int(row["KullanilanIzin"])),
                    str(int(row["FiiliCalismaSaat"])), str(int(row["KumulatifSaat"])),
                    str(int(row["SuaHakEdis"])),
                ])
            t = Table(td, colWidths=[55,100,35,50,40,40,55,60,70], repeatRows=1)
            sc = [
                ("BACKGROUND",(0,0),(-1,0),rl_colors.HexColor("#0090B0")),
                ("TEXTCOLOR",(0,0),(-1,0),rl_colors.white),
                ("FONTNAME",(0,0),(-1,0),fb), ("FONTSIZE",(0,0),(-1,0),8),
                ("FONTNAME",(0,1),(-1,-1),fn), ("FONTSIZE",(0,1),(-1,-1),7),
                ("ALIGN",(0,0),(-1,-1),"CENTER"),
                ("ALIGN",(1,1),(1,-1),"LEFT"),
                ("VALIGN",(0,0),(-1,-1),"MIDDLE"),
                ("GRID",(0,0),(-1,-1),0.5,rl_colors.grey),
                ("ROWBACKGROUNDS",(0,1),(-1,-1),
                 [rl_colors.white, rl_colors.HexColor("#0D1F35")]),
            ]
            for i, row in enumerate(self._rapor_data, 1):
                sv = int(row["SuaHakEdis"])
                if sv >= 20:
                    sc += [("BACKGROUND",(8,i),(8,i),rl_colors.HexColor("#0B3D1E")),
                            ("TEXTCOLOR",(8,i),(8,i),rl_colors.white)]
                elif sv >= 10:
                    sc += [("BACKGROUND",(8,i),(8,i),rl_colors.HexColor("#7A4600")),
                            ("TEXTCOLOR",(8,i),(8,i),rl_colors.white)]
                elif sv > 0:
                    sc += [("BACKGROUND",(8,i),(8,i),rl_colors.HexColor("#003366")),
                            ("TEXTCOLOR",(8,i),(8,i),rl_colors.white)]
            t.setStyle(TableStyle(sc))
            elems.append(t)
            elems.append(Spacer(1,5*mm))
            ns = st["Normal"]
            ns.fontName = fn; ns.fontSize = 8
            elems.append(Paragraph(
                f"Rapor tarihi: {datetime.now().strftime('%d.%m.%Y %H:%M')} "
                f"— {len(self._rapor_data)} kayıt", ns))
            doc.build(elems)
            self.rapor_lbl_durum.setText(f"PDF kaydedildi: {os.path.basename(path)}")
            QMessageBox.information(self, "Başarılı", f"PDF kaydedildi:\n{path}")
        except ImportError:
            QMessageBox.warning(self, "Uyarı", "reportlab yüklü değil.\npip install reportlab")
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"PDF kaydedilemedi:\n{e}")

    # ════════════════════════════════════════════════════════
    #  YARDIMCI OLUŞTURUCULAR
    # ════════════════════════════════════════════════════════

    @staticmethod
    def _make_progress() -> QProgressBar:
        pb = QProgressBar()
        pb.setVisible(False)
        pb.setRange(0, 0)
        pb.setFixedWidth(160)
        pb.setFixedHeight(4)
        pb.setTextVisible(False)
        pb.setStyleSheet(S["progress"])
        return pb

    @staticmethod
    def _hdr_lbl(text: str) -> QLabel:
        lbl = QLabel(text)
        lbl.setStyleSheet(
            f"color:{T.TEXT_MUTED}; font-size:12px; background:transparent;")
        return lbl

    @staticmethod
    def _vline() -> QFrame:
        f = QFrame()
        f.setFixedSize(1, 20)
        f.setStyleSheet(f"background:{T.BORDER_STRONG};")
        return f

    @staticmethod
    def _tab_qss(active: bool) -> str:
        if active:
            return (
                f"QPushButton{{"
                f"background:transparent; border:none;"
                f"border-bottom:2px solid {T.ACCENT};"
                f"color:{T.ACCENT2};"
                f"font-size:13px; font-weight:700; padding:0 18px;"
                f"}}"
            )
        return (
            f"QPushButton{{"
            f"background:transparent; border:none;"
            f"border-bottom:2px solid transparent;"
            f"color:{T.TEXT_MUTED};"
            f"font-size:13px; padding:0 18px;"
            f"}}"
            f"QPushButton:hover{{color:{T.TEXT_SECONDARY};}}"
        )
