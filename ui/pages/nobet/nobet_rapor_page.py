# -*- coding: utf-8 -*-
"""Nöbet Rapor Sayfası — PDF ve Excel çıktı."""
import datetime
from typing import Optional

from PySide6.QtCore import Qt, QThread, Signal
from PySide6.QtGui import QCursor
from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QFrame, QLabel,
    QPushButton, QComboBox, QGroupBox, QGridLayout,
    QRadioButton, QButtonGroup, QApplication, QTableView,
    QAbstractItemView,
)

from core.di import get_nobet_service
from core.logger import logger
from core.hata_yonetici import hata_goster, uyari_goster
from ui.components.base_table_model import BaseTableModel
from ui.styles.icons import IconRenderer, IconColors

_AY_ADLARI = ["","Ocak","Şubat","Mart","Nisan","Mayıs","Haziran",
               "Temmuz","Ağustos","Eylül","Ekim","Kasım","Aralık"]


# ─── PDF şablonu ─────────────────────────────────────────

def _css():
    return ("body{font-family:'Times New Roman',serif;font-size:10.5pt;color:#000;}"
            "h1{text-align:center;font-size:13pt;font-weight:bold;margin-bottom:4px;}"
            "h2{font-size:11pt;font-weight:bold;margin-top:12px;margin-bottom:3px;"
            "background:#ddd;padding:2px 6px;}"
            ".c{text-align:center;font-size:10pt;margin-bottom:8px;}"
            "table{width:100%;border-collapse:collapse;margin-top:6px;font-size:9.5pt;}"
            "th,td{border:1px solid #000;padding:3px 5px;vertical-align:middle;}"
            "th{background:#eee;font-weight:bold;text-align:center;}"
            ".l{text-align:left;} .c2{text-align:center;}"
            "tr:nth-child(even){background:#f9f9f9;}"
            ".fm{color:#c00;font-weight:600;}")


def html_nobet_listesi(veriler: list[dict], baslik: str,
                       donem: str, gruplu: bool = False) -> str:
    tarih = datetime.datetime.now().strftime("%d.%m.%Y")
    baslik_sat = ("<tr><th width='5%'>Sıra</th><th width='12%'>Tarih</th>"
                  "<th width='8%'>Gün</th><th width='20%'>Birim</th>"
                  "<th width='15%'>Vardiya</th><th width='25%'>Personel</th>"
                  "<th width='15%'>Tür</th></tr>")
    gunler = ["Pzt","Sal","Çar","Per","Cum","Cmt","Paz"]

    def _satirlar(rows, baslangic=1):
        result = []
        for i, r in enumerate(rows, baslangic):
            try:
                t  = datetime.date.fromisoformat(str(r.get("NobetTarihi","")))
                gs = gunler[t.weekday()]
                tarih_str = t.strftime("%d.%m.%Y")
            except Exception:
                gs = ""; tarih_str = str(r.get("NobetTarihi",""))
            fm_cls = ' class="fm"' if r.get("NobetTuru") == "fazla_mesai" else ""
            result.append(
                f"<tr><td class='c2'>{i}</td>"
                f"<td class='c2'>{tarih_str}</td>"
                f"<td class='c2'>{gs}</td>"
                f"<td class='l'>{r.get('BirimAdi','')}</td>"
                f"<td class='l'>{r.get('VardiyaAdi','')}</td>"
                f"<td class='l'{fm_cls}>{r.get('AdSoyad','')}</td>"
                f"<td class='c2'>{r.get('NobetTuru','').replace('_',' ').capitalize()}</td>"
                f"</tr>"
            )
        return "".join(result)

    if gruplu:
        from itertools import groupby
        sirali = sorted(veriler, key=lambda r: r.get("NobetTarihi",""))
        icerik = ""
        for tarih_g, tg in groupby(sirali, key=lambda r: r.get("NobetTarihi","")):
            t_list = list(tg)
            try:
                t   = datetime.date.fromisoformat(tarih_g)
                hdr = f"{t.strftime('%d.%m.%Y')} {gunler[t.weekday()]}"
            except Exception:
                hdr = tarih_g
            icerik += (f"<h2>{hdr}</h2>"
                       f"<table><thead>{baslik_sat}</thead>"
                       f"<tbody>{_satirlar(t_list)}</tbody></table>")
    else:
        icerik = (f"<table><thead>{baslik_sat}</thead>"
                  f"<tbody>{_satirlar(veriler)}</tbody></table>")

    return (f"<html><head><meta charset='utf-8'><style>{_css()}</style></head><body>"
            f"<h1>{baslik}</h1>"
            f"<div class='c'>{donem} &nbsp;|&nbsp; Rapor Tarihi: {tarih} "
            f"&nbsp;|&nbsp; Toplam: {len(veriler)} kayıt</div>"
            f"{icerik}"
            "<table style='border:none;margin-top:40px;'><tr>"
            "<td style='border:none;text-align:center;'>"
            "<b>Hazırlayan</b><br><br>_______________</td>"
            "<td style='border:none;text-align:center;'>"
            "<b>Onaylayan</b><br><br>_______________</td>"
            "</tr></table>"
            "</body></html>")


_RAPOR_COLUMNS = [
    ("NobetTarihi", "Tarih", 95),
    ("Gun", "Gün", 70),
    ("BirimAdi", "Birim", 180),
    ("VardiyaAdi", "Vardiya", 150),
    ("SaatAraligi", "Saat", 95),
    ("AdSoyad", "Personel", 220),
    ("NobetTuru", "Tür", 110),
    ("PlanDurum", "Plan", 105),
]


class _OnayliNobetModel(BaseTableModel):
    DATE_KEYS = frozenset({"NobetTarihi"})
    ALIGN_CENTER = frozenset({"NobetTarihi", "Gun", "SaatAraligi", "NobetTuru", "PlanDurum"})

    def __init__(self, rows=None, parent=None):
        super().__init__(_RAPOR_COLUMNS, rows, parent)

    def _display(self, key: str, row: dict) -> str:
        if key == "NobetTuru":
            return str(row.get(key, "")).replace("_", " ").title()
        if key == "PlanDurum":
            val = str(row.get(key, "")).strip().lower()
            if val == "yururlukte":
                return "Yürürlükte"
            if val == "onaylandi":
                return "Onaylandı"
        return super()._display(key, row)

    def _fg(self, key: str, row: dict):
        if key == "NobetTuru" and str(row.get("NobetTuru", "")) == "fazla_mesai":
            return self.status_fg("Beklemede")
        if key == "PlanDurum":
            durum = str(row.get("PlanDurum", "")).strip().lower()
            if durum == "yururlukte":
                return self.status_fg("Aktif")
            if durum == "onaylandi":
                return self.status_fg("Onaylandı")
        return None


# ─── Worker ──────────────────────────────────────────────

class _RaporWorker(QThread):
    log   = Signal(str)
    bitti = Signal()

    def __init__(self, mod: int, tablo: list[dict], ozet: list[dict],
                 context: dict, donem: str, kayit: str, gruplu: bool = True):
        super().__init__()
        self._mod    = mod
        self._tablo  = tablo    # gün bazlı satırlar
        self._ozet   = ozet     # personel özet satırları
        self._ctx    = context
        self._donem  = donem
        self._kayit  = kayit
        self._gruplu = gruplu

    def run(self):
        try:
            from core.rapor_servisi import RaporServisi
            if self._mod == 1:  # PDF — inline HTML
                baslik = f"{self._ctx.get('BirimAdi','')} Nöbet Listesi"
                html   = self._html_olustur(baslik)
                from PySide6.QtGui import QTextDocument, QPdfWriter, QPageSize, QPageLayout
                from PySide6.QtCore import QMarginsF
                doc = QTextDocument()
                doc.setHtml(html)
                w   = QPdfWriter(self._kayit)
                w.setPageSize(QPageSize(QPageSize.PageSizeId.A4))
                lay = QPageLayout()
                lay.setPageSize(QPageSize(QPageSize.PageSizeId.A4))
                lay.setOrientation(QPageLayout.Orientation.Landscape)
                lay.setMargins(QMarginsF(10, 10, 10, 10))
                w.setPageLayout(lay)
                doc.print_(w)
                self.log.emit(f"PDF hazır: {self._kayit}")
                RaporServisi.ac(self._kayit)

            else:  # Excel — şablon veya openpyxl
                yol = RaporServisi.excel("nobet_listesi", self._ctx,
                                          self._tablo, self._kayit)
                if not yol:
                    # Şablon yoksa openpyxl ile doğrudan oluştur
                    yol = self._excel_olustur()
                if yol:
                    self.log.emit(f"Excel hazır: {yol}")
                    RaporServisi.ac(yol)
                else:
                    self.log.emit("Excel oluşturulamadı.")

        except Exception as e:
            self.log.emit(f"Hata: {e}")
            logger.error(f"Nöbet rapor worker: {e}")
        finally:
            self.bitti.emit()

    def _html_olustur(self, baslik: str) -> str:
        ctx  = self._ctx
        ay   = ctx.get("Ay","")
        yil  = ctx.get("Yil","")
        birim = ctx.get("BirimAdi","")
        tarih_str = datetime.datetime.now().strftime("%d.%m.%Y")

        # Tüm vardiya adları
        v_adlar: list[str] = []
        for satir in self._tablo:
            for k in satir:
                if k not in ("Tarih","Gun") and k[0].isupper() and k[1].isdigit():
                    prefix = k[0]
                    if prefix not in v_adlar:
                        v_adlar.append(prefix)
        v_adlar = sorted(set(v_adlar))

        th_vardiya = "".join(
            f"<th colspan='4'>{chr(65+i+2)}</th>"
            for i, _ in enumerate(v_adlar)
        ) if v_adlar else ""
        th_kisi = "<th>Tarih</th><th>Gün</th>" + "".join(
            "<th>1</th><th>2</th><th>3</th><th>4</th>"
            for _ in v_adlar
        )

        rows_html = ""
        _GUN_K = {"Cumartesi","Pazar"}
        for satir in self._tablo:
            gn  = satir.get("Gun","")
            bg  = " style='background:#fff3cd'" if gn in _GUN_K else ""
            row = f"<tr{bg}><td>{satir.get('Tarih','')}</td><td><b>{gn}</b></td>"
            for prefix in v_adlar:
                for i in range(1, 5):
                    row += f"<td>{satir.get(f'{prefix}{i}','')}</td>"
            row += "</tr>"
            rows_html += row

        ozet_rows = ""
        for o in self._ozet:
            ozet_rows += (f"<tr><td class='l'>{o.get('AdSoyad','')}</td>"
                          f"<td>{o.get('ToplamSayisi',0)}</td>"
                          f"<td>{o.get('HedefSaat','')}</td>"
                          f"<td>{o.get('CalisanSaat','')}</td>"
                          f"<td>{o.get('FazlaMesai','')}</td></tr>")

        css = (_css() + "td{min-width:80px}"
               " @page{size:A4 landscape;margin:10mm}")
        return (f"<html><head><meta charset='utf-8'><style>{css}</style></head><body>"
                f"<h1>{birim} — {ay} {yil} Nöbet Listesi</h1>"
                f"<div class='c'>Rapor Tarihi: {tarih_str}</div>"
                "<table><thead>"
                f"<tr><th colspan='2'></th>{th_vardiya}</tr>"
                f"<tr>{th_kisi}</tr></thead><tbody>{rows_html}</tbody></table>"
                "<br><h2>Personel Özeti</h2>"
                "<table><thead><tr><th class='l'>Ad Soyad</th>"
                "<th>Nöbet</th><th>Hedef</th><th>Çalışılan</th><th>Fazla</th>"
                f"</tr></thead><tbody>{ozet_rows}</tbody></table>"
                "</body></html>")

    def _excel_olustur(self) -> str:
        """Şablon yoksa openpyxl ile doğrudan oluştur."""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            ctx  = self._ctx
            birim = ctx.get("BirimAdi","")
            ay   = ctx.get("Ay","")
            yil  = ctx.get("Yil","")

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Nöbet Listesi"
            ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE

            thin = Side(style="thin")
            bd   = Border(left=thin, right=thin, top=thin, bottom=thin)

            def sc(r, c, v, bold=False, bg=None, color="000000",
                   center=True, sz=10):
                cell = ws.cell(row=r, column=c, value=v)
                cell.font      = Font(name="Calibri", size=sz,
                                      bold=bold, color=color)
                cell.alignment = Alignment(
                    horizontal="center" if center else "left",
                    vertical="center", wrap_text=True)
                cell.border = bd
                if bg:
                    cell.fill = PatternFill("solid", fgColor=bg)
                return cell

            # Tüm vardiya prefix'leri
            v_adlar: list[str] = []
            for satir in self._tablo:
                for k in satir:
                    if k not in ("Tarih","Gun") and len(k)==2 and k[0].isupper() and k[1].isdigit():
                        p = k[0]
                        if p not in v_adlar:
                            v_adlar.append(p)
            v_adlar = sorted(set(v_adlar))

            # Sütun genişlikleri
            ws.column_dimensions["A"].width = 12
            ws.column_dimensions["B"].width = 10
            col = 3
            for _ in v_adlar:
                for i in range(4):
                    ws.column_dimensions[chr(64+col+i)].width = 18
                col += 4

            # Başlık
            total_cols = 2 + len(v_adlar) * 4
            ws.merge_cells(start_row=1, start_column=1,
                           end_row=1, end_column=total_cols)
            sc(1, 1, f"{birim} — {ay} {yil} Nöbet Listesi",
               bold=True, bg="1F4E79", color="FFFFFF", sz=13)
            ws.row_dimensions[1].height = 22

            # Vardiya başlıkları (2. satır)
            sc(2, 1, "", bg="2E75B6", color="FFFFFF")
            sc(2, 2, "", bg="2E75B6", color="FFFFFF")
            col = 3
            for vi, prefix in enumerate(v_adlar):
                # Vardiya adını context'ten al
                v_lbl = (ctx.get("GunduzBas","") + "-" + ctx.get("GunduzBit","")
                         if vi == 0 else
                         ctx.get("GeceBas","") + "-" + ctx.get("GeceBit","")
                         if vi == 1 else
                         ctx.get("SeyyarAdi",""))
                ws.merge_cells(start_row=2, start_column=col,
                               end_row=2, end_column=col+3)
                sc(2, col, v_lbl, bold=True, bg="2E75B6", color="FFFFFF")
                col += 4
            ws.row_dimensions[2].height = 18

            # Sütun başlıkları (3. satır)
            HDR = "D6E4F0"
            sc(3, 1, "Tarih",  bold=True, bg=HDR)
            sc(3, 2, "Gün",    bold=True, bg=HDR)
            col = 3
            for _ in v_adlar:
                for i in range(1, 5):
                    sc(3, col, str(i), bold=True, bg=HDR)
                    col += 1
            ws.row_dimensions[3].height = 16

            # Veri satırları
            _GUN_K = {"Cumartesi", "Pazar"}
            for ri, satir in enumerate(self._tablo, start=4):
                gn  = satir.get("Gun","")
                bg_g = "FFF3CD" if gn in _GUN_K else None
                sc(ri, 1, satir.get("Tarih",""), bg=bg_g)
                sc(ri, 2, gn, bold=True, bg=bg_g)
                col = 3
                for prefix in v_adlar:
                    for i in range(1, 5):
                        sc(ri, col, satir.get(f"{prefix}{i}",""),
                           bg=bg_g, center=False)
                        col += 1
                ws.row_dimensions[ri].height = 18

            # Özet sayfası
            ws2 = wb.create_sheet("Personel Özeti")
            ozet_hdrs = ["Ad Soyad","Toplam Nöbet","Hedef Saat",
                         "Çalışılan Saat","Fazla Mesai"]
            for ci, h in enumerate(ozet_hdrs, 1):
                c = ws2.cell(row=1, column=ci, value=h)
                c.font = Font(name="Calibri", size=10, bold=True)
                c.fill = PatternFill("solid", fgColor="D6E4F0")
                c.border = bd
                c.alignment = Alignment(horizontal="center",
                                        vertical="center")
                ws2.column_dimensions[chr(64+ci)].width = 20
            for ri, o in enumerate(self._ozet, start=2):
                vals = [o.get("AdSoyad",""),
                        o.get("ToplamSayisi",0),
                        o.get("HedefSaat",""),
                        o.get("CalisanSaat",""),
                        o.get("FazlaMesai","")]
                for ci, v in enumerate(vals, 1):
                    c = ws2.cell(row=ri, column=ci, value=v)
                    c.font   = Font(name="Calibri", size=10)
                    c.border = bd
                    c.alignment = Alignment(
                        horizontal="left" if ci == 1 else "center",
                        vertical="center")

            wb.save(self._kayit)
            return self._kayit
        except Exception as e:
            logger.error(f"Excel oluşturma: {e}")
            return ""
        except Exception as e:
            self.log.emit(f"Hata: {e}")
        finally:
            self.bitti.emit()


# ─── SAYFA ───────────────────────────────────────────────

class NobetRaporPage(QWidget):

    def __init__(self, db=None, action_guard=None, parent=None):
        super().__init__(parent)
        self.setProperty("bg-role", "page")
        self._db     = db
        self._ag     = action_guard
        self._worker: Optional[_RaporWorker] = None
        self._yil    = datetime.date.today().year
        self._ay     = datetime.date.today().month
        self._svc    = get_nobet_service(db) if db else None
        self._rapor_planlar: list[dict] = []
        self._rapor_satirlari: list[dict] = []
        self._rapor_ozet: list[dict] = []
        self._rapor_izinli = True
        self._setup_ui()
        self.spn_yil.valueChanged.connect(self._istatistik_goster)
        self.cmb_ay.currentIndexChanged.connect(self._istatistik_goster)
        self.cmb_birim.currentIndexChanged.connect(self._istatistik_goster)
        if db:
            self.load_data()

    def _setup_ui(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(0)
        root.addWidget(self._build_kontrol_paneli())
        root.addWidget(self._build_ozet_alan(), 1)
        root.addWidget(self._build_footer())

    def _build_kontrol_paneli(self) -> QFrame:
        frame = QFrame()
        frame.setProperty("bg-role", "panel")
        frame.setFixedHeight(120)
        lay = QHBoxLayout(frame)
        lay.setContentsMargins(16, 12, 16, 12)
        lay.setSpacing(12)

        # Dönem
        grp_donem = QGroupBox("Dönem")
        grp_donem.setProperty("style-role", "group")
        dl = QGridLayout(grp_donem)
        dl.setContentsMargins(8, 4, 8, 8)
        dl.setSpacing(6)

        def _lbl(t):
            l = QLabel(t); l.setProperty("color-role","muted")
            l.setProperty("style-role","stat-label"); return l

        dl.addWidget(_lbl("Yıl"), 0, 0)
        from PySide6.QtWidgets import QSpinBox
        self.spn_yil = QSpinBox()
        self.spn_yil.setRange(2020, 2099)
        self.spn_yil.setValue(self._yil)
        dl.addWidget(self.spn_yil, 1, 0)

        dl.addWidget(_lbl("Ay"), 0, 1)
        self.cmb_ay = QComboBox()
        for i, a in enumerate(_AY_ADLARI[1:], 1):
            self.cmb_ay.addItem(a, userData=i)
        self.cmb_ay.setCurrentIndex(self._ay - 1)
        dl.addWidget(self.cmb_ay, 1, 1)
        lay.addWidget(grp_donem, 1)

        # Birim + rapor türü
        grp_filtre = QGroupBox("Filtreler")
        grp_filtre.setProperty("style-role", "group")
        fl = QGridLayout(grp_filtre)
        fl.setContentsMargins(8, 4, 8, 8)
        fl.setSpacing(6)

        fl.addWidget(_lbl("Birim"), 0, 0)
        self.cmb_birim = QComboBox()
        self.cmb_birim.addItem("Tüm Birimler", userData=None)
        fl.addWidget(self.cmb_birim, 1, 0)

        fl.addWidget(_lbl("Rapor Türü"), 0, 1)
        self._bg = QButtonGroup(self)
        self.rb_pdf   = QRadioButton("PDF")
        self.rb_excel = QRadioButton("Excel")
        self.rb_pdf.setChecked(True)
        self.rb_pdf.setProperty("style-role", "radio")
        self.rb_excel.setProperty("style-role", "radio")
        for rb in (self.rb_pdf, self.rb_excel):
            self._bg.addButton(rb)
        rb_row = QHBoxLayout()
        rb_row.addWidget(self.rb_pdf); rb_row.addWidget(self.rb_excel)
        fl.addLayout(rb_row, 1, 1)
        lay.addWidget(grp_filtre, 2)

        # İşlemler
        grp_isl = QGroupBox("İşlemler")
        grp_isl.setProperty("style-role", "group")
        il = QVBoxLayout(grp_isl)
        il.setContentsMargins(8, 4, 8, 8)
        il.setSpacing(6)

        self.btn_yenile = QPushButton(" Yenile")
        self.btn_yenile.setProperty("style-role", "refresh")
        self.btn_yenile.setFixedHeight(28)
        self.btn_yenile.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        IconRenderer.set_button_icon(self.btn_yenile, "refresh",
                                     color=IconColors.MUTED, size=14)
        self.btn_yenile.clicked.connect(self.load_data)
        il.addWidget(self.btn_yenile)

        self.btn_rapor = QPushButton(" Rapor Oluştur")
        self.btn_rapor.setProperty("style-role", "action")
        self.btn_rapor.setFixedHeight(28)
        self.btn_rapor.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        IconRenderer.set_button_icon(self.btn_rapor, "file_text",
                                     color=IconColors.PRIMARY, size=14)
        self.btn_rapor.clicked.connect(self._rapor_olustur)
        if self._ag:
            self._ag.disable_if_unauthorized(self.btn_rapor, "nobet.write")
        self._rapor_izinli = self.btn_rapor.isEnabled()
        il.addWidget(self.btn_rapor)
        lay.addWidget(grp_isl, 1)
        return frame

    def _build_ozet_alan(self) -> QFrame:
        frame = QFrame()
        frame.setProperty("bg-role", "page")
        lay = QVBoxLayout(frame)
        lay.setContentsMargins(16, 16, 16, 16)
        lay.setSpacing(8)

        lbl_ozet = QLabel("Onaylı Plan Önizlemesi")
        lbl_ozet.setProperty("style-role", "section-title")
        lbl_ozet.setProperty("color-role", "primary")
        lay.addWidget(lbl_ozet)

        self.lbl_istatistik = QLabel("Onaylanmış veya yürürlükteki nöbet planları yükleniyor…")
        self.lbl_istatistik.setProperty("color-role", "muted")
        self.lbl_istatistik.setWordWrap(True)
        lay.addWidget(self.lbl_istatistik)

        kart_row = QHBoxLayout()
        kart_row.setSpacing(10)
        self.k_plan = self._istat_karti("Onaylı Plan", "—", "stat-highlight")
        self.k_nobet = self._istat_karti("Toplam Nöbet", "—", "stat-value")
        self.k_personel = self._istat_karti("Personel", "—", "stat-green")
        self.k_fm = self._istat_karti("Fazla Mesai", "—", "stat-red")
        self.k_onay = self._istat_karti("Son Onay", "—", "stat-value")
        for kart in (self.k_plan, self.k_nobet, self.k_personel, self.k_fm, self.k_onay):
            kart_row.addWidget(kart)
        lay.addLayout(kart_row)

        self.lbl_onay_bilgi = QLabel("")
        self.lbl_onay_bilgi.setProperty("color-role", "secondary")
        self.lbl_onay_bilgi.setWordWrap(True)
        lay.addWidget(self.lbl_onay_bilgi)

        self.table_preview = QTableView()
        self.table_preview.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.table_preview.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.table_preview.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)
        self.table_preview.verticalHeader().setVisible(False)
        self.table_preview.setAlternatingRowColors(True)
        self._preview_model = _OnayliNobetModel([])
        self.table_preview.setModel(self._preview_model)
        self._preview_model.setup_columns(self.table_preview, stretch_keys=["AdSoyad"])
        lay.addWidget(self.table_preview, 1)
        return frame

    def _istat_karti(self, baslik: str, deger: str, value_role: str) -> QFrame:
        frame = QFrame()
        frame.setProperty("bg-role", "panel")
        lay = QVBoxLayout(frame)
        lay.setContentsMargins(14, 10, 14, 10)
        lay.setSpacing(3)
        lbl_baslik = QLabel(baslik)
        lbl_baslik.setProperty("style-role", "stat-label")
        lbl_deger = QLabel(deger)
        lbl_deger.setProperty("style-role", value_role)
        lay.addWidget(lbl_baslik)
        lay.addWidget(lbl_deger)
        frame._value_label = lbl_deger
        return frame

    def _istat_guncelle(self, kart: QFrame, deger: str):
        if hasattr(kart, "_value_label"):
            kart._value_label.setText(deger)

    def _build_footer(self) -> QFrame:
        frame = QFrame()
        frame.setFixedHeight(36)
        frame.setProperty("bg-role", "panel")
        lay = QHBoxLayout(frame)
        lay.setContentsMargins(16, 0, 16, 0)
        self.lbl_durum = QLabel("")
        self.lbl_durum.setProperty("color-role", "muted")
        lay.addWidget(self.lbl_durum)
        lay.addStretch()
        return frame

    # ─── Veri ────────────────────────────────────────────

    def load_data(self):
        if not self._svc:
            return
        try:
            svc   = self._svc
            sonuc = svc.get_birimler()
            cur   = self.cmb_birim.currentData()
            self.cmb_birim.blockSignals(True)
            self.cmb_birim.clear()
            self.cmb_birim.addItem("Tüm Birimler", userData=None)
            for b in (sonuc.veri or []):
                if isinstance(b, dict):
                    adi = b.get("BirimAdi", "")
                else:
                    adi = str(b)
                self.cmb_birim.addItem(adi, userData=adi)
            for i in range(self.cmb_birim.count()):
                if self.cmb_birim.itemData(i) == cur:
                    self.cmb_birim.setCurrentIndex(i)
                    break
            self.cmb_birim.blockSignals(False)
            self._istatistik_goster()
        except Exception as e:
            logger.error(f"Rapor sayfa yükleme: {e}")

    def _istatistik_goster(self):
        try:
            yil  = self.spn_yil.value()
            ay   = self.cmb_ay.currentData()
            brid = self.cmb_birim.currentData()
            svc  = self._svc
            sonuc = svc.get_onayli_rapor_verisi(yil, ay, brid)
            if not sonuc.basarili:
                self._bos_durum(sonuc.mesaj or "Onaylı plan bilgisi alınamadı.")
                return
            veri = sonuc.veri or {}
            self._rapor_planlar = veri.get("planlar", []) or []
            self._rapor_satirlari = veri.get("satirlar", []) or []
            self._rapor_ozet = veri.get("ozet", []) or []

            if not self._rapor_planlar:
                self._bos_durum("Bu dönem için onaylanmış veya yürürlükte plan bulunamadı.")
                return

            self._preview_model.set_data(self._rapor_satirlari)

            toplam = len(self._rapor_satirlari)
            plan_adet = len(self._rapor_planlar)
            fm = sum(1 for r in self._rapor_satirlari if r.get("NobetTuru") == "fazla_mesai")
            personel = len({str(r.get("PersonelID", "")) for r in self._rapor_satirlari if str(r.get("PersonelID", ""))})
            son_onay = sorted([
                str(p.get("OnayTarihi", "") or "") for p in self._rapor_planlar if str(p.get("OnayTarihi", "") or "")
            ])[-1] if any(str(p.get("OnayTarihi", "") or "") for p in self._rapor_planlar) else "—"

            self._istat_guncelle(self.k_plan, str(plan_adet))
            self._istat_guncelle(self.k_nobet, str(toplam))
            self._istat_guncelle(self.k_personel, str(personel))
            self._istat_guncelle(self.k_fm, str(fm))
            self._istat_guncelle(self.k_onay, self._fmt_dt(son_onay))

            tek_birim = brid is not None or plan_adet == 1
            self.btn_rapor.setEnabled(self._rapor_izinli and tek_birim and bool(self._rapor_satirlari))
            if tek_birim:
                self.lbl_onay_bilgi.setText(
                    "Onaylı plan önizlemesi hazır. PDF veya Excel çıktısı üretebilirsiniz."
                )
            else:
                self.lbl_onay_bilgi.setText(
                    "Birden fazla onaylı birim bulundu. Resmi çıktı almak için tek birim seçin."
                )

            self.lbl_istatistik.setText(
                f"{_AY_ADLARI[ay]} {yil} dönemi için {plan_adet} onaylı plan bulundu. "
                f"Önizleme yalnız onaylanmış/yürürlükte nöbet satırlarını gösterir."
            )
            self.lbl_durum.setText(f"{toplam} onaylı nöbet satırı listelendi.")
        except Exception as e:
            logger.error(f"İstatistik: {e}")

    def _bos_durum(self, mesaj: str):
        self._rapor_planlar = []
        self._rapor_satirlari = []
        self._rapor_ozet = []
        self._preview_model.set_data([])
        self._istat_guncelle(self.k_plan, "0")
        self._istat_guncelle(self.k_nobet, "0")
        self._istat_guncelle(self.k_personel, "0")
        self._istat_guncelle(self.k_fm, "0")
        self._istat_guncelle(self.k_onay, "—")
        self.lbl_istatistik.setText(mesaj)
        self.lbl_onay_bilgi.setText("Çıktı üretmek için önce onaylı bir nöbet planı bulunmalı.")
        self.lbl_durum.setText(mesaj)
        self.btn_rapor.setEnabled(False)

    @staticmethod
    def _fmt_dt(value: str) -> str:
        try:
            if not value:
                return "—"
            return datetime.datetime.fromisoformat(value).strftime("%d.%m.%Y %H:%M")
        except Exception:
            return value or "—"

    # ─── Rapor ───────────────────────────────────────────

    def _rapor_olustur(self):
        import collections, datetime as _dt
        yil  = self.spn_yil.value()
        ay   = self.cmb_ay.currentData()
        brid = self.cmb_birim.currentData()
        try:
            if not self._rapor_satirlari:
                uyari_goster(self, "Bu dönem için onaylı nöbet kaydı bulunamadı.")
                return
            if brid is None and len(self._rapor_planlar) > 1:
                uyari_goster(self, "Resmi çıktı almak için tek birim seçin.")
                return

            plan = list(self._rapor_satirlari)

            # Gün → vardiya prefix → kişiler listesi
            gun_plan: dict[str, dict] = collections.defaultdict(
                lambda: collections.defaultdict(list)
            )
            # Vardiya adı → prefix harfi (sıralı)
            v_adi_idx: dict[str, str] = {}
            prefix_idx = 0
            for r in sorted(plan, key=lambda x: x.get("NobetTarihi","")):
                v_adi = str(r.get("VardiyaAdi", ""))
                if v_adi and v_adi not in v_adi_idx:
                    v_adi_idx[v_adi] = chr(ord("A") + prefix_idx)
                    prefix_idx += 1
                prefix = v_adi_idx.get(v_adi, "Z")
                ad     = str(r.get("AdSoyad", ""))
                gun_plan[r.get("NobetTarihi","")][prefix].append(ad)

            _GUN = ["Pazartesi","Salı","Çarşamba","Perşembe",
                    "Cuma","Cumartesi","Pazar"]
            tablo = []
            for t in sorted(gun_plan.keys()):
                try:
                    d  = _dt.date.fromisoformat(t)
                    gn = _GUN[d.weekday()]
                    ts = d.strftime("%d.%m.%Y")
                except Exception:
                    gn = ""; ts = t
                satir = {"Tarih": ts, "Gun": gn}
                for prefix, kisiler in gun_plan[t].items():
                    for ki, kisi in enumerate(kisiler[:4], 1):
                        satir[f"{prefix}{ki}"] = kisi
                tablo.append(satir)

            ozet_tablo = list(self._rapor_ozet)

            # Vardiya saat bilgilerini context'e al (ilk 3 vardiya)
            v_list = sorted(v_adi_idx.keys(), key=lambda n: v_adi_idx[n])
            v_rows = []
            for v_adi in v_list:
                ilk = next((r for r in plan if str(r.get("VardiyaAdi", "")) == v_adi), None)
                if not ilk:
                    continue
                v_rows.append({
                    "VardiyaAdi": v_adi,
                    "BasSaat": str(ilk.get("BasSaat", "")),
                    "BitSaat": str(ilk.get("BitSaat", "")),
                })

            def _saat(idx, alan):
                if idx < len(v_list):
                    vn = v_list[idx]
                    vr = next((v for v in v_rows
                               if v.get("VardiyaAdi") == vn), {})
                    return vr.get(alan, "")
                return ""

        except Exception as e:
            hata_goster(self, str(e)); return

        from core.rapor_servisi import RaporServisi
        if brid is None and len(self._rapor_planlar) == 1:
            birim_adi = str(self._rapor_planlar[0].get("BirimAdi", "Tüm Birimler"))
        else:
            birim_adi = self.cmb_birim.currentText()
        mod  = 1 if self.rb_pdf.isChecked() else 2
        isim = f"Nobet_{yil}_{ay:02d}"
        tur  = "pdf" if mod == 1 else "excel"

        kayit = RaporServisi.kaydet_diyalogu(self, isim, tur=tur)
        if not kayit:
            return

        context = {
            "BirimAdi":  birim_adi,
            "Ay":        _AY_ADLARI[ay],
            "Yil":       str(yil),
            "GunduzBas": _saat(0, "BasSaat"),
            "GunduzBit": _saat(0, "BitSaat"),
            "GeceBas":   _saat(1, "BasSaat"),
            "GeceBit":   _saat(1, "BitSaat"),
            "SeyyarAdi": v_list[2] if len(v_list) > 2 else "Seyyar Grafi",
        }

        self.btn_rapor.setEnabled(False)
        self.lbl_durum.setText("Oluşturuluyor…")
        QApplication.setOverrideCursor(Qt.CursorShape.WaitCursor)

        self._worker = _RaporWorker(
            mod=mod, tablo=tablo, ozet=ozet_tablo,
            context=context, donem=f"{_AY_ADLARI[ay]} {yil}",
            kayit=kayit,
        )
        self._worker.log.connect(self.lbl_durum.setText)
        self._worker.bitti.connect(self._rapor_tamam)
        self._worker.start()

    def _rapor_tamam(self):
        QApplication.restoreOverrideCursor()
        tek_birim = self.cmb_birim.currentData() is not None or len(self._rapor_planlar) == 1
        self.btn_rapor.setEnabled(self._rapor_izinli and tek_birim and bool(self._rapor_satirlari))

    def closeEvent(self, event):
        if self._worker and self._worker.isRunning():
            self._worker.quit(); self._worker.wait(500)
        event.accept()
